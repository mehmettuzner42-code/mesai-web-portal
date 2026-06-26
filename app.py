import csv
import glob
import io
import json
import os
import re
import secrets
import smtplib
import threading
import time
import urllib.error
import zipfile
import urllib.request
from datetime import date, datetime, timedelta
from email.message import EmailMessage
from functools import wraps
from types import SimpleNamespace
from collections import Counter, defaultdict

from flask import Flask, flash, g, has_request_context, jsonify, redirect, render_template, request, send_file, session, url_for
from markupsafe import escape
from flask_sqlalchemy import SQLAlchemy
from openpyxl.cell.cell import MergedCell
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.worksheet.page import PageMargins
from sqlalchemy.exc import OperationalError, ProgrammingError
from werkzeug.security import check_password_hash, generate_password_hash
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired


def _load_local_env_file():
    env_path = os.path.join(os.path.dirname(__file__), ".env.local")
    if not os.path.exists(env_path):
        return
    try:
        with open(env_path, "r", encoding="utf-8") as f:
            for raw in f:
                line = raw.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                key = key.strip()
                value = value.strip().strip('"').strip("'")
                if key and key not in os.environ:
                    os.environ[key] = value
    except Exception:
        # .env okunamasa bile uygulama env vars ile devam eder
        pass


_load_local_env_file()

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-change-me")
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get("DATABASE_URL", "sqlite:///mesai_web.db")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["RESET_TOKEN_EXPIRE_MIN"] = int(os.environ.get("RESET_TOKEN_EXPIRE_MIN", "30"))
app.config["APK_URL"] = os.environ.get("APK_URL", "/download-apk")
app.config["UPDATE_MANIFEST_URL"] = os.environ.get(
    "UPDATE_MANIFEST_URL",
    "https://github.com/mehmettuzner42-code/mesai-app/releases/latest/download/update.json",
)
app.config["SMTP_HOST"] = os.environ.get("SMTP_HOST", "")
app.config["SMTP_PORT"] = int(os.environ.get("SMTP_PORT", "587"))
app.config["SMTP_USERNAME"] = os.environ.get("SMTP_USERNAME", "")
app.config["SMTP_PASSWORD"] = os.environ.get("SMTP_PASSWORD", "")
app.config["SMTP_FROM"] = os.environ.get("SMTP_FROM", "")
app.config["SMTP_USE_TLS"] = os.environ.get("SMTP_USE_TLS", "true").lower() == "true"
app.config["SITE_BASE_URL"] = os.environ.get("SITE_BASE_URL", "http://127.0.0.1:5000")
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("COOKIE_SECURE", "false").lower() == "true"
app.config["MAX_CONTENT_LENGTH"] = int(os.environ.get("MAX_UPLOAD_MB", "5")) * 1024 * 1024

db = SQLAlchemy(app)
token_serializer = URLSafeTimedSerializer(app.config["SECRET_KEY"])
_RATE_LIMIT_STATE = {}
FOUNDER_EMAIL = "mehmettuzner42@gmail.com"
_HOLIDAY_BG_REFRESH_RUNNING = set()
_DELEGATE_PERM_CACHE = {}
_DELEGATE_PERM_CACHE_TTL_SEC = 3600
_FOUNDER_ID_CACHE = {"value": None, "expires_at": 0.0}
_DELEGATE_VIEW_CACHE = {}
# Yetkili kullanici agir HTML sayfalarinda tekrar tekrar tam render maliyetini azaltir (cok islemci = kisa TTL).
_DELEGATE_VIEW_CACHE_TTL_SEC = 180.0
_AUDIT_TABLE_READY_CACHE = {"value": None, "checked_at": 0.0}

_RQ_USER_CACHE_UNSET = object()
DAIRE_OPTIONS = ["Abone İşleri Dairesi Başkanlığı"]
SUBE_OPTIONS = [
    "Sayaç İşleri Şube Müdürlüğü",
    "Abone İşleri Şube Müdürlüğü",
    "Müşteri Hizmetleri Şube Müdürlüğü",
    "Tahakkuk İşleri Şube Müdürlüğü",
]


@app.get("/healthz")
def healthz():
    # Keep-alive pingi icin hafif endpoint.
    return jsonify({"ok": True, "service": "mesai-web-portal"}), 200


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False, index=True)
    email = db.Column(db.String(120), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(255), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)


class UserProfile(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, unique=True, index=True)
    daire_baskanligi = db.Column(db.String(255), default="", nullable=False)
    sube_mudurlugu = db.Column(db.String(255), default="", nullable=False)
    ad_soyad = db.Column(db.String(255), default="", nullable=False)
    sicil_no = db.Column(db.String(100), default="", nullable=False)
    ekip_kodu = db.Column(db.String(100), default="", nullable=False)
    employment_end_date = db.Column(db.Date, nullable=True, index=True)


class OvertimeEntry(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    work_date = db.Column(db.Date, nullable=False, index=True)
    start_time = db.Column(db.String(5), nullable=False)
    end_time = db.Column(db.String(5), nullable=False)
    pct60 = db.Column(db.Float, default=0.0, nullable=False)
    pct15 = db.Column(db.Float, default=0.0, nullable=False)
    pazar = db.Column(db.Float, default=0.0, nullable=False)
    bayram = db.Column(db.Float, default=0.0, nullable=False)
    description = db.Column(db.String(500), default="", nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)


class AuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    event_time = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    actor_user_id = db.Column(db.Integer, nullable=True, index=True)
    actor_label = db.Column(db.String(255), default="", nullable=False)
    target_user_id = db.Column(db.Integer, nullable=True, index=True)
    target_label = db.Column(db.String(255), default="", nullable=False)
    action = db.Column(db.String(20), default="", nullable=False, index=True)  # create/update/delete
    source = db.Column(db.String(40), default="", nullable=False, index=True)  # web/apk/bulk/import/backup
    work_date = db.Column(db.Date, nullable=True, index=True)
    period_start_year = db.Column(db.Integer, nullable=True, index=True)
    period_start_month = db.Column(db.Integer, nullable=True, index=True)
    daire_baskanligi = db.Column(db.String(255), default="", nullable=False, index=True)
    sube_mudurlugu = db.Column(db.String(255), default="", nullable=False, index=True)
    old_data_json = db.Column(db.Text, default="{}", nullable=False)
    new_data_json = db.Column(db.Text, default="{}", nullable=False)
    note = db.Column(db.String(255), default="", nullable=False)
    ip_address = db.Column(db.String(64), default="", nullable=False)


class AppSetting(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    setting_key = db.Column(db.String(120), unique=True, nullable=False, index=True)
    setting_value = db.Column(db.Text, default="", nullable=False)


class PeriodLock(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    start_year = db.Column(db.Integer, nullable=False, index=True)
    start_month = db.Column(db.Integer, nullable=False, index=True)
    is_locked = db.Column(db.Boolean, default=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)


class DelegatedAdminPermission(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    owner_user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    delegate_user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, unique=True, index=True)
    allowed_user_ids_json = db.Column(db.Text, default="[]", nullable=False)
    # Legacy kolon (eski surumlerle uyumluluk icin tutuluyor)
    can_view_passwords = db.Column(db.Boolean, default=False, nullable=False)
    can_reset_password = db.Column(db.Boolean, default=False, nullable=False)
    can_view_users_screen = db.Column(db.Boolean, default=False, nullable=False)
    can_view_charts = db.Column(db.Boolean, default=False, nullable=False)
    can_view_filters = db.Column(db.Boolean, default=False, nullable=False)
    can_add_user = db.Column(db.Boolean, default=False, nullable=False)
    can_change_email = db.Column(db.Boolean, default=False, nullable=False)
    can_period_lock = db.Column(db.Boolean, default=False, nullable=False)
    can_bulk_entry = db.Column(db.Boolean, default=False, nullable=False)
    can_view_terminated_users = db.Column(db.Boolean, default=False, nullable=False)
    can_unit_change = db.Column(db.Boolean, default=False, nullable=False)
    scope_daire_baskanligi = db.Column(db.String(255), default="", nullable=False)
    scope_sube_mudurlugu = db.Column(db.String(255), default="", nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)


class UnitChange(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    transfer_date = db.Column(db.Date, nullable=False, index=True)
    from_daire_baskanligi = db.Column(db.String(255), default="", nullable=False)
    from_sube_mudurlugu = db.Column(db.String(255), default="", nullable=False)
    to_daire_baskanligi = db.Column(db.String(255), default="", nullable=False)
    to_sube_mudurlugu = db.Column(db.String(255), default="", nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)


def clear_bulk_mesai_template_person_slot(ws, row60: int):
    """Toplu mesai Excel sablonunda bir personel satir ciftinin veri hucrelerini temizler."""
    row15 = row60 + 1
    cols = [3, 4] + list(range(7, 41))
    for row in (row60, row15):
        for c in cols:
            cell = ws.cell(row=row, column=c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def fmt_num(value: float) -> str:
    if value is None:
        return ""
    if abs(value) < 1e-9:
        return ""
    if float(value).is_integer():
        return str(int(value))
    out = f"{value:.2f}".rstrip("0").rstrip(".")
    return out.replace(".", ",")


def parse_float(value: str) -> float:
    t = (value or "").strip()
    if not t:
        return 0.0
    return float(t.replace(",", "."))


def parse_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def format_dmy(d: date) -> str:
    return d.strftime("%d.%m.%Y")


def weekday_tr(d: date) -> str:
    names = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
    return names[d.weekday()]


def overtime_entry_payload(entry: OvertimeEntry):
    if not entry:
        return {}
    wd = entry.work_date.isoformat() if entry.work_date else ""
    return {
        "id": int(entry.id or 0),
        "user_id": int(entry.user_id or 0),
        "work_date": wd,
        "start_time": str(entry.start_time or ""),
        "end_time": str(entry.end_time or ""),
        "pct60": float(entry.pct60 or 0),
        "pct15": float(entry.pct15 or 0),
        "pazar": float(entry.pazar or 0),
        "bayram": float(entry.bayram or 0),
        "description": str(entry.description or ""),
    }


def _user_label_for(uid: int) -> str:
    if not uid:
        return ""
    u = User.query.get(int(uid))
    if not u:
        return f"#{uid}"
    p = UserProfile.query.filter_by(user_id=u.id).first()
    name = (p.ad_soyad if p else "") or ""
    return str(name).strip() if str(name).strip() else str(u.email or f"#{uid}")


def _is_audit_table_ready() -> bool:
    now = time.monotonic()
    cached = _AUDIT_TABLE_READY_CACHE.get("value")
    checked_at = float(_AUDIT_TABLE_READY_CACHE.get("checked_at") or 0.0)
    if cached is not None and (now - checked_at) < 30.0:
        return bool(cached)
    try:
        inspector = db.inspect(db.engine)
        if not inspector.has_table("audit_log"):
            _AUDIT_TABLE_READY_CACHE["value"] = False
            _AUDIT_TABLE_READY_CACHE["checked_at"] = now
            return False
        cols = {c["name"] for c in inspector.get_columns("audit_log")}
        required = {"event_time", "actor_user_id", "target_user_id", "action", "source", "old_data_json", "new_data_json"}
        ready = required.issubset(cols)
        _AUDIT_TABLE_READY_CACHE["value"] = bool(ready)
        _AUDIT_TABLE_READY_CACHE["checked_at"] = now
        return bool(ready)
    except Exception:
        _AUDIT_TABLE_READY_CACHE["value"] = False
        _AUDIT_TABLE_READY_CACHE["checked_at"] = now
        return False


def write_overtime_audit_log(
    *,
    action: str,
    actor_user_id: int,
    target_user_id: int,
    old_entry: OvertimeEntry = None,
    new_entry: OvertimeEntry = None,
    source: str = "web",
    note: str = "",
):
    try:
        if not _is_audit_table_ready():
            return
        chosen = new_entry or old_entry
        wd = chosen.work_date if chosen else None
        ps = period_start_for_date(wd) if wd else None
        profile = UserProfile.query.filter_by(user_id=int(target_user_id or 0)).first() if target_user_id else None
        actor_label = _user_label_for(int(actor_user_id or 0)) if actor_user_id else "Sistem"
        target_label = _user_label_for(int(target_user_id or 0)) if target_user_id else ""
        ip = ""
        if has_request_context():
            ip = str(request.headers.get("X-Forwarded-For", "") or request.remote_addr or "").split(",")[0].strip()
        log = AuditLog(
            actor_user_id=int(actor_user_id or 0) or None,
            actor_label=actor_label,
            target_user_id=int(target_user_id or 0) or None,
            target_label=target_label,
            action=str(action or "").strip().lower(),
            source=str(source or "").strip().lower(),
            work_date=wd,
            period_start_year=(ps.year if ps else None),
            period_start_month=(ps.month if ps else None),
            daire_baskanligi=str((profile.daire_baskanligi if profile else "") or ""),
            sube_mudurlugu=str((profile.sube_mudurlugu if profile else "") or ""),
            old_data_json=json.dumps(overtime_entry_payload(old_entry), ensure_ascii=False),
            new_data_json=json.dumps(overtime_entry_payload(new_entry), ensure_ascii=False),
            note=str(note or ""),
            ip_address=ip,
        )
        db.session.add(log)
    except Exception:
        _AUDIT_TABLE_READY_CACHE["value"] = False
        _AUDIT_TABLE_READY_CACHE["checked_at"] = time.monotonic()
        # Audit failure should not block core transaction.
        pass


def hhmm_to_minutes(hhmm: str):
    try:
        h, m = hhmm.split(":")
        h, m = int(h), int(m)
        if h < 0 or h > 23 or m < 0 or m > 59:
            return None
        return h * 60 + m
    except Exception:
        return None


MESAI_TIME_OVERLAP_MESSAGE = "Aynı gün için yazılmış mesai kaydınız bulunmaktadır."


def mesai_time_ranges_overlap(start_a: str, end_a: str, start_b: str, end_b: str) -> bool:
    """Aynı takvim günü için iki [başlama, bitiş] aralığı çakışıyor mu (bitişik uçlar hariç)."""
    s1 = hhmm_to_minutes((start_a or "").strip())
    e1 = hhmm_to_minutes((end_a or "").strip())
    s2 = hhmm_to_minutes((start_b or "").strip())
    e2 = hhmm_to_minutes((end_b or "").strip())
    if s1 is None or e1 is None or s2 is None or e2 is None:
        return False
    a0, a1 = s1, e1
    if a1 <= a0:
        a1 += 1440
    b0, b1 = s2, e2
    if b1 <= b0:
        b1 += 1440
    return max(a0, b0) < min(a1, b1)


def find_overlapping_overtime_for_user(
    user_id: int,
    work_date,
    start_time: str,
    end_time: str,
    exclude_entry_id=None,
):
    q = OvertimeEntry.query.filter_by(user_id=int(user_id), work_date=work_date)
    if exclude_entry_id is not None:
        q = q.filter(OvertimeEntry.id != int(exclude_entry_id))
    for row in q.all():
        if mesai_time_ranges_overlap(start_time, end_time, row.start_time or "", row.end_time or ""):
            return row
    return None


def calc_total_hours(start_hhmm: str, end_hhmm: str):
    s = hhmm_to_minutes(start_hhmm)
    e = hhmm_to_minutes(end_hhmm)
    if s is None or e is None:
        return None
    if e <= s:
        e += 1440
    return (e - s) / 60.0


def overlap(a0: int, a1: int, b0: int, b1: int) -> int:
    return max(0, min(a1, b1) - max(a0, b0))


def calc_night_20_06(start_hhmm: str, end_hhmm: str):
    s = hhmm_to_minutes(start_hhmm)
    e = hhmm_to_minutes(end_hhmm)
    if s is None or e is None:
        return None
    if e <= s:
        e += 1440
    total = 0
    max_day = (e // 1440) + 1
    for k in range(max_day + 2):
        d0 = k * 1440
        total += overlap(s, e, d0, d0 + 6 * 60)
        total += overlap(s, e, d0 + 20 * 60, d0 + 24 * 60)
    return total / 60.0


def calc_lunch_12_13(start_hhmm: str, end_hhmm: str):
    s = hhmm_to_minutes(start_hhmm)
    e = hhmm_to_minutes(end_hhmm)
    if s is None or e is None:
        return None
    if e <= s:
        e += 1440
    total = 0
    max_day = (e // 1440) + 1
    for k in range(max_day + 2):
        d0 = k * 1440
        total += overlap(s, e, d0 + 12 * 60, d0 + 13 * 60)
    return total / 60.0


def calc_lunch_1230_1330(start_hhmm: str, end_hhmm: str):
    s = hhmm_to_minutes(start_hhmm)
    e = hhmm_to_minutes(end_hhmm)
    if s is None or e is None:
        return None
    if e <= s:
        e += 1440
    total = 0
    max_day = (e // 1440) + 1
    for k in range(max_day + 2):
        d0 = k * 1440
        total += overlap(s, e, d0 + 12 * 60 + 30, d0 + 13 * 60 + 30)
    return total / 60.0


def _minutes_to_hhmm(m: int) -> str:
    m = ((int(m) % 1440) + 1440) % 1440
    return f"{m // 60:02d}:{m % 60:02d}"


def compute_mesai_split(start: str, end: str, wd: int, holiday_kind):
    """Başlama/bitiş ve güne göre %60, %15, pazar, bayram (tatil/pazar/cumartesi/hafta içi kuralları)."""
    start = (start or "").strip()
    end = (end or "").strip()
    hk = (holiday_kind or "").strip().lower() or None
    if hk not in (None, "half", "full"):
        hk = None
    is_half = hk == "half"
    is_full = hk == "full"
    is_holiday = is_half or is_full

    sm = hhmm_to_minutes(start)
    em = hhmm_to_minutes(end)
    if sm is None or em is None:
        return {"pct60": 0.0, "pct15": 0.0, "pazar": 0.0, "bayram": 0.0}

    G = calc_total_hours(start, end) or 0.0
    L1230 = calc_lunch_1230_1330(start, end) or 0.0
    net = max(0.0, G - L1230)
    night = calc_night_20_06(start, end) or 0.0
    crosses = em <= sm

    T8 = 8 * 60
    T13 = 13 * 60
    T16 = 16 * 60
    T17 = 17 * 60
    T20 = 20 * 60

    def pack(p60, p15, pz, br):
        return {
            "pct60": float(p60),
            "pct15": float(p15),
            "pazar": float(pz),
            "bayram": float(br),
        }

    def extra_from_17_to(end_hhmm: str) -> float:
        return max(
            0.0,
            (calc_total_hours("17:00", end_hhmm) or 0.0) - (calc_lunch_1230_1330("17:00", end_hhmm) or 0.0),
        )

    def irregular_full_or_sunday_pct60() -> float:
        if crosses:
            return net
        ecap = min(em, T20)
        if ecap <= sm:
            return 0.0
        end_c = _minutes_to_hhmm(ecap)
        tot = (ecap - sm) / 60.0
        lsub = calc_lunch_1230_1330(start, end_c) or 0.0
        return max(0.0, tot - lsub)

    if is_holiday:
        if crosses:
            return pack(net, night, 0.0, 0.0)
        if is_half:
            if sm != T13:
                return pack(net, night, 0.0, 0.0)
            if em > T17:
                return pack(extra_from_17_to(end), night, 0.0, 0.5)
            return pack(0.0, night, 0.0, 0.5)
        # tam gün resmi tatil
        if em <= T16:
            return pack(net, night, 0.0, 0.0)
        if sm == T8 and em > T17:
            return pack(extra_from_17_to(end), night, 0.0, 1.0)
        if sm == T8 and em == T17:
            return pack(0.0, night, 0.0, 1.0)
        if sm == T8 and T16 < em < T17:
            return pack(net, night, 0.0, 0.0)
        return pack(irregular_full_or_sunday_pct60(), night, 0.0, 0.0)

    if wd == 6:
        if crosses:
            return pack(net, night, 0.0, 0.0)
        if em <= T16:
            return pack(net, night, 0.0, 0.0)
        if sm == T8 and em > T17:
            return pack(extra_from_17_to(end), night, 1.0, 0.0)
        if sm == T8 and em == T17:
            return pack(0.0, night, 1.0, 0.0)
        if sm == T8 and T16 < em < T17:
            return pack(net, night, 0.0, 0.0)
        return pack(irregular_full_or_sunday_pct60(), night, 0.0, 0.0)

    if wd == 5:
        return pack(max(0.0, G - L1230), night, 0.0, 0.0)
    return pack(max(0.0, G - L1230), night, 0.0, 0.0)


def add_hours_hhmm(hhmm: str, hours: float) -> str:
    base = hhmm_to_minutes(hhmm)
    if base is None:
        return hhmm
    mins = int(round(float(hours or 0.0) * 60))
    end_minutes = (base + mins) % (24 * 60)
    hh = end_minutes // 60
    mm = end_minutes % 60
    return f"{hh:02d}:{mm:02d}"


def saturday_net_mesai_hours(start_hhmm: str, end_hhmm: str) -> float:
    """Cumartesi net %60 suresi: brut - 12:30-13:30 ogle (vardiya ile kesisim)."""
    t = calc_total_hours(start_hhmm, end_hhmm) or 0.0
    l = calc_lunch_1230_1330(start_hhmm, end_hhmm) or 0.0
    return max(0.0, t - l)


def end_hhmm_for_saturday_net(start_hhmm: str, net_hours: float) -> str:
    """Baslangic sabit; net mesai saati net_hours olacak sekilde en yakin bitis (HH:MM)."""
    sm = hhmm_to_minutes(start_hhmm) or 480
    target = float(net_hours)
    if target <= 0:
        return str(start_hhmm or "08:00")
    best_end = "18:00"
    best_err = 1e9
    for D in range(1, 36 * 60 + 1):
        em = (sm + D) % (24 * 60)
        end_str = f"{em // 60:02d}:{em % 60:02d}"
        n = saturday_net_mesai_hours(start_hhmm, end_str)
        err = abs(n - target)
        if err < best_err - 1e-9:
            best_err = err
            best_end = end_str
    return best_end


def end_hhmm_for_bulk_special_target_pct60(start_hhmm: str, target_pct60: float, wd: int, holiday_kind) -> str:
    """Pazar / resmi tatilde hücredeki sayı doğrudan %60 saatidir; bitiş, öğle ve tatil kurallarıyla bu nete uyan şekilde aranır."""
    target = float(target_pct60 or 0.0)
    if target <= 0:
        return str(start_hhmm or "08:00")
    sm = hhmm_to_minutes(start_hhmm) or 480
    best_end = add_hours_hhmm(start_hhmm, target)
    best_err = 1e9
    for D in range(1, 36 * 60 + 1):
        em = (sm + D) % (24 * 60)
        end_str = f"{em // 60:02d}:{em % 60:02d}"
        sp = compute_mesai_split(str(start_hhmm), end_str, int(wd), holiday_kind)
        if float(sp.get("pazar", 0) or 0) > 1e-6 or float(sp.get("bayram", 0) or 0) > 1e-6:
            continue
        p60 = float(sp.get("pct60", 0) or 0)
        err = abs(p60 - target)
        if err < best_err - 1e-9:
            best_err = err
            best_end = end_str
    if best_err > 0.02:
        return add_hours_hhmm(start_hhmm, target)
    return best_end


def tr_upper(text: str) -> str:
    # Turkce buyuk harf donusumu: i->I degil, i->I ve ı->I kurallarini dogru uygular.
    if text is None:
        return ""
    trans = str.maketrans({"i": "İ", "ı": "I"})
    return str(text).translate(trans).upper()


def period_start_for_date(d: date) -> date:
    if d.day >= 24:
        return date(d.year, d.month, 24)
    if d.month == 1:
        return date(d.year - 1, 12, 24)
    return date(d.year, d.month - 1, 24)


def add_month(year: int, month: int):
    if month == 12:
        return year + 1, 1
    return year, month + 1


def period_for_start(year: int, month: int):
    start = date(year, month, 24)
    ey, em = add_month(year, month)
    end = date(ey, em, 23)
    return start, end


def period_year(start_year: int, start_month: int) -> int:
    return start_year + 1 if start_month == 12 else start_year


def year_period_workdate_bounds(selected_year: int):
    # Donem yilina gore kapsanan tum takvim gunleri:
    # 24.12.(Y-1) - 23.12.Y
    y = int(selected_year)
    return date(y - 1, 12, 24), date(y, 12, 23)


def build_start_options_from_date_range(min_work_date: date, max_work_date: date):
    if not min_work_date or not max_work_date:
        ps = period_start_for_date(date.today())
        return [(ps.year, ps.month)]
    cur = period_start_for_date(max_work_date)
    min_ps = period_start_for_date(min_work_date)
    out = []
    seen = set()
    while (cur.year, cur.month) >= (min_ps.year, min_ps.month):
        k = (cur.year, cur.month)
        if k not in seen:
            out.append(k)
            seen.add(k)
        if cur.month == 1:
            cur = date(cur.year - 1, 12, 24)
        else:
            cur = date(cur.year, cur.month - 1, 24)
    return out or [(period_start_for_date(date.today()).year, period_start_for_date(date.today()).month)]


def resolve_period_start_year(selected_year: int, start_month: int) -> int:
    # Donem yili kurali: Aralikta baslayan donem bir sonraki yilin donemine yazilir.
    # Ornek: 24.12.2024-23.01.2025 donemi, 2025 donem yilina aittir.
    return selected_year - 1 if int(start_month) == 12 else selected_year


def fixed_holiday_set(year: int):
    return {
        date(year, 1, 1),
        date(year, 4, 23),
        date(year, 5, 1),
        date(year, 5, 19),
        date(year, 7, 15),
        date(year, 8, 30),
        date(year, 10, 29),
    }


def half_holiday_set(year: int):
    # Yarım gün resmi tatiller (arife günleri) - içe aktarma kuralı için.
    # Gerektikçe yeni yıllar eklenebilir.
    mapping = {
        2024: {date(2024, 4, 9), date(2024, 6, 15)},
        2025: {date(2025, 3, 29), date(2025, 6, 5)},
        2026: {date(2026, 3, 19), date(2026, 5, 26)},
        2027: {date(2027, 3, 8), date(2027, 5, 15)},
    }
    return mapping.get(year, set())


def religious_full_holiday_set(year: int):
    # Dini bayram tam gun tarihleri (TR)
    mapping = {
        2024: {
            date(2024, 4, 10), date(2024, 4, 11), date(2024, 4, 12),  # Ramazan
            date(2024, 6, 16), date(2024, 6, 17), date(2024, 6, 18), date(2024, 6, 19),  # Kurban
        },
        2025: {
            date(2025, 3, 30), date(2025, 3, 31), date(2025, 4, 1),  # Ramazan
            date(2025, 6, 6), date(2025, 6, 7), date(2025, 6, 8), date(2025, 6, 9),  # Kurban
        },
        2026: {
            date(2026, 3, 20), date(2026, 3, 21), date(2026, 3, 22),  # Ramazan
            date(2026, 5, 27), date(2026, 5, 28), date(2026, 5, 29), date(2026, 5, 30),  # Kurban
        },
        2027: {
            date(2027, 3, 9), date(2027, 3, 10), date(2027, 3, 11),  # Ramazan
            date(2027, 5, 16), date(2027, 5, 17), date(2027, 5, 18), date(2027, 5, 19),  # Kurban
        },
    }
    return mapping.get(year, set())


def _holiday_cache_key(year: int) -> str:
    return f"holiday_cache_tr_{int(year)}"


def _holiday_cache_fetched_key(year: int) -> str:
    return f"holiday_cache_tr_fetched_{int(year)}"


def fetch_public_holidays_tr_rows(year: int):
    y = int(year)
    url = f"https://date.nager.at/api/v3/PublicHolidays/{y}/TR"
    req = urllib.request.Request(url, headers={"User-Agent": "MesaiWeb/1.0"})
    with urllib.request.urlopen(req, timeout=10) as resp:
        raw = resp.read().decode("utf-8")
    arr = json.loads(raw)
    if not isinstance(arr, list):
        return []

    full = {}
    ramazan_start = None
    kurban_start = None
    for item in arr:
        if not isinstance(item, dict):
            continue
        day = str(item.get("date", "")).strip()
        if len(day) != 10:
            continue
        name = str(item.get("localName") or item.get("name") or "").strip()
        full[day] = name
        low = name.lower()
        if ("ramazan" in low) or ("eid al-fitr" in low) or ("şeker" in low):
            ramazan_start = day if (ramazan_start is None or day < ramazan_start) else ramazan_start
        if ("kurban" in low) or ("eid al-adha" in low):
            kurban_start = day if (kurban_start is None or day < kurban_start) else kurban_start

    rows = [{"day": day, "kind": "full", "name": name} for day, name in full.items()]

    if f"{y:04d}-10-29" in full:
        rows.append({"day": f"{y:04d}-10-28", "kind": "half", "name": "Cumhuriyet Bayramı Arefesi"})

    def minus_one_day(ymd: str):
        try:
            d = datetime.strptime(ymd, "%Y-%m-%d").date()
            return (d - timedelta(days=1)).isoformat()
        except Exception:
            return ymd

    if ramazan_start:
        eve = minus_one_day(ramazan_start)
        if eve.startswith(f"{y:04d}-"):
            rows.append({"day": eve, "kind": "half", "name": "Ramazan Bayramı Arefesi"})
    if kurban_start:
        eve = minus_one_day(kurban_start)
        if eve.startswith(f"{y:04d}-"):
            rows.append({"day": eve, "kind": "half", "name": "Kurban Bayramı Arefesi"})

    return rows


def normalize_holiday_rows(rows):
    by_day = {}
    for r in rows if isinstance(rows, list) else []:
        day = str((r or {}).get("day", "")).strip()
        kind = str((r or {}).get("kind", "")).strip().lower()
        name = str((r or {}).get("name", "")).strip()
        if len(day) != 10 or kind not in ("full", "half"):
            continue
        prev = by_day.get(day)
        # Aynı güne full ve half gelirse full onceliklidir.
        if prev is None or (prev.get("kind") == "half" and kind == "full"):
            by_day[day] = {"day": day, "kind": kind, "name": name}
    return [by_day[k] for k in sorted(by_day.keys())]


def ensure_holiday_cache_tr(year: int, force_refresh: bool = False, allow_network: bool = True):
    y = int(year)
    key = _holiday_cache_key(y)
    fetched_key = _holiday_cache_fetched_key(y)
    now_iso = datetime.utcnow().isoformat()

    current_raw = get_setting_value(key, "")
    try:
        current_rows = normalize_holiday_rows(json.loads(current_raw or "[]"))
    except Exception:
        current_rows = []

    # 24 saatten eskiyse internetten yenilemeyi dener; hata olursa eldeki cache korunur.
    fetched_at = get_setting_value(fetched_key, "")
    should_refresh = True
    if fetched_at:
        try:
            last = datetime.fromisoformat(fetched_at)
            should_refresh = (datetime.utcnow() - last).total_seconds() > 24 * 3600
        except Exception:
            should_refresh = True

    if (not force_refresh) and (not should_refresh) and current_rows:
        return

    if allow_network:
        try:
            rows_remote = normalize_holiday_rows(fetch_public_holidays_tr_rows(y))
            # Sabit gunleri de her zaman dahil et (eksik kaynak verisine karsi).
            full_fallback_days = sorted(fixed_holiday_set(y) | religious_full_holiday_set(y))
            fixed_rows = [{"day": d.isoformat(), "kind": "full", "name": ""} for d in full_fallback_days]
            fixed_rows.extend([{"day": d.isoformat(), "kind": "half", "name": ""} for d in sorted(half_holiday_set(y))])
            merged_rows = normalize_holiday_rows(rows_remote + fixed_rows)
            if merged_rows:
                if merged_rows != current_rows:
                    set_setting_value(key, json.dumps(merged_rows, ensure_ascii=False))
                set_setting_value(fetched_key, now_iso)
                db.session.commit()
                return
        except Exception:
            db.session.rollback()

    # Ağ yokken mevcut cache varsa direkt kullan.
    if current_rows:
        return

    # İlk kurulumda cache boşsa fallback oluştur; internet geldiğinde arka planda güncellenir.
    full_fallback_days = sorted(fixed_holiday_set(y) | religious_full_holiday_set(y))
    fallback = [{"day": d.isoformat(), "kind": "full", "name": ""} for d in full_fallback_days]
    fallback.extend([{"day": d.isoformat(), "kind": "half", "name": ""} for d in sorted(half_holiday_set(y))])
    fallback = normalize_holiday_rows(fallback)
    try:
        set_setting_value(key, json.dumps(fallback, ensure_ascii=False))
        db.session.commit()
    except Exception:
        db.session.rollback()


def _refresh_holiday_cache_tr_bg(year: int):
    y = int(year)
    try:
        with app.app_context():
            ensure_holiday_cache_tr(y, allow_network=True)
    finally:
        _HOLIDAY_BG_REFRESH_RUNNING.discard(y)


def schedule_holiday_cache_refresh(year: int):
    y = int(year)
    if y in _HOLIDAY_BG_REFRESH_RUNNING:
        return
    _HOLIDAY_BG_REFRESH_RUNNING.add(y)
    t = threading.Thread(target=_refresh_holiday_cache_tr_bg, args=(y,), daemon=True)
    t.start()


def holiday_kind_tr(target_date: date):
    # Hesaplama sırasında internet bekletmesi olmasın: sadece kayıtlı cache/fallback kullan.
    ensure_holiday_cache_tr(target_date.year, allow_network=False)
    # Güncelleme ihtiyacı varsa arka planda dene.
    schedule_holiday_cache_refresh(target_date.year)
    key = _holiday_cache_key(target_date.year)
    day_iso = target_date.isoformat()

    def find_kind():
        try:
            rows = json.loads(get_setting_value(key, "[]") or "[]")
        except Exception:
            rows = []
        for r in rows if isinstance(rows, list) else []:
            if str((r or {}).get("day", "")).strip() == day_iso:
                kind = str((r or {}).get("kind", "")).strip().lower()
                if kind in ("full", "half"):
                    return kind
        return None

    found = find_kind()
    if found:
        return found

    return None


def day_defaults(target_date: date, end_time_override: str = None, start_time_override: str = None):
    wd = target_date.weekday()  # 0 pazartesi ... 6 pazar
    kind = holiday_kind_tr(target_date)
    is_half_holiday = kind == "half"
    is_full_holiday = kind == "full"
    is_holiday = is_full_holiday or is_half_holiday

    if is_holiday:
        def_start, def_end = ("13:00", "17:00") if is_half_holiday else ("08:00", "17:00")
    elif wd == 6:
        def_start, def_end = "08:00", "17:00"
    elif wd == 5:
        def_start, def_end = "08:00", "18:00"
    else:
        def_start, def_end = "18:00", "21:00"

    start = (start_time_override or "").strip() or def_start
    end = (end_time_override or "").strip() or def_end

    split = compute_mesai_split(start, end, wd, kind)
    return {
        "start": start,
        "end": end,
        "pct60": split["pct60"],
        "pct15": split["pct15"],
        "pazar": split["pazar"],
        "bayram": split["bayram"],
        "isHoliday": is_holiday,
        "isHalfHoliday": is_half_holiday,
        "weekday": wd,
    }


def send_reset_email(to_email: str, reset_url: str) -> bool:
    host = app.config["SMTP_HOST"]
    username = app.config["SMTP_USERNAME"]
    password = app.config["SMTP_PASSWORD"]
    sender = app.config["SMTP_FROM"] or username
    if not host or not sender:
        return False
    msg = EmailMessage()
    msg["Subject"] = "Mesai Portal - Sifre Sifirlama"
    msg["From"] = sender
    msg["To"] = to_email
    msg.set_content(
        "Sifre sifirlama baglantiniz:\n\n"
        f"{reset_url}\n\n"
        f"Baglanti {app.config['RESET_TOKEN_EXPIRE_MIN']} dakika gecerlidir."
    )
    port = app.config["SMTP_PORT"]
    use_tls = app.config["SMTP_USE_TLS"]
    if port == 465:
        with smtplib.SMTP_SSL(host, port, timeout=20) as server:
            if username:
                server.login(username, password)
            server.send_message(msg)
    else:
        with smtplib.SMTP(host, port, timeout=20) as server:
            if use_tls:
                server.starttls()
            if username:
                server.login(username, password)
            server.send_message(msg)
    return True


def is_rate_limited(key: str, limit: int, window_sec: int) -> bool:
    now = datetime.utcnow().timestamp()
    rec = _RATE_LIMIT_STATE.get(key, [])
    rec = [t for t in rec if now - t < window_sec]
    if len(rec) >= limit:
        _RATE_LIMIT_STATE[key] = rec
        return True
    rec.append(now)
    _RATE_LIMIT_STATE[key] = rec
    return False


def login_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        try:
            uid = int(session.get("user_id") or 0)
        except Exception:
            uid = 0
        if uid and is_user_terminated(uid):
            session.clear()
            flash("İşten ayrılış kaydınız bulunduğu için sisteme erişiminiz kapatılmıştır.", "error")
            return redirect(url_for("login"))
        return view_func(*args, **kwargs)

    return wrapped


def is_founder_user(user: User) -> bool:
    return bool(user and (user.email or "").strip().lower() == FOUNDER_EMAIL)


def founder_user_id() -> int:
    now = time.monotonic()
    if _FOUNDER_ID_CACHE["value"] is not None and now < float(_FOUNDER_ID_CACHE["expires_at"]):
        return int(_FOUNDER_ID_CACHE["value"] or 0)
    u = User.query.filter(db.func.lower(db.func.trim(User.email)) == FOUNDER_EMAIL.lower()).first()
    val = int(u.id) if u else 0
    _FOUNDER_ID_CACHE["value"] = val
    _FOUNDER_ID_CACHE["expires_at"] = now + 300.0
    return val


def invalidate_delegate_permission_cache(delegate_user_id: int = None):
    if delegate_user_id is None:
        _DELEGATE_PERM_CACHE.clear()
        _DELEGATE_VIEW_CACHE.clear()
        _FOUNDER_ID_CACHE["value"] = None
        _FOUNDER_ID_CACHE["expires_at"] = 0.0
        return
    _DELEGATE_PERM_CACHE.pop(int(delegate_user_id), None)
    uid = int(delegate_user_id)
    for k in list(_DELEGATE_VIEW_CACHE.keys()):
        if f"|u:{uid}|" in k:
            _DELEGATE_VIEW_CACHE.pop(k, None)


def _delegate_view_cache_key(route_name: str, login_user_id: int):
    q = request.query_string.decode("utf-8", errors="ignore")
    imp = str(session.get("admin_impersonate_user_id") or "")
    return f"{route_name}|u:{int(login_user_id)}|imp:{imp}|q:{q}"


def get_delegate_view_cache(route_name: str, login_user_id: int):
    key = _delegate_view_cache_key(route_name, login_user_id)
    rec = _DELEGATE_VIEW_CACHE.get(key)
    if not rec:
        return None
    if time.monotonic() >= float(rec.get("expires_at", 0)):
        _DELEGATE_VIEW_CACHE.pop(key, None)
        return None
    return rec.get("html")


def set_delegate_view_cache(route_name: str, login_user_id: int, html: str):
    key = _delegate_view_cache_key(route_name, login_user_id)
    _DELEGATE_VIEW_CACHE[key] = {
        "expires_at": time.monotonic() + _DELEGATE_VIEW_CACHE_TTL_SEC,
        "html": html,
    }


def _perm_to_cache_dict(perm: DelegatedAdminPermission):
    return {
        "owner_user_id": int(perm.owner_user_id),
        "delegate_user_id": int(perm.delegate_user_id),
        "allowed_user_ids_json": str(perm.allowed_user_ids_json or "[]"),
        "can_view_passwords": bool(perm.can_view_passwords),
        "can_reset_password": bool(perm.can_reset_password),
        "can_view_users_screen": bool(perm.can_view_users_screen),
        "can_view_charts": bool(perm.can_view_charts),
        "can_view_filters": bool(perm.can_view_filters),
        "can_add_user": bool(perm.can_add_user),
        "can_change_email": bool(perm.can_change_email),
        "can_period_lock": bool(perm.can_period_lock),
        "can_bulk_entry": bool(perm.can_bulk_entry),
        "can_view_terminated_users": bool(perm.can_view_terminated_users),
        "can_unit_change": bool(perm.can_unit_change),
        "scope_daire_baskanligi": str(perm.scope_daire_baskanligi or ""),
        "scope_sube_mudurlugu": str(perm.scope_sube_mudurlugu or ""),
    }


def get_delegate_permission(user_id: int):
    fid = founder_user_id()
    if not fid:
        return None
    now = time.monotonic()
    cached = _DELEGATE_PERM_CACHE.get(int(user_id))
    if cached and now < float(cached.get("expires_at", 0)):
        data = cached.get("data")
        if data is None:
            return None
        if int(data.get("owner_user_id", 0)) == int(fid):
            return SimpleNamespace(**data)
    try:
        perm = DelegatedAdminPermission.query.filter_by(owner_user_id=fid, delegate_user_id=user_id).first()
    except (OperationalError, ProgrammingError):
        db.session.rollback()
        # Canlıda kolonlar henüz oluşmadıysa otomatik tamamla ve tekrar dene.
        ensure_delegated_permission_columns()
        perm = DelegatedAdminPermission.query.filter_by(owner_user_id=fid, delegate_user_id=user_id).first()
    if not perm:
        _DELEGATE_PERM_CACHE[int(user_id)] = {"expires_at": now + _DELEGATE_PERM_CACHE_TTL_SEC, "data": None}
        return None
    data = _perm_to_cache_dict(perm)
    _DELEGATE_PERM_CACHE[int(user_id)] = {"expires_at": now + _DELEGATE_PERM_CACHE_TTL_SEC, "data": data}
    return SimpleNamespace(**data)


def allowed_user_ids_for(user: User):
    if not user:
        return set()
    if is_founder_user(user):
        return None  # None => all users
    perm = get_delegate_permission(user.id)
    if not perm:
        return set()
    try:
        raw = json.loads(perm.allowed_user_ids_json or "[]")
        return {int(x) for x in raw if str(x).isdigit()}
    except Exception:
        return set()


def can_access_admin_area(user: User) -> bool:
    return bool(is_founder_user(user) or get_delegate_permission(user.id if user else 0))


def delegate_can(user: User, capability: str) -> bool:
    if not user:
        return False
    if is_founder_user(user):
        return True
    perm = get_delegate_permission(user.id)
    if not perm:
        return False
    if capability == "users":
        return bool(perm.can_view_users_screen)
    if capability == "charts":
        return bool(perm.can_view_charts)
    if capability == "filters":
        return bool(perm.can_view_filters)
    if capability == "add_user":
        return bool(perm.can_add_user)
    if capability == "change_email":
        return bool(perm.can_change_email)
    if capability == "period_lock":
        return bool(perm.can_period_lock)
    if capability == "bulk_entry":
        return bool(perm.can_bulk_entry)
    if capability == "terminated_users":
        return bool(perm.can_view_terminated_users)
    if capability == "unit_change":
        return bool(perm.can_unit_change)
    if capability == "reset_password":
        return bool(perm.can_reset_password)
    if capability == "impersonate":
        return bool(perm.can_view_users_screen)
    return False


def session_login_user():
    """Oturumdaki kullaniciyi dondurur; tek istek icinde tekrarlanan User.query.get cagrilarini g ile onler."""
    if has_request_context():
        hit = getattr(g, "_session_login_user_hit", _RQ_USER_CACHE_UNSET)
        if hit is not _RQ_USER_CACHE_UNSET:
            return g._session_login_user
    uid = session.get("user_id")
    if not uid:
        if has_request_context():
            g._session_login_user_hit = True
            g._session_login_user = None
        return None
    u = User.query.get(uid)
    if has_request_context():
        g._session_login_user_hit = True
        g._session_login_user = u
    return u


def is_user_terminated(user_id: int) -> bool:
    p = UserProfile.query.filter_by(user_id=user_id).first()
    if not p or not p.employment_end_date:
        return False
    return date.today() >= p.employment_end_date


def include_user_for_selected_year(profile: UserProfile, selected_year: int) -> bool:
    if not profile or not profile.employment_end_date:
        return True
    return int(selected_year) <= int(profile.employment_end_date.year)


def include_user_for_selected_period(profile: UserProfile, period_start_year: int, period_start_month: int) -> bool:
    if not profile or not profile.employment_end_date:
        return True
    end_ps = period_start_for_date(profile.employment_end_date)
    return (int(period_start_year), int(period_start_month)) <= (int(end_ps.year), int(end_ps.month))


def unit_changes_map_for_users(user_ids):
    ids = [int(x) for x in (user_ids or []) if str(x).isdigit()]
    if not ids:
        return {}
    rows = (
        UnitChange.query.filter(UnitChange.user_id.in_(ids))
        .order_by(UnitChange.user_id.asc(), UnitChange.transfer_date.asc(), UnitChange.id.asc())
        .all()
    )
    out = {}
    for r in rows:
        out.setdefault(int(r.user_id), []).append(r)
    return out


def unit_at_date_for_user(user_id: int, target_date: date, profile: UserProfile = None, changes=None):
    p = profile or UserProfile.query.filter_by(user_id=user_id).first() or UserProfile(user_id=user_id)
    if changes is None:
        changes = UnitChange.query.filter_by(user_id=user_id).order_by(UnitChange.transfer_date.asc(), UnitChange.id.asc()).all()
    if not changes:
        return {
            "daire_baskanligi": (p.daire_baskanligi or "").strip(),
            "sube_mudurlugu": (p.sube_mudurlugu or "").strip(),
        }
    current_daire = (changes[0].from_daire_baskanligi or "").strip()
    current_sube = (changes[0].from_sube_mudurlugu or "").strip()
    for c in changes:
        if target_date < c.transfer_date:
            return {"daire_baskanligi": current_daire, "sube_mudurlugu": current_sube}
        current_daire = (c.to_daire_baskanligi or "").strip()
        current_sube = (c.to_sube_mudurlugu or "").strip()
    return {"daire_baskanligi": current_daire, "sube_mudurlugu": current_sube}


def unit_scope_allows_user(viewer: User, target_user_id: int, ref_date: date, profile: UserProfile = None, perm=None, changes=None) -> bool:
    if not viewer:
        return False
    if is_founder_user(viewer):
        return True
    perm = perm or get_delegate_permission(viewer.id)
    if not perm:
        return False
    daire_scope = (perm.scope_daire_baskanligi or "").strip()
    sube_scope = (perm.scope_sube_mudurlugu or "").strip()
    if not daire_scope and not sube_scope:
        return True
    u = unit_at_date_for_user(target_user_id, ref_date, profile=profile, changes=changes)
    u_daire = (u.get("daire_baskanligi") or "").strip()
    u_sube = (u.get("sube_mudurlugu") or "").strip()
    if daire_scope and u_daire != daire_scope:
        return False
    if sube_scope and u_sube != sube_scope:
        return False
    return True


def unit_scope_allows_user_for_year(viewer: User, target_user_id: int, selected_year: int, profile: UserProfile = None, perm=None, changes=None) -> bool:
    if not viewer:
        return False
    if is_founder_user(viewer):
        return True
    y = int(selected_year)
    d1 = date(y, 1, 1)
    d2 = date(y, 12, 31)
    return unit_scope_allows_user(viewer, target_user_id, d1, profile=profile, perm=perm, changes=changes) or unit_scope_allows_user(
        viewer, target_user_id, d2, profile=profile, perm=perm, changes=changes
    )


def aggregate_entries_with_scope(entries, profiles_map, unit_changes_map, daire_scope: str, sube_scope: str):
    daire_scope = (daire_scope or "").strip()
    sube_scope = (sube_scope or "").strip()
    state = {}
    out = {}
    for e in entries:
        uid = int(e.user_id)
        st = state.get(uid)
        if st is None:
            changes = list(unit_changes_map.get(uid) or [])
            p = profiles_map.get(uid) or UserProfile(user_id=uid)
            if changes:
                cur_d = (changes[0].from_daire_baskanligi or "").strip()
                cur_s = (changes[0].from_sube_mudurlugu or "").strip()
            else:
                cur_d = (p.daire_baskanligi or "").strip()
                cur_s = (p.sube_mudurlugu or "").strip()
            st = {"changes": changes, "idx": 0, "cur_d": cur_d, "cur_s": cur_s}
            state[uid] = st

        changes = st["changes"]
        idx = int(st["idx"])
        while idx < len(changes) and e.work_date >= changes[idx].transfer_date:
            st["cur_d"] = (changes[idx].to_daire_baskanligi or "").strip()
            st["cur_s"] = (changes[idx].to_sube_mudurlugu or "").strip()
            idx += 1
        st["idx"] = idx

        if daire_scope and st["cur_d"] != daire_scope:
            continue
        if sube_scope and st["cur_s"] != sube_scope:
            continue

        d = out.setdefault(uid, {"pct60": 0.0, "pct15": 0.0, "pazar": 0.0, "bayram": 0.0})
        d["pct60"] += float(e.pct60 or 0)
        d["pct15"] += float(e.pct15 or 0)
        d["pazar"] += float(e.pazar or 0)
        d["bayram"] += float(e.bayram or 0)
    return out


def filter_entries_with_scope(entries, user_id: int, profile: UserProfile, changes, daire_scope: str, sube_scope: str):
    daire_scope = (daire_scope or "").strip()
    sube_scope = (sube_scope or "").strip()
    if not daire_scope and not sube_scope:
        return list(entries)
    changes = list(changes or [])
    if changes:
        cur_d = (changes[0].from_daire_baskanligi or "").strip()
        cur_s = (changes[0].from_sube_mudurlugu or "").strip()
    else:
        cur_d = ((profile.daire_baskanligi if profile else "") or "").strip()
        cur_s = ((profile.sube_mudurlugu if profile else "") or "").strip()
    idx = 0
    out = []
    for e in entries:
        while idx < len(changes) and e.work_date >= changes[idx].transfer_date:
            cur_d = (changes[idx].to_daire_baskanligi or "").strip()
            cur_s = (changes[idx].to_sube_mudurlugu or "").strip()
            idx += 1
        if daire_scope and cur_d != daire_scope:
            continue
        if sube_scope and cur_s != sube_scope:
            continue
        out.append(e)
    return out


def period_start_key_for_date(target_date: date):
    ps = period_start_for_date(target_date)
    return ps.year, ps.month


def is_period_locked(target_date: date) -> bool:
    sy, sm = period_start_key_for_date(target_date)
    lock = PeriodLock.query.filter_by(start_year=sy, start_month=sm).first()
    return bool(lock and lock.is_locked)


def can_bypass_period_lock(user: User) -> bool:
    return bool(is_founder_user(user) or delegate_can(user, "period_lock"))


def current_user():
    """login_user + burunme; tek istekte tekrarlanan sorgulari g ile onler."""
    if has_request_context():
        hit = getattr(g, "_current_user_hit", _RQ_USER_CACHE_UNSET)
        if hit is not _RQ_USER_CACHE_UNSET:
            return g._current_user
    login_user = session_login_user()
    if login_user is None:
        if has_request_context():
            g._current_user_hit = True
            g._current_user = None
        return None
    # Kurucu kullanici "kullaniciya burunme" modunda ise ekrandaki tum veriler secilen kisiye gore akar.
    if can_access_admin_area(login_user):
        imp_uid = session.get("admin_impersonate_user_id")
        if imp_uid:
            imp_user = User.query.get(imp_uid)
            if imp_user:
                allowed = allowed_user_ids_for(login_user)
                if allowed is None or imp_user.id in allowed:
                    if has_request_context():
                        g._current_user_hit = True
                        g._current_user = imp_user
                    return imp_user
                session.pop("admin_impersonate_user_id", None)
    if has_request_context():
        g._current_user_hit = True
        g._current_user = login_user
    return login_user


def ensure_user_or_redirect():
    user = current_user()
    if user is None:
        session.clear()
        return None
    return user


def admin_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        login_user = session_login_user()
        if not login_user:
            return redirect(url_for("login"))
        if not is_founder_user(login_user):
            flash("Bu alan sadece kurucu kullanıcıya açıktır.", "error")
            return redirect(url_for("dashboard"))
        return view_func(*args, **kwargs)

    return wrapped


def admin_or_delegate_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        login_user = session_login_user()
        if not login_user:
            return redirect(url_for("login"))
        if not can_access_admin_area(login_user):
            flash("Bu alan sadece yetkili kullanıcılara açıktır.", "error")
            return redirect(url_for("dashboard"))
        return view_func(*args, **kwargs)

    return wrapped


def get_or_create_profile(user_id: int):
    p = UserProfile.query.filter_by(user_id=user_id).first()
    if p:
        return p
    p = UserProfile(user_id=user_id)
    db.session.add(p)
    db.session.commit()
    return p


def get_setting_value(key: str, default_value: str = "") -> str:
    row = AppSetting.query.filter_by(setting_key=key).first()
    return row.setting_value if row and row.setting_value is not None else default_value


def set_setting_value(key: str, value: str):
    row = AppSetting.query.filter_by(setting_key=key).first()
    if row:
        row.setting_value = value
    else:
        row = AppSetting(setting_key=key, setting_value=value)
        db.session.add(row)


def entry_to_dict(entry: OvertimeEntry):
    return {
        "id": entry.id,
        "workDate": entry.work_date.isoformat(),
        "startTime": entry.start_time,
        "endTime": entry.end_time,
        "pct60": entry.pct60,
        "pct15": entry.pct15,
        "pazar": entry.pazar,
        "bayram": entry.bayram,
        "description": entry.description,
        "updatedAt": entry.updated_at.isoformat(),
    }


def grouped_period_rows(entries):
    day_map = {}
    for e in entries:
        key = e.work_date.isoformat()
        if key not in day_map:
            day_map[key] = {
                "work_date": e.work_date,
                "start_time": e.start_time,
                "end_time": e.end_time,
                "pct60": e.pct60,
                "pct15": e.pct15,
                "pazar": e.pazar,
                "bayram": e.bayram,
                "description": e.description.strip(),
                "entry_id": e.id,
            }
        else:
            r = day_map[key]
            r["start_time"] = min(r["start_time"], e.start_time)
            r["end_time"] = max(r["end_time"], e.end_time)
            r["pct60"] += e.pct60
            r["pct15"] += e.pct15
            r["pazar"] += e.pazar
            r["bayram"] += e.bayram
            if e.description.strip():
                r["description"] = " | ".join([x for x in [r["description"], e.description.strip()] if x])
    return [day_map[k] for k in sorted(day_map.keys())]


def build_recent_ui_items(entries):
    """Dönemler en yeniden eskiye; her dönem içinde kayıtlar en yeni tarihten eskiye."""
    by_period = defaultdict(list)
    for e in entries:
        ps = period_start_for_date(e.work_date)
        by_period[(ps.year, ps.month)].append(e)
    period_keys = sorted(by_period.keys(), reverse=True)
    out = []
    for y, m in period_keys:
        p_start, p_end = period_for_start(y, m)
        out.append(
            {
                "kind": "header",
                "label": f"{format_dmy(p_start)} - {format_dmy(p_end)}",
            }
        )
        period_entries = sorted(
            by_period[(y, m)],
            key=lambda e: (e.work_date, e.start_time or "", e.id),
            reverse=True,
        )
        for e in period_entries:
            out.append({"kind": "entry", "entry": e})
    return out


@app.context_processor
def inject_helpers():
    login_user = session_login_user()
    effective_user = current_user() if login_user else None
    is_founder = is_founder_user(effective_user)
    is_delegate_admin = bool(effective_user and get_delegate_permission(effective_user.id))
    is_impersonating = bool(session.get("admin_impersonate_user_id"))
    # Impersonation geri donus butonu, aktif impersonate oturumu varsa her zaman gorunsun.
    can_return_to_self = bool(session.get("admin_original_user_id") and session.get("admin_impersonate_user_id"))
    return {
        "fmt_num": fmt_num,
        "apk_url": app.config.get("APK_URL", "/download-apk"),
        "is_founder": is_founder,
        "is_real_founder": is_founder_user(login_user),
        "is_delegate_admin": is_delegate_admin,
        "can_view_users_screen": delegate_can(effective_user, "users"),
        "can_view_charts": delegate_can(effective_user, "charts"),
        "can_view_filters": delegate_can(effective_user, "filters"),
        "is_impersonating": is_impersonating,
        "can_return_to_self": can_return_to_self,
    }


@app.after_request
def apply_security_headers(resp):
    resp.headers["X-Frame-Options"] = "DENY"
    resp.headers["X-Content-Type-Options"] = "nosniff"
    resp.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    resp.headers["Content-Security-Policy"] = "default-src 'self'; style-src 'self' 'unsafe-inline'; script-src 'self' 'unsafe-inline'; img-src 'self' data:; frame-ancestors 'none'"
    # Dinamik sayfalar cache'lenmesin: farkli kullaniciya geciste eski profil/veri gorunmesini engeller.
    if not request.path.startswith("/static/"):
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
    return resp


@app.get("/")
def root():
    if "user_id" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/register", methods=["GET", "POST"])
def register():
    flash("Web kayıt olma ekranı kapatıldı.", "error")
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    # Aktif oturum varken /login acilirsa ust menu + giris formu birlikte gorunmesin;
    # gecerli kullanici varsa panele yonlendir, yoksa (silinmis/gecersiz id) oturumu temizle.
    if request.method == "GET":
        raw_uid = session.get("user_id")
        if raw_uid is not None and raw_uid != "":
            try:
                uid_int = int(raw_uid)
            except (TypeError, ValueError):
                uid_int = 0
            if uid_int and User.query.get(uid_int):
                return redirect(url_for("dashboard"))
            session.clear()
    if request.method == "POST":
        ip = request.headers.get("X-Forwarded-For", request.remote_addr or "unknown")
        if is_rate_limited(f"login:{ip}", limit=15, window_sec=60):
            flash("Çok fazla deneme. Lütfen 1 dakika sonra tekrar deneyin.", "error")
            return render_template("login.html")
        identity = request.form.get("email", request.form.get("username_or_email", "")).strip()
        password = request.form.get("password", "")
        user = User.query.filter((User.username == identity) | (User.email == identity)).first()
        if not user or not check_password_hash(user.password_hash, password):
            flash("E-posta veya şifre hatalı.", "error")
            return render_template("login.html")
        if is_user_terminated(int(user.id)):
            flash("İşten ayrılış kaydınız bulunduğu için sisteme giriş yapamazsınız.", "error")
            return render_template("login.html")
        session.clear()
        session["user_id"] = user.id
        session["api_token"] = token_serializer.dumps({"uid": user.id, "nonce": secrets.token_hex(8)})
        return redirect(url_for("dashboard"))
    return render_template("login.html")


@app.get("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


def build_period_options_for_entries(entries):
    start_options = sorted({(period_start_for_date(e.work_date).year, period_start_for_date(e.work_date).month) for e in entries}, reverse=True)
    if not start_options:
        ps = period_start_for_date(date.today())
        start_options = [(ps.year, ps.month)]
    years = sorted({period_year(y, m) for (y, m) in start_options}, reverse=True)
    selected_year = years[0]
    period_options = [(y, m) for (y, m) in start_options if period_year(y, m) == selected_year] or [start_options[0]]
    return years, selected_year, period_options, start_options[0]


@app.get("/admin/users")
@login_required
@admin_or_delegate_required
def admin_users():
    login_user = session_login_user()
    if login_user and (not is_founder_user(login_user)):
        cached_html = get_delegate_view_cache("admin_users", int(login_user.id))
        if cached_html:
            return app.response_class(cached_html)
    effective_user = current_user() if login_user else None
    can_users_screen = delegate_can(effective_user, "users")
    can_charts_screen = delegate_can(effective_user, "charts")
    can_filters = delegate_can(effective_user, "filters")
    can_add_user = delegate_can(effective_user, "add_user")
    can_change_email = delegate_can(effective_user, "change_email")
    can_reset_password = delegate_can(effective_user, "reset_password")
    can_period_lock = delegate_can(effective_user, "period_lock")
    can_bulk_entry = delegate_can(effective_user, "bulk_entry")
    can_terminated_users = delegate_can(effective_user, "terminated_users")
    can_unit_change = delegate_can(effective_user, "unit_change")
    allowed_ids = allowed_user_ids_for(effective_user)
    can_impersonate = delegate_can(effective_user, "impersonate")
    if not can_users_screen:
        flash("Kullanıcı ekranını görme yetkiniz yok.", "error")
        return redirect(url_for("dashboard"))
    # Tum kullanicilari profil ile birlikte listele
    users_query = User.query.order_by(User.created_at.desc())
    users = users_query.all() if allowed_ids is None else users_query.filter(User.id.in_(list(allowed_ids) or [0])).all()
    user_ids = [u.id for u in users]
    profiles = {
        p.user_id: p
        for p in UserProfile.query.filter(UserProfile.user_id.in_(user_ids or [0])).all()
    }
    delegate_perm = None if is_founder_user(effective_user) else get_delegate_permission(effective_user.id if effective_user else 0)
    need_unit_scope = bool(delegate_perm and ((delegate_perm.scope_daire_baskanligi or "").strip() or (delegate_perm.scope_sube_mudurlugu or "").strip()))
    unit_changes_map = unit_changes_map_for_users(user_ids) if need_unit_scope else {}
    entry_counts = {
        uid: cnt
        for uid, cnt in db.session.query(OvertimeEntry.user_id, db.func.count(OvertimeEntry.id))
        .filter(OvertimeEntry.user_id.in_(user_ids or [0]))
        .group_by(OvertimeEntry.user_id)
        .all()
    }
    owner_user = effective_user or login_user
    founder_work_dates = (
        db.session.query(OvertimeEntry.work_date)
        .filter(OvertimeEntry.user_id == owner_user.id)
        .all()
    )
    start_options = sorted(
        {
            (period_start_for_date(wd).year, period_start_for_date(wd).month)
            for (wd,) in founder_work_dates
            if wd is not None
        },
        reverse=True,
    )
    if not start_options:
        ps = period_start_for_date(date.today())
        start_options = [(ps.year, ps.month)]
    years = sorted({period_year(y, m) for (y, m) in start_options}, reverse=True)
    default_year = years[0]
    selected_year = request.args.get("year", type=int) or default_year
    if selected_year not in years:
        selected_year = default_year
    period_options = [(y, m) for (y, m) in start_options if period_year(y, m) == selected_year] or [start_options[0]]
    selected_period = request.args.get("period", "").strip()
    active_start = period_options[0]
    if selected_period and "-" in selected_period:
        try:
            sy, sm = (int(x) for x in selected_period.split("-"))
            if (sy, sm) in period_options:
                active_start = (sy, sm)
        except Exception:
            pass
    visible_users = []
    for u in users:
        p = profiles.get(u.id) or UserProfile(user_id=u.id)
        if p.employment_end_date and not include_user_for_selected_year(p, selected_year):
            continue
        if need_unit_scope and not unit_scope_allows_user_for_year(
            effective_user,
            u.id,
            selected_year,
            profile=p,
            perm=delegate_perm,
            changes=unit_changes_map.get(u.id),
        ):
            continue
        visible_users.append(u)
    rows = []
    founder_email_norm = FOUNDER_EMAIL.strip().lower()
    viewer_is_founder = bool(is_founder_user(effective_user))
    for u in visible_users:
        p = profiles.get(u.id) or UserProfile(user_id=u.id)
        target_is_founder_account = (str(u.email or "").strip().lower() == founder_email_norm)
        can_open_as_user = bool(can_impersonate and (allowed_ids is None or u.id in allowed_ids))
        if target_is_founder_account and not viewer_is_founder:
            can_open_as_user = False
        rows.append(
            {
                "user": u,
                "profile": p,
                "entry_count": int(entry_counts.get(u.id, 0)),
                "can_manage_permissions": bool(is_founder_user(effective_user)),
                "can_reset_password": bool(can_reset_password),
                "can_open_user": can_open_as_user,
                "can_change_email": bool(can_change_email),
                "can_terminated_users": bool(can_terminated_users),
                "can_unit_change": bool(can_unit_change),
            }
        )
    sig_prefix = f"bulk_excel_sign_{owner_user.id}"
    default_title = "" if not is_founder_user(owner_user) else "Ambarlar Şefi"
    default_manager_title = "" if not is_founder_user(owner_user) else "Ambarlar Şube Müdürü"
    default_director_title = "" if not is_founder_user(owner_user) else "Daire Başkanı"
    sign_fields = {
        "chef_title": get_setting_value(f"{sig_prefix}_chef_title", default_title),
        "chef_name": get_setting_value(f"{sig_prefix}_chef_name", ""),
        "manager_title": get_setting_value(f"{sig_prefix}_manager_title", default_manager_title),
        "manager_name": get_setting_value(f"{sig_prefix}_manager_name", ""),
        "director_title": get_setting_value(f"{sig_prefix}_director_title", default_director_title),
        "director_name": get_setting_value(f"{sig_prefix}_director_name", ""),
    }
    html = render_template(
        "admin_users.html",
        rows=rows,
        can_users_screen=can_users_screen,
        can_charts_screen=can_charts_screen,
        can_filters=can_filters,
        can_add_user=can_add_user,
        can_reset_password=can_reset_password,
        can_change_email=can_change_email,
        can_period_lock=can_period_lock,
        can_bulk_entry=can_bulk_entry,
        can_terminated_users=can_terminated_users,
        can_unit_change=can_unit_change,
        daire_options=DAIRE_OPTIONS,
        sube_options=SUBE_OPTIONS,
        years=years,
        selected_year=selected_year,
        period_options=period_options,
        period_value=f"{active_start[0]:04d}-{active_start[1]:02d}",
        sign_fields=sign_fields,
        has_audit_logs_route=bool("admin_audit_logs" in app.view_functions),
    )
    if login_user and (not is_founder_user(login_user)):
        set_delegate_view_cache("admin_users", int(login_user.id), html)
    return html


@app.get("/admin/audit-logs")
@login_required
@admin_or_delegate_required
def admin_audit_logs():
    login_user = session_login_user()
    effective_user = current_user() if login_user else None

    # Personel ekranındaki yil/donem secimi ile ayni davranis.
    period_pairs = db.session.query(AuditLog.period_start_year, AuditLog.period_start_month).filter(
        AuditLog.period_start_year.isnot(None),
        AuditLog.period_start_month.isnot(None),
    ).distinct().all()
    start_options = sorted(
        {(int(y), int(m)) for y, m in period_pairs if y and m and 1 <= int(m) <= 12},
        reverse=True,
    )
    if not start_options:
        ps0 = period_start_for_date(date.today())
        start_options = [(ps0.year, ps0.month)]
    years = sorted({period_year(y, m) for (y, m) in start_options}, reverse=True)
    year_raw = (request.args.get("year") or "").strip()
    year = int(year_raw) if year_raw.isdigit() and int(year_raw) in years else None
    period_options = [(y, m) for (y, m) in start_options if (year is None or period_year(y, m) == year)] or list(start_options)
    period = (request.args.get("period") or "").strip()
    active_start = None
    if period and "-" in period:
        try:
            sy, sm = (int(x) for x in period.split("-"))
            if (sy, sm) in period_options:
                active_start = (sy, sm)
        except Exception:
            pass
    period_value = f"{active_start[0]:04d}-{active_start[1]:02d}" if active_start else ""

    day = (request.args.get("day") or "").strip()
    actor_user_id = request.args.get("actor_user_id", type=int)
    target_user_id = request.args.get("target_user_id", type=int)
    action = (request.args.get("action") or "").strip().lower()
    daire = (request.args.get("daire") or "").strip()
    sube = (request.args.get("sube") or "").strip()

    q = AuditLog.query.order_by(AuditLog.event_time.desc(), AuditLog.id.desc())
    if actor_user_id:
        q = q.filter(AuditLog.actor_user_id == actor_user_id)
    if target_user_id:
        q = q.filter(AuditLog.target_user_id == target_user_id)
    if action in {"create", "update", "delete"}:
        q = q.filter(AuditLog.action == action)
    if daire:
        q = q.filter(AuditLog.daire_baskanligi == daire)
    if sube:
        q = q.filter(AuditLog.sube_mudurlugu == sube)
    if day:
        try:
            d = parse_date(day)
            q = q.filter(AuditLog.work_date == d)
        except Exception:
            pass
    if active_start:
        q = q.filter(
            AuditLog.period_start_year == active_start[0],
            AuditLog.period_start_month == active_start[1],
        )
    elif year is not None:
        candidate_pairs = [(y, m) for (y, m) in start_options if period_year(y, m) == year]
        if candidate_pairs:
            cond = db.or_(*[
                db.and_(AuditLog.period_start_year == y, AuditLog.period_start_month == m)
                for (y, m) in candidate_pairs
            ])
            q = q.filter(cond)

    raw_rows = q.limit(2000).all()
    field_labels = {
        "start_time": "Başlama",
        "end_time": "Bitiş",
        "pct60": "%60 mesai",
        "pct15": "%15 mesai",
        "pazar": "Pazar",
        "bayram": "Bayram",
        "description": "Açıklama",
        "work_date": "Tarih",
        "daire_baskanligi": "Daire Başkanlığı",
        "sube_mudurlugu": "Şube Müdürlüğü",
        "ad_soyad": "Ad Soyad",
        "sicil_no": "Sicil No",
        "ekip_kodu": "Ekip Kodu",
        "email": "E-posta",
    }

    field_aliases = {
        "start_time": {"start_time", "startTime"},
        "end_time": {"end_time", "endTime"},
        "pct60": {"pct60"},
        "pct15": {"pct15"},
        "pazar": {"pazar"},
        "bayram": {"bayram"},
        "description": {"description"},
        "work_date": {"work_date", "workDate"},
        "daire_baskanligi": {"daire_baskanligi", "daireBaskanligi"},
        "sube_mudurlugu": {"sube_mudurlugu", "subeMudurlugu"},
        "ad_soyad": {"ad_soyad", "adSoyad"},
        "sicil_no": {"sicil_no", "sicilNo"},
        "ekip_kodu": {"ekip_kodu", "ekipKodu"},
        "email": {"email"},
    }

    field_order = [
        "start_time",
        "end_time",
        "pct60",
        "pct15",
        "pazar",
        "bayram",
        "description",
        "daire_baskanligi",
        "sube_mudurlugu",
        "ad_soyad",
        "sicil_no",
        "ekip_kodu",
        "email",
    ]
    numeric_diff_fields = {"pct60", "pct15", "pazar", "bayram"}

    def _fmt_val(v):
        if isinstance(v, float):
            if abs(v - round(v)) < 1e-9:
                return str(int(round(v)))
            return str(v).replace(".", ",")
        return str(v or "")

    def _build_diff_text(old_json: str, new_json: str):
        try:
            o = json.loads(old_json or "{}") if old_json else {}
        except Exception:
            o = {}
        try:
            n = json.loads(new_json or "{}") if new_json else {}
        except Exception:
            n = {}

        def _val_from(payload: dict, canonical_key: str):
            for alias in field_aliases.get(canonical_key, {canonical_key}):
                if alias in payload:
                    return payload.get(alias)
            return ""

        def _norm_for_compare(canonical_key: str, value):
            if canonical_key in numeric_diff_fields:
                s = str(value or "").strip().replace(",", ".")
                if not s:
                    return 0.0
                try:
                    return float(s)
                except Exception:
                    return 0.0
            return _fmt_val(value)

        changed = []
        for k in field_order:
            ov_cmp = _norm_for_compare(k, _val_from(o, k))
            nv_cmp = _norm_for_compare(k, _val_from(n, k))
            if k in numeric_diff_fields:
                if abs(float(ov_cmp) - float(nv_cmp)) > 1e-9:
                    changed.append(k)
            elif ov_cmp != nv_cmp:
                changed.append(k)

        # Beklenmedik/ek alanlar da degisti ise kacirmayalim.
        known_aliases = {a for s in field_aliases.values() for a in s}
        extra_keys = set(o.keys()) | set(n.keys())
        for raw_k in sorted(extra_keys):
            if raw_k in known_aliases or str(raw_k).lower() in {"id", "user_id", "userid"}:
                continue
            ov = _fmt_val(o.get(raw_k, ""))
            nv = _fmt_val(n.get(raw_k, ""))
            if ov != nv:
                changed.append(raw_k)

        if not changed:
            return "-", "-"

        before_parts = []
        after_parts = []
        for k in changed:
            ov = _fmt_val(_val_from(o, k) if k in field_aliases else o.get(k, ""))
            nv = _fmt_val(_val_from(n, k) if k in field_aliases else n.get(k, ""))
            if ov:
                label = escape(field_labels.get(k, k))
                val = escape(ov)
                before_parts.append(f'{label} <span class="audit-diff-value">{val}</span>')
            if nv:
                label = escape(field_labels.get(k, k))
                val = escape(nv)
                after_parts.append(f'{label} <span class="audit-diff-value">{val}</span>')

        if not before_parts:
            before_parts = ["-"]
        if not after_parts:
            after_parts = ["-"]
        return "; ".join(before_parts), "; ".join(after_parts)

    def _action_tr(a: str):
        a0 = str(a or "").lower()
        if a0 == "create":
            return "Ekleme"
        if a0 == "update":
            return "Güncelleme"
        if a0 == "delete":
            return "Silme"
        return a

    def _source_tr(s: str):
        m = {
            "web": "Web",
            "apk": "APK",
            "bulk": "Toplu Mesai",
            "import": "Excel İçe Aktar",
            "backup": "Yedek İçe Aktar",
            "admin": "Yönetim",
        }
        return m.get(str(s or "").lower(), s or "-")

    rows = []
    for r in raw_rows:
        before_text, after_text = _build_diff_text(r.old_data_json, r.new_data_json)
        tr_time = (r.event_time + timedelta(hours=3)) if r.event_time else None
        rows.append(
            {
                "id": int(r.id or 0),
                "event_time": tr_time,
                "action_tr": _action_tr(r.action),
                "actor_label": r.actor_label,
                "target_label": r.target_label,
                "work_date": r.work_date,
                "unit_text": f"{(r.daire_baskanligi or '-')}" + " / " + f"{(r.sube_mudurlugu or '-')}",
                "source_tr": _source_tr(r.source),
                "note_tr": "Açıklama",
                "before_text": before_text,
                "after_text": after_text,
            }
        )
    users = User.query.order_by(User.email.asc()).all()
    profiles = {p.user_id: p for p in UserProfile.query.filter(UserProfile.user_id.in_([u.id for u in users] or [0])).all()}
    daire_options = sorted({(p.daire_baskanligi or "").strip() for p in profiles.values() if (p.daire_baskanligi or "").strip()})
    sube_options = sorted({(p.sube_mudurlugu or "").strip() for p in profiles.values() if (p.sube_mudurlugu or "").strip()})
    return render_template(
        "admin_audit_logs.html",
        rows=rows,
        users=users,
        profiles=profiles,
        selected={
            "year": year or "",
            "period": period_value,
            "day": day,
            "actor_user_id": actor_user_id or "",
            "target_user_id": target_user_id or "",
            "action": action,
            "daire": daire,
            "sube": sube,
        },
        years=years,
        period_options=period_options,
        daire_options=daire_options,
        sube_options=sube_options,
        can_audit_clear=bool(is_founder_user(effective_user)),
    )


@app.post("/admin/audit-logs/clear")
@login_required
@admin_required
def admin_audit_logs_clear():
    effective_user = current_user()
    if not is_founder_user(effective_user):
        flash("Audit temizleme yalnızca kurucu hesabında kullanılabilir.", "error")
        return redirect(url_for("admin_audit_logs"))
    selected_ids = [int(v) for v in request.form.getlist("selected_audit_ids") if str(v).isdigit()]
    if selected_ids:
        deleted = AuditLog.query.filter(AuditLog.id.in_(selected_ids)).delete(synchronize_session=False)
        db.session.commit()
        flash(f"İşaretli audit kayıtlarından {deleted} satır silindi.", "success")
        return redirect(url_for("admin_audit_logs"))
    actor_user_id = request.form.get("actor_user_id", type=int)
    target_user_id = request.form.get("target_user_id", type=int)
    action = (request.form.get("action") or "").strip().lower()
    day = (request.form.get("day") or "").strip()
    daire = (request.form.get("daire") or "").strip()
    sube = (request.form.get("sube") or "").strip()
    year_raw = (request.form.get("year") or "").strip()
    period_raw = (request.form.get("period") or "").strip()

    q = AuditLog.query
    if actor_user_id:
        q = q.filter(AuditLog.actor_user_id == actor_user_id)
    if target_user_id:
        q = q.filter(AuditLog.target_user_id == target_user_id)
    if action in {"create", "update", "delete"}:
        q = q.filter(AuditLog.action == action)
    if daire:
        q = q.filter(AuditLog.daire_baskanligi == daire)
    if sube:
        q = q.filter(AuditLog.sube_mudurlugu == sube)
    if day:
        try:
            q = q.filter(AuditLog.work_date == parse_date(day))
        except Exception:
            pass
    if period_raw and "-" in period_raw:
        try:
            sy, sm = (int(x) for x in period_raw.split("-"))
            q = q.filter(
                AuditLog.period_start_year == sy,
                AuditLog.period_start_month == sm,
            )
        except Exception:
            pass
    elif year_raw.isdigit():
        q = q.filter(AuditLog.period_start_year == int(year_raw))

    deleted = q.delete(synchronize_session=False)
    db.session.commit()
    flash(f"Audit kayıtlarından {deleted} satır silindi.", "success")
    return redirect(url_for("admin_audit_logs"))


@app.route("/admin/users/bulk-entry", methods=["GET", "POST"])
@login_required
@admin_or_delegate_required
def admin_users_bulk_entry():
    login_user = session_login_user()
    if request.method == "GET" and login_user:
        cached_html = get_delegate_view_cache("admin_users_bulk_entry", int(login_user.id))
        if cached_html:
            return app.response_class(cached_html)
    if not delegate_can(login_user, "bulk_entry"):
        flash("Toplu mesai girişi yetkiniz yok.", "error")
        return redirect(url_for("admin_users"))
    allowed_ids = allowed_user_ids_for(login_user)
    users_q = User.query.order_by(User.created_at.desc())
    users = users_q.all() if allowed_ids is None else users_q.filter(User.id.in_(list(allowed_ids) or [0])).all()
    user_ids = [u.id for u in users]
    profiles = {
        p.user_id: p
        for p in UserProfile.query.filter(UserProfile.user_id.in_(user_ids or [0])).all()
    }
    rows = [{"user": u, "profile": profiles.get(u.id) or UserProfile(user_id=u.id)} for u in users]

    sort_key = (request.values.get("sort_key") or "").strip()
    sort_dir = (request.values.get("sort_dir") or "asc").strip().lower()
    if sort_dir not in ("asc", "desc"):
        sort_dir = "asc"

    def _row_sort_value(row, key_name: str):
        p = row.get("profile") or UserProfile()
        u = row.get("user")
        if key_name == "sicil_no":
            return (p.sicil_no or "").strip().lower()
        if key_name == "ad_soyad":
            return (p.ad_soyad or "").strip().lower()
        if key_name == "daire_baskanligi":
            return (p.daire_baskanligi or "").strip().lower()
        if key_name == "sube_mudurlugu":
            return (p.sube_mudurlugu or "").strip().lower()
        if key_name == "ekip_kodu":
            return (p.ekip_kodu or "").strip().lower()
        if key_name == "email":
            return (u.email or "").strip().lower() if u else ""
        return ""

    allowed_sort_keys = {"sicil_no", "ad_soyad", "daire_baskanligi", "sube_mudurlugu", "ekip_kodu", "email"}
    if sort_key in allowed_sort_keys:
        rows = sorted(rows, key=lambda r: _row_sort_value(r, sort_key), reverse=(sort_dir == "desc"))

    owner_work_dates = (
        db.session.query(OvertimeEntry.work_date)
        .filter(OvertimeEntry.user_id == login_user.id)
        .all()
    )
    start_options = sorted(
        {
            (
                period_start_for_date(wd).year,
                period_start_for_date(wd).month,
            )
            for (wd,) in owner_work_dates
            if wd is not None
        },
        reverse=True,
    )
    if not start_options:
        ps = period_start_for_date(date.today())
        start_options = [(ps.year, ps.month)]
    years = sorted({period_year(y, m) for (y, m) in start_options}, reverse=True)
    selected_year = request.values.get("year", type=int) or years[0]
    if selected_year not in years:
        selected_year = years[0]
    period_options = [(y, m) for (y, m) in start_options if period_year(y, m) == selected_year] or [start_options[0]]
    period_value = (request.values.get("period") or "").strip()
    active_start = period_options[0]
    if period_value and "-" in period_value:
        try:
            sy, sm = (int(x) for x in period_value.split("-"))
            if (sy, sm) in period_options:
                active_start = (sy, sm)
        except Exception:
            pass
    period_value = f"{active_start[0]:04d}-{active_start[1]:02d}"

    rows = [
        r
        for r in rows
        if include_user_for_selected_year((r.get("profile") or UserProfile(user_id=0)), selected_year)
    ]
    row_user_ids = [int((r.get("user") or User()).id or 0) for r in rows]
    delegate_perm = None if is_founder_user(login_user) else get_delegate_permission(login_user.id if login_user else 0)
    need_unit_scope = bool(delegate_perm and ((delegate_perm.scope_daire_baskanligi or "").strip() or (delegate_perm.scope_sube_mudurlugu or "").strip()))
    unit_changes_map = unit_changes_map_for_users(row_user_ids) if need_unit_scope else {}
    if need_unit_scope:
        rows = [
            r
            for r in rows
            if unit_scope_allows_user_for_year(
                login_user,
                int((r.get("user") or User()).id or 0),
                selected_year,
                profile=(r.get("profile") or None),
                perm=delegate_perm,
                changes=unit_changes_map.get(int((r.get("user") or User()).id or 0)),
            )
        ]

    selected_filter_daire = (request.values.get("bulk_filter_daire") or "").strip()
    selected_filter_sube = (request.values.get("bulk_filter_sube") or "").strip()
    if selected_filter_daire:
        rows = [
            r for r in rows
            if str(((r.get("profile") or UserProfile(user_id=0)).daire_baskanligi or "")).strip() == selected_filter_daire
        ]
    if selected_filter_sube:
        rows = [
            r for r in rows
            if str(((r.get("profile") or UserProfile(user_id=0)).sube_mudurlugu or "")).strip() == selected_filter_sube
        ]

    selected_user_ids = {int(v) for v in request.values.getlist("selected_user_ids") if str(v).isdigit()}
    # Guvenlik ve performans: sadece o anki filtreli listede gorunen kullanicilar islenir.
    row_user_id_set = {int((r.get("user") or User()).id or 0) for r in rows}
    selected_user_ids = {uid for uid in selected_user_ids if uid in row_user_id_set}
    action = (request.values.get("action") or "").strip().lower()
    show_grid = action in ("preview", "save")

    p_start, p_end = period_for_start(active_start[0], active_start[1])
    day_columns = []
    day_styles = {}
    day_defaults_map = {}
    cur = p_start
    while cur <= p_end:
        day_columns.append(cur)
        defaults = day_defaults(cur)
        day_defaults_map[cur.isoformat()] = defaults
        day_styles[cur.isoformat()] = {
            "is_weekend": cur.weekday() >= 5,
            "is_holiday": bool(defaults.get("isHoliday")),
        }
        cur += timedelta(days=1)

    if show_grid and not selected_user_ids:
        flash("Lütfen en az bir kullanıcı seçin.", "error")
        return redirect(
            url_for(
                "admin_users_bulk_entry",
                year=selected_year,
                period=period_value,
                bulk_filter_daire=selected_filter_daire,
                bulk_filter_sube=selected_filter_sube,
            )
        )

    if action == "save":
        try:
            old_rows = OvertimeEntry.query.filter(
                OvertimeEntry.user_id.in_(list(selected_user_ids)),
                OvertimeEntry.work_date >= p_start,
                OvertimeEntry.work_date <= p_end,
            ).all()
            to_insert = []
            # Tum matrisi (kullanici x gun) dolasmak yerine sadece gelen dolu hucreleri isle.
            selected_uid_set = {int(x) for x in selected_user_ids}
            valid_day_map = {d.isoformat(): d for d in day_columns}
            for key, value in request.form.items():
                if not key.startswith("cell_"):
                    continue
                raw_val = (value or "").strip()
                if not raw_val:
                    continue
                parts = key.split("_", 2)
                if len(parts) != 3:
                    continue
                try:
                    uid = int(parts[1])
                except Exception:
                    continue
                if uid not in selected_uid_set:
                    continue
                d = valid_day_map.get(parts[2])
                if not d:
                    continue
                raw = raw_val.replace(",", ".")
                defaults = day_defaults_map.get(d.isoformat()) or day_defaults(d)
                is_special_day = bool(defaults.get("isHoliday")) or int(defaults.get("weekday", -1)) == 6
                start = defaults["start"]
                end = start
                pct60 = 0.0
                pct15 = 0.0
                pazar = 0.0
                bayram = 0.0

                if is_special_day and "+" in raw:
                    left, right = (s.strip() for s in raw.split("+", 1))
                    try:
                        base_val = float(left)
                        extra_hours = float(right)
                    except Exception:
                        continue
                    # Ozel gun hucrelerinde 0+X degerini destekle:
                    # 0 kismi pazar/bayrama yazilmaz, + sonrasi %60'a gider.
                    if extra_hours < 0 or base_val not in (0.0, 0.5, 1.0):
                        continue
                    base_end = str(defaults.get("end") or "17:00")
                    end = add_hours_hhmm(base_end, extra_hours)
                    pct60 = float(extra_hours)
                    pct15 = float(calc_night_20_06(start, end) or 0.0)
                    if base_val == 0.0:
                        pazar = 0.0
                        bayram = 0.0
                    elif bool(defaults.get("isHoliday")):
                        bayram = float(base_val)
                        pazar = 0.0
                    else:
                        pazar = float(base_val)
                        bayram = 0.0
                elif is_special_day:
                    try:
                        val = float(raw)
                    except Exception:
                        continue
                    if val <= 0:
                        continue
                    hk_cell = holiday_kind_tr(d)
                    if val in (1.0, 0.5):
                        end = str(defaults.get("end") or start)
                        sp_b = compute_mesai_split(start, end, d.weekday(), hk_cell)
                        pct60 = float(sp_b.get("pct60", 0) or 0)
                        pct15 = float(sp_b.get("pct15", 0) or 0)
                        pazar = float(sp_b.get("pazar", 0) or 0)
                        bayram = float(sp_b.get("bayram", 0) or 0)
                        if bool(defaults.get("isHoliday")):
                            bayram = float(val)
                            pazar = 0.0
                        else:
                            pazar = float(val)
                            bayram = 0.0
                    else:
                        end = end_hhmm_for_bulk_special_target_pct60(start, val, d.weekday(), hk_cell)
                        sp_b = compute_mesai_split(start, end, d.weekday(), hk_cell)
                        pct60 = float(val)
                        pct15 = float(sp_b.get("pct15", 0) or 0)
                        pazar = 0.0
                        bayram = 0.0
                else:
                    try:
                        hours = float(raw)
                    except Exception:
                        continue
                    if hours <= 0:
                        continue
                    wd = d.weekday()
                    is_sat_no_holiday = wd == 5 and not bool(defaults.get("isHoliday"))
                    if is_sat_no_holiday:
                        end = end_hhmm_for_saturday_net(start, hours)
                    else:
                        end = add_hours_hhmm(start, hours)
                    calc = day_defaults(d, end, start)
                    pct60 = float(calc.get("pct60", 0) or 0)
                    pct15 = float(calc.get("pct15", 0) or 0)
                    pazar = float(calc.get("pazar", 0) or 0)
                    bayram = float(calc.get("bayram", 0) or 0)
                to_insert.append(
                    OvertimeEntry(
                        user_id=uid,
                        work_date=d,
                        start_time=start,
                        end_time=end,
                        pct60=pct60,
                        pct15=pct15,
                        pazar=pazar,
                        bayram=bayram,
                        description="",
                    )
                )
            # Sadece gercek degisiklikleri audit'e yaz.
            def _entry_sig(e: OvertimeEntry):
                return (
                    e.work_date.isoformat() if e.work_date else "",
                    str(e.start_time or ""),
                    str(e.end_time or ""),
                    round(float(e.pct60 or 0), 4),
                    round(float(e.pct15 or 0), 4),
                    round(float(e.pazar or 0), 4),
                    round(float(e.bayram or 0), 4),
                    str(e.description or ""),
                )

            old_by_user = {}
            for r in old_rows:
                old_by_user.setdefault(int(r.user_id), []).append(r)
            new_by_user = {}
            for r in to_insert:
                new_by_user.setdefault(int(r.user_id), []).append(r)

            rows_to_delete = []
            rows_to_add = []
            all_uids = set(old_by_user.keys()) | set(new_by_user.keys())
            for uid in all_uids:
                olds = old_by_user.get(uid, [])
                news = new_by_user.get(uid, [])
                old_counter = Counter(_entry_sig(x) for x in olds)
                new_counter = Counter(_entry_sig(x) for x in news)

                removed = old_counter - new_counter
                added = new_counter - old_counter

                old_bucket = {}
                for x in olds:
                    old_bucket.setdefault(_entry_sig(x), []).append(x)
                new_bucket = {}
                for x in news:
                    new_bucket.setdefault(_entry_sig(x), []).append(x)

                for sig, cnt in removed.items():
                    for old_row in (old_bucket.get(sig) or [])[:cnt]:
                        rows_to_delete.append(old_row)
                        write_overtime_audit_log(
                            action="delete",
                            actor_user_id=(login_user.id if login_user else 0),
                            target_user_id=uid,
                            old_entry=old_row,
                            new_entry=None,
                            source="bulk",
                            note="bulk_delete_changed_only",
                        )
                for sig, cnt in added.items():
                    for new_row in (new_bucket.get(sig) or [])[:cnt]:
                        rows_to_add.append(new_row)
                        write_overtime_audit_log(
                            action="create",
                            actor_user_id=(login_user.id if login_user else 0),
                            target_user_id=uid,
                            old_entry=None,
                            new_entry=new_row,
                            source="bulk",
                            note="bulk_add_changed_only",
                        )

            # Sadece degisen satirlari uygula (degismeyenlere dokunma).
            for old_row in rows_to_delete:
                db.session.delete(old_row)
            if rows_to_add:
                db.session.add_all(rows_to_add)
            db.session.commit()
            flash("Toplu mesai girişi kaydedildi.", "success")
        except Exception as exc:
            db.session.rollback()
            flash(f"Toplu mesai kaydı başarısız: {exc}", "error")

    input_values = {}
    if show_grid:
        existing = (
            db.session.query(
                OvertimeEntry.user_id,
                OvertimeEntry.work_date,
                OvertimeEntry.pct60,
                OvertimeEntry.pazar,
                OvertimeEntry.bayram,
            )
            .filter(
                OvertimeEntry.user_id.in_(list(selected_user_ids)),
                OvertimeEntry.work_date >= p_start,
                OvertimeEntry.work_date <= p_end,
            )
            .order_by(OvertimeEntry.user_id.asc(), OvertimeEntry.work_date.asc(), OvertimeEntry.id.asc())
            .all()
        )
        sums_by_day = {}

        def _fmt_cell_num(v: float) -> str:
            n = float(v or 0)
            if abs(n) < 1e-9:
                return "0"
            if float(n).is_integer():
                return str(int(n))
            return f"{n:.2f}".rstrip("0").rstrip(".").replace(".", ",")

        for uid, work_date, pct60_val, pazar_val, bayram_val in existing:
            k = (int(uid), work_date.isoformat())
            if k not in sums_by_day:
                sums_by_day[k] = {"pct60": 0.0, "pazar": 0.0, "bayram": 0.0}
            sums_by_day[k]["pct60"] += float(pct60_val or 0)
            sums_by_day[k]["pazar"] += float(pazar_val or 0)
            sums_by_day[k]["bayram"] += float(bayram_val or 0)
        # Sadece mevcut kaydi olan hucreleri doldur; uid x gun carpimiyla bos hucreleri dolasma.
        for (uid, day_iso), rec in sums_by_day.items():
            if uid not in selected_user_ids:
                continue
            k = f"cell_{uid}_{day_iso}"
            pct60 = float(rec.get("pct60", 0) or 0)
            pazar = float(rec.get("pazar", 0) or 0)
            bayram = float(rec.get("bayram", 0) or 0)

            if pazar > 0:
                base = _fmt_cell_num(pazar)
                if pct60 > 0:
                    input_values[k] = f"{base}+{_fmt_cell_num(pct60)}"
                else:
                    input_values[k] = base
                continue

            if bayram > 0:
                base = _fmt_cell_num(bayram)
                if pct60 > 0:
                    input_values[k] = f"{base}+{_fmt_cell_num(pct60)}"
                else:
                    input_values[k] = base
                continue

            if pct60 > 0:
                input_values[k] = _fmt_cell_num(pct60)

    bulk_grid_day_meta = []
    if show_grid:
        for d in day_columns:
            dd = day_defaults_map.get(d.isoformat()) or day_defaults(d)
            bulk_grid_day_meta.append(
                {
                    "iso": d.isoformat(),
                    "wd": int(dd.get("weekday", 0)),
                    "hol": bool(dd.get("isHoliday")),
                    "halfHol": bool(dd.get("isHalfHoliday")),
                    "start": str(dd.get("start") or "08:00"),
                    "defEnd": str(dd.get("end") or "17:00"),
                }
            )

    html = render_template(
        "admin_bulk_entry.html",
        rows=rows,
        years=years,
        selected_year=selected_year,
        period_options=period_options,
        period_value=period_value,
        selected_user_ids=selected_user_ids,
        show_grid=show_grid,
        day_columns=day_columns,
        day_styles=day_styles,
        input_values=input_values,
        bulk_grid_day_meta=bulk_grid_day_meta,
        selected_filter_daire=selected_filter_daire,
        selected_filter_sube=selected_filter_sube,
        format_dmy=format_dmy,
        sort_key=sort_key,
        sort_dir=sort_dir,
    )
    if request.method == "GET" and login_user:
        set_delegate_view_cache("admin_users_bulk_entry", int(login_user.id), html)
    return html


@app.get("/admin/backup/export")
@login_required
@admin_required
def admin_backup_export():
    def dt(v):
        return v.isoformat() if v else None

    payload = {
        "exported_at": datetime.utcnow().isoformat(),
        "version": 1,
        "users": [
            {
                "id": u.id,
                "username": u.username,
                "email": u.email,
                "password_hash": u.password_hash,
                "created_at": dt(u.created_at),
            }
            for u in User.query.order_by(User.id.asc()).all()
        ],
        "profiles": [
            {
                "id": p.id,
                "user_id": p.user_id,
                "daire_baskanligi": p.daire_baskanligi,
                "sube_mudurlugu": p.sube_mudurlugu,
                "ad_soyad": p.ad_soyad,
                "sicil_no": p.sicil_no,
                "ekip_kodu": p.ekip_kodu,
                "employment_end_date": p.employment_end_date.isoformat() if p.employment_end_date else None,
            }
            for p in UserProfile.query.order_by(UserProfile.id.asc()).all()
        ],
        "entries": [
            {
                "id": e.id,
                "user_id": e.user_id,
                "work_date": e.work_date.isoformat(),
                "start_time": e.start_time,
                "end_time": e.end_time,
                "pct60": e.pct60,
                "pct15": e.pct15,
                "pazar": e.pazar,
                "bayram": e.bayram,
                "description": e.description,
                "created_at": dt(e.created_at),
                "updated_at": dt(e.updated_at),
            }
            for e in OvertimeEntry.query.order_by(OvertimeEntry.id.asc()).all()
        ],
        "delegated_permissions": [
            {
                "id": p.id,
                "owner_user_id": p.owner_user_id,
                "delegate_user_id": p.delegate_user_id,
                "allowed_user_ids_json": p.allowed_user_ids_json,
                "can_view_passwords": bool(p.can_view_passwords),
                "can_reset_password": bool(p.can_reset_password),
                "can_view_users_screen": bool(p.can_view_users_screen),
                "can_view_charts": bool(p.can_view_charts),
                "can_view_filters": bool(p.can_view_filters),
                "can_add_user": bool(p.can_add_user),
                "can_change_email": bool(p.can_change_email),
                "can_period_lock": bool(p.can_period_lock),
                "can_bulk_entry": bool(p.can_bulk_entry),
                "can_view_terminated_users": bool(p.can_view_terminated_users),
                "can_unit_change": bool(p.can_unit_change),
                "scope_daire_baskanligi": str(p.scope_daire_baskanligi or ""),
                "scope_sube_mudurlugu": str(p.scope_sube_mudurlugu or ""),
                "created_at": dt(p.created_at),
                "updated_at": dt(p.updated_at),
            }
            for p in DelegatedAdminPermission.query.order_by(DelegatedAdminPermission.id.asc()).all()
        ],
        "unit_changes": [
            {
                "id": r.id,
                "user_id": r.user_id,
                "transfer_date": r.transfer_date.isoformat(),
                "from_daire_baskanligi": r.from_daire_baskanligi,
                "from_sube_mudurlugu": r.from_sube_mudurlugu,
                "to_daire_baskanligi": r.to_daire_baskanligi,
                "to_sube_mudurlugu": r.to_sube_mudurlugu,
                "created_at": dt(r.created_at),
            }
            for r in UnitChange.query.order_by(UnitChange.id.asc()).all()
        ],
        "period_locks": [
            {
                "id": r.id,
                "start_year": r.start_year,
                "start_month": r.start_month,
                "is_locked": bool(r.is_locked),
                "created_at": dt(r.created_at),
                "updated_at": dt(r.updated_at),
            }
            for r in PeriodLock.query.order_by(PeriodLock.id.asc()).all()
        ],
        "app_settings": [
            {
                "id": s.id,
                "setting_key": s.setting_key,
                "setting_value": s.setting_value,
            }
            for s in AppSetting.query.order_by(AppSetting.id.asc()).all()
        ],
    }
    mem = io.BytesIO(json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8"))
    mem.seek(0)
    name = f"mesai_tam_yedek_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    return send_file(mem, mimetype="application/json", as_attachment=True, download_name=name)


@app.route("/admin/backup/import", methods=["GET", "POST"])
@login_required
@admin_required
def admin_backup_import():
    if request.method == "GET":
        users = User.query.order_by(User.email.asc()).all()
        profiles = {p.user_id: p for p in UserProfile.query.filter(UserProfile.user_id.in_([u.id for u in users] or [0])).all()}
        rows = [{"user": u, "profile": profiles.get(u.id) or UserProfile(user_id=u.id)} for u in users]
        return render_template("admin_backup_import.html", rows=rows)

    f = request.files.get("backup_file")
    if f is None or not (f.filename or "").strip():
        flash("Yedek dosyası seçin.", "error")
        return redirect(url_for("admin_backup_import"))
    try:
        raw = f.read()
        payload = json.loads(raw.decode("utf-8-sig"))
        if not isinstance(payload, dict):
            raise ValueError("Geçersiz yedek formatı")
    except Exception as exc:
        flash(f"Yedek dosyası okunamadı: {exc}", "error")
        return redirect(url_for("admin_backup_import"))

    def parse_dt(v):
        t = str(v or "").strip()
        if not t:
            return None
        try:
            return datetime.fromisoformat(t)
        except Exception:
            return None

    import_scope = (request.form.get("import_scope") or "selected").strip().lower()
    selected_ids = [int(v) for v in request.form.getlist("selected_user_ids") if str(v).isdigit()]
    selected_id_set = set(selected_ids)
    if import_scope == "selected" and not selected_id_set:
        flash("Kısmi içe aktarma için en az bir personel seçin.", "error")
        return redirect(url_for("admin_backup_import"))

    try:
        if import_scope == "all":
            OvertimeEntry.query.delete(synchronize_session=False)
            UserProfile.query.delete(synchronize_session=False)
            DelegatedAdminPermission.query.delete(synchronize_session=False)
            UnitChange.query.delete(synchronize_session=False)
            PeriodLock.query.delete(synchronize_session=False)
            AppSetting.query.delete(synchronize_session=False)
            User.query.delete(synchronize_session=False)

            for u in payload.get("users", []):
                db.session.add(
                    User(
                        id=int(u.get("id")),
                        username=str(u.get("username", "")).strip(),
                        email=str(u.get("email", "")).strip().lower(),
                        password_hash=str(u.get("password_hash", "")),
                        created_at=parse_dt(u.get("created_at")) or datetime.utcnow(),
                    )
                )
            db.session.flush()

            for p in payload.get("profiles", []):
                db.session.add(
                    UserProfile(
                        id=int(p.get("id")),
                        user_id=int(p.get("user_id")),
                        daire_baskanligi=str(p.get("daire_baskanligi", "")),
                        sube_mudurlugu=str(p.get("sube_mudurlugu", "")),
                        ad_soyad=str(p.get("ad_soyad", "")),
                        sicil_no=str(p.get("sicil_no", "")),
                        ekip_kodu=str(p.get("ekip_kodu", "")),
                        employment_end_date=parse_date(str(p.get("employment_end_date", ""))) if str(p.get("employment_end_date", "")).strip() else None,
                    )
                )

            for e in payload.get("entries", []):
                db.session.add(
                    OvertimeEntry(
                        id=int(e.get("id")),
                        user_id=int(e.get("user_id")),
                        work_date=parse_date(str(e.get("work_date", ""))),
                        start_time=str(e.get("start_time", "")),
                        end_time=str(e.get("end_time", "")),
                        pct60=float(e.get("pct60", 0) or 0),
                        pct15=float(e.get("pct15", 0) or 0),
                        pazar=float(e.get("pazar", 0) or 0),
                        bayram=float(e.get("bayram", 0) or 0),
                        description=str(e.get("description", "")),
                        created_at=parse_dt(e.get("created_at")) or datetime.utcnow(),
                        updated_at=parse_dt(e.get("updated_at")) or datetime.utcnow(),
                    )
                )

            for p in payload.get("delegated_permissions", []):
                db.session.add(
                    DelegatedAdminPermission(
                        id=int(p.get("id")),
                        owner_user_id=int(p.get("owner_user_id")),
                        delegate_user_id=int(p.get("delegate_user_id")),
                        allowed_user_ids_json=str(p.get("allowed_user_ids_json", "[]")),
                        can_view_passwords=bool(p.get("can_view_passwords", False)),
                        can_reset_password=bool(p.get("can_reset_password", False)),
                        can_view_users_screen=bool(p.get("can_view_users_screen", False)),
                        can_view_charts=bool(p.get("can_view_charts", False)),
                        can_view_filters=bool(p.get("can_view_filters", False)),
                        can_add_user=bool(p.get("can_add_user", False)),
                        can_change_email=bool(p.get("can_change_email", False)),
                        can_period_lock=bool(p.get("can_period_lock", False)),
                        can_bulk_entry=bool(p.get("can_bulk_entry", False)),
                        can_view_terminated_users=bool(p.get("can_view_terminated_users", False)),
                        can_unit_change=bool(p.get("can_unit_change", False)),
                        scope_daire_baskanligi=str(p.get("scope_daire_baskanligi", "")),
                        scope_sube_mudurlugu=str(p.get("scope_sube_mudurlugu", "")),
                        created_at=parse_dt(p.get("created_at")) or datetime.utcnow(),
                        updated_at=parse_dt(p.get("updated_at")) or datetime.utcnow(),
                    )
                )

            for r in payload.get("unit_changes", []):
                db.session.add(
                    UnitChange(
                        id=int(r.get("id")),
                        user_id=int(r.get("user_id")),
                        transfer_date=parse_date(str(r.get("transfer_date", ""))),
                        from_daire_baskanligi=str(r.get("from_daire_baskanligi", "")),
                        from_sube_mudurlugu=str(r.get("from_sube_mudurlugu", "")),
                        to_daire_baskanligi=str(r.get("to_daire_baskanligi", "")),
                        to_sube_mudurlugu=str(r.get("to_sube_mudurlugu", "")),
                        created_at=parse_dt(r.get("created_at")) or datetime.utcnow(),
                    )
                )

            for r in payload.get("period_locks", []):
                db.session.add(
                    PeriodLock(
                        id=int(r.get("id")),
                        start_year=int(r.get("start_year")),
                        start_month=int(r.get("start_month")),
                        is_locked=bool(r.get("is_locked", False)),
                        created_at=parse_dt(r.get("created_at")) or datetime.utcnow(),
                        updated_at=parse_dt(r.get("updated_at")) or datetime.utcnow(),
                    )
                )

            for s in payload.get("app_settings", []):
                db.session.add(
                    AppSetting(
                        id=int(s.get("id")),
                        setting_key=str(s.get("setting_key", "")),
                        setting_value=str(s.get("setting_value", "")),
                    )
                )
            db.session.commit()
            invalidate_delegate_permission_cache()
            flash("Tam yedek içe aktarıldı. Tüm veriler geri yüklendi.", "success")
        else:
            payload_users = [u for u in payload.get("users", []) if int(u.get("id") or 0) in selected_id_set]
            payload_profiles = [p for p in payload.get("profiles", []) if int(p.get("user_id") or 0) in selected_id_set]
            payload_entries = [e for e in payload.get("entries", []) if int(e.get("user_id") or 0) in selected_id_set]

            existing_users = {u.id: u for u in User.query.filter(User.id.in_(list(selected_id_set) or [0])).all()}
            for u in payload_users:
                uid = int(u.get("id"))
                row = existing_users.get(uid)
                if row is None:
                    row = User(id=uid)
                    db.session.add(row)
                row.username = str(u.get("username", "")).strip()
                row.email = str(u.get("email", "")).strip().lower()
                row.password_hash = str(u.get("password_hash", ""))
                row.created_at = parse_dt(u.get("created_at")) or row.created_at or datetime.utcnow()

            existing_profiles = {
                p.user_id: p for p in UserProfile.query.filter(UserProfile.user_id.in_(list(selected_id_set) or [0])).all()
            }
            for p in payload_profiles:
                uid = int(p.get("user_id"))
                row = existing_profiles.get(uid)
                if row is None:
                    row = UserProfile(user_id=uid)
                    db.session.add(row)
                row.daire_baskanligi = str(p.get("daire_baskanligi", ""))
                row.sube_mudurlugu = str(p.get("sube_mudurlugu", ""))
                row.ad_soyad = str(p.get("ad_soyad", ""))
                row.sicil_no = str(p.get("sicil_no", ""))
                row.ekip_kodu = str(p.get("ekip_kodu", ""))
                row.employment_end_date = (
                    parse_date(str(p.get("employment_end_date", ""))) if str(p.get("employment_end_date", "")).strip() else None
                )

            actor_user = session_login_user()
            old_rows = OvertimeEntry.query.filter(OvertimeEntry.user_id.in_(list(selected_id_set) or [0])).all()
            for old_row in old_rows:
                write_overtime_audit_log(
                    action="delete",
                    actor_user_id=(actor_user.id if actor_user else 0),
                    target_user_id=old_row.user_id,
                    old_entry=old_row,
                    new_entry=None,
                    source="backup",
                    note="backup_import_selected_replace",
                )
                db.session.delete(old_row)
            for e in payload_entries:
                row = OvertimeEntry(
                    user_id=int(e.get("user_id")),
                    work_date=parse_date(str(e.get("work_date", ""))),
                    start_time=str(e.get("start_time", "")),
                    end_time=str(e.get("end_time", "")),
                    pct60=float(e.get("pct60", 0) or 0),
                    pct15=float(e.get("pct15", 0) or 0),
                    pazar=float(e.get("pazar", 0) or 0),
                    bayram=float(e.get("bayram", 0) or 0),
                    description=str(e.get("description", "")),
                    created_at=parse_dt(e.get("created_at")) or datetime.utcnow(),
                    updated_at=parse_dt(e.get("updated_at")) or datetime.utcnow(),
                )
                db.session.add(row)
                write_overtime_audit_log(
                    action="create",
                    actor_user_id=(actor_user.id if actor_user else 0),
                    target_user_id=row.user_id,
                    old_entry=None,
                    new_entry=row,
                    source="backup",
                    note="backup_import_selected_add",
                )

            db.session.commit()
            flash(
                f"Kısmi yedek içe aktarma tamamlandı. {len(selected_id_set)} seçili personel işlendi, diğer kullanıcılara dokunulmadı.",
                "success",
            )
    except Exception as exc:
        db.session.rollback()
        flash(f"Yedek içe aktarma başarısız: {exc}", "error")
    return redirect(url_for("admin_backup_import"))


@app.get("/admin/users/charts")
@login_required
@admin_or_delegate_required
def admin_users_charts():
    login_user = session_login_user()
    if login_user and (not is_founder_user(login_user)):
        cached_html = get_delegate_view_cache("admin_users_charts", int(login_user.id))
        if cached_html:
            return app.response_class(cached_html)
    if not delegate_can(login_user, "charts"):
        flash("Grafik ekranını görme yetkiniz yok.", "error")
        return redirect(url_for("admin_users"))
    allowed_ids = allowed_user_ids_for(login_user)
    date_range_query = db.session.query(db.func.min(OvertimeEntry.work_date), db.func.max(OvertimeEntry.work_date))
    if allowed_ids is not None:
        date_range_query = date_range_query.filter(OvertimeEntry.user_id.in_(list(allowed_ids) or [0]))
    min_wd, max_wd = date_range_query.first() or (None, None)
    start_options = build_start_options_from_date_range(min_wd, max_wd)
    start_options = sorted(set(start_options), reverse=True)
    years = sorted({period_year(y, m) for (y, m) in start_options}, reverse=True)
    default_year = years[0]
    selected_year = request.args.get("year", type=int) or default_year
    if selected_year not in years:
        selected_year = default_year
    period_options = [(y, m) for (y, m) in start_options if period_year(y, m) == selected_year] or [start_options[0]]
    default_start = period_options[0]
    selected_period = request.args.get("period", "").strip()
    selection_applied = request.args.get("selection_applied") == "1"
    selected_user_ids = {int(v) for v in request.args.getlist("selected_user_ids") if str(v).isdigit()}
    selected_daire = request.args.get("daire", "").strip()
    selected_sube = request.args.get("sube", "").strip()
    active_start = default_start
    if selected_period and "-" in selected_period:
        try:
            sy, sm = (int(x) for x in selected_period.split("-"))
            if (sy, sm) in period_options:
                active_start = (sy, sm)
        except Exception:
            pass
    p_start, p_end = period_for_start(active_start[0], active_start[1])

    users_query = User.query.order_by(User.email.asc())
    users = users_query.all() if allowed_ids is None else users_query.filter(User.id.in_(list(allowed_ids) or [0])).all()
    all_users_for_table = list(users)
    if selected_user_ids:
        users = [u for u in users if u.id in selected_user_ids]
    elif selection_applied:
        users = []
    candidate_ids = list({u.id for u in all_users_for_table} | {u.id for u in users})
    profiles = {
        p.user_id: p
        for p in UserProfile.query.filter(UserProfile.user_id.in_(candidate_ids or [0])).all()
    }
    delegate_perm = None if is_founder_user(login_user) else get_delegate_permission(login_user.id if login_user else 0)
    need_unit_scope = bool(delegate_perm and ((delegate_perm.scope_daire_baskanligi or "").strip() or (delegate_perm.scope_sube_mudurlugu or "").strip()))
    unit_changes_map = unit_changes_map_for_users(candidate_ids) if need_unit_scope else {}
    all_users_for_table = [
        u
        for u in all_users_for_table
        if include_user_for_selected_year((profiles.get(u.id) or UserProfile(user_id=u.id)), selected_year)
        and (
            (not need_unit_scope)
            or unit_scope_allows_user_for_year(
                login_user,
                u.id,
                selected_year,
                profile=(profiles.get(u.id) or UserProfile(user_id=u.id)),
                perm=delegate_perm,
                changes=unit_changes_map.get(u.id),
            )
        )
    ]
    users = [
        u
        for u in users
        if include_user_for_selected_year((profiles.get(u.id) or UserProfile(user_id=u.id)), selected_year)
        and (
            (not need_unit_scope)
            or unit_scope_allows_user_for_year(
                login_user,
                u.id,
                selected_year,
                profile=(profiles.get(u.id) or UserProfile(user_id=u.id)),
                perm=delegate_perm,
                changes=unit_changes_map.get(u.id),
            )
        )
    ]
    users_ids = [u.id for u in users]

    if need_unit_scope:
        period_entries = (
            OvertimeEntry.query.filter(
                OvertimeEntry.work_date >= p_start,
                OvertimeEntry.work_date <= p_end,
                OvertimeEntry.user_id.in_(users_ids or [0]),
            )
            .order_by(OvertimeEntry.user_id.asc(), OvertimeEntry.work_date.asc(), OvertimeEntry.id.asc())
            .all()
        )
        period_agg = aggregate_entries_with_scope(
            period_entries,
            profiles,
            unit_changes_map,
            (delegate_perm.scope_daire_baskanligi if delegate_perm else ""),
            (delegate_perm.scope_sube_mudurlugu if delegate_perm else ""),
        )
    else:
        period_agg_rows = (
            db.session.query(
                OvertimeEntry.user_id,
                db.func.sum(OvertimeEntry.pct60),
                db.func.sum(OvertimeEntry.pct15),
                db.func.sum(OvertimeEntry.pazar),
                db.func.sum(OvertimeEntry.bayram),
            )
            .filter(
                OvertimeEntry.work_date >= p_start,
                OvertimeEntry.work_date <= p_end,
                OvertimeEntry.user_id.in_(users_ids or [0]),
            )
            .group_by(OvertimeEntry.user_id)
            .all()
        )
        period_agg = {
            int(uid): {
                "pct60": float(s60 or 0),
                "pct15": float(s15 or 0),
                "pazar": float(sp or 0),
                "bayram": float(sb or 0),
            }
            for uid, s60, s15, sp, sb in period_agg_rows
        }
    # Yil grafigi, rapor sayfasindaki "donem yili" kuraliyla ayni olmali:
    # Aralikta baslayan donem bir sonraki yila yazilir.
    y_from, y_to = year_period_workdate_bounds(selected_year)
    if need_unit_scope:
        all_year_entries = (
            OvertimeEntry.query.filter(
                OvertimeEntry.user_id.in_(users_ids or [0]),
                OvertimeEntry.work_date >= y_from,
                OvertimeEntry.work_date <= y_to,
            )
            .order_by(OvertimeEntry.user_id.asc(), OvertimeEntry.work_date.asc(), OvertimeEntry.id.asc())
            .all()
        )
        year_agg = aggregate_entries_with_scope(
            all_year_entries,
            profiles,
            unit_changes_map,
            (delegate_perm.scope_daire_baskanligi if delegate_perm else ""),
            (delegate_perm.scope_sube_mudurlugu if delegate_perm else ""),
        )
    else:
        year_agg_rows = (
            db.session.query(
                OvertimeEntry.user_id,
                db.func.sum(OvertimeEntry.pct60),
                db.func.sum(OvertimeEntry.pct15),
                db.func.sum(OvertimeEntry.pazar),
                db.func.sum(OvertimeEntry.bayram),
            )
            .filter(
                OvertimeEntry.user_id.in_(users_ids or [0]),
                OvertimeEntry.work_date >= y_from,
                OvertimeEntry.work_date <= y_to,
            )
            .group_by(OvertimeEntry.user_id)
            .all()
        )
        year_agg = {
            int(uid): {
                "pct60": float(s60 or 0),
                "pct15": float(s15 or 0),
                "pazar": float(sp or 0),
                "bayram": float(sb or 0),
            }
            for uid, s60, s15, sp, sb in year_agg_rows
        }

    rows = []
    rows_period_only = []
    for u in users:
        p = profiles.get(u.id) or UserProfile(user_id=u.id)
        pa = period_agg.get(u.id, {"pct60": 0.0, "pct15": 0.0, "pazar": 0.0, "bayram": 0.0})
        ya = year_agg.get(u.id, {"pct60": 0.0, "pct15": 0.0, "pazar": 0.0, "bayram": 0.0})
        row_data = {
            "email": u.email,
            "name": p.ad_soyad or "-",
            "period": pa,
            "year": ya,
            # Grafik metriği: sadece %60 mesai
            "period_hours": pa["pct60"],
            "year_hours": ya["pct60"],
        }
        rows.append(row_data)
        if abs(float(pa.get("pct60", 0) or 0)) > 1e-9 or abs(float(pa.get("pct15", 0) or 0)) > 1e-9 or abs(float(pa.get("pazar", 0) or 0)) > 1e-9 or abs(float(pa.get("bayram", 0) or 0)) > 1e-9:
            rows_period_only.append(row_data)

    rows_period = sorted(rows_period_only, key=lambda x: x["period_hours"], reverse=True)
    rows_year = sorted(rows, key=lambda x: x["year_hours"], reverse=True)
    rows_year_pazar = sorted(rows, key=lambda x: float(x["year"].get("pazar", 0) or 0), reverse=True)
    rows_year_bayram = sorted(rows, key=lambda x: float(x["year"].get("bayram", 0) or 0), reverse=True)
    max_period = max([r["period_hours"] for r in rows_period] + [1.0])
    max_year = max([r["year_hours"] for r in rows_year] + [1.0])
    max_year_pazar = max([float(r["year"].get("pazar", 0) or 0) for r in rows_year_pazar] + [1.0])
    max_year_bayram = max([float(r["year"].get("bayram", 0) or 0) for r in rows_year_bayram] + [1.0])
    year_total_pazar = sum(float(r["year"].get("pazar", 0) or 0) for r in rows_year)
    year_total_bayram = sum(float(r["year"].get("bayram", 0) or 0) for r in rows_year)

    html = render_template(
        "admin_users_charts.html",
        all_users=all_users_for_table,
        profiles=profiles,
        selected_user_ids=selected_user_ids,
        selected_daire=selected_daire,
        selected_sube=selected_sube,
        can_view_filters=delegate_can(login_user, "filters"),
        rows_period=rows_period,
        rows_year=rows_year,
        rows_year_pazar=rows_year_pazar,
        rows_year_bayram=rows_year_bayram,
        max_period=max_period,
        max_year=max_year,
        max_year_pazar=max_year_pazar,
        max_year_bayram=max_year_bayram,
        years=years,
        selected_year=selected_year,
        period_options=period_options,
        period_value=f"{active_start[0]:04d}-{active_start[1]:02d}",
        period_start=p_start,
        period_end=p_end,
        format_dmy=format_dmy,
        year_total_pazar=year_total_pazar,
        year_total_bayram=year_total_bayram,
    )
    if login_user and (not is_founder_user(login_user)):
        set_delegate_view_cache("admin_users_charts", int(login_user.id), html)
    return html


_PAGE_MARGIN_HALF_CM_RE = re.compile(r"<pageMargins\b[^>]*/>", re.IGNORECASE)


def _xlsx_patch_worksheets_page_margins_half_cm(data: bytes) -> bytes:
    """Zip içindeki OOXML sayfalarında kenar boşluklarını 0,5 cm (inç) olarak yazar.

    openpyxl çoğu zaman doğru yazar; bazı Excel sürümleri açılışta 'Normal' önayarına kayabiliyor.
    Paket düzeyinde sabitlemek şablon + dışa aktarma sonrası Ctrl+P ile uyumu güçlendirir.
    """
    inch = 0.5 / 2.54
    s = f"{inch:.14f}".rstrip("0").rstrip(".")
    tag = (
        f'<pageMargins left="{s}" right="{s}" top="{s}" bottom="{s}" '
        f'header="{s}" footer="{s}" />'
    )
    out = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(data), "r") as zin, zipfile.ZipFile(
        out, "w", compression=zipfile.ZIP_DEFLATED
    ) as zout:
        for info in zin.infolist():
            chunk = zin.read(info.filename)
            if info.filename.startswith("xl/worksheets/") and info.filename.endswith(".xml"):
                text = chunk.decode("utf-8")
                if _PAGE_MARGIN_HALF_CM_RE.search(text):
                    text = _PAGE_MARGIN_HALF_CM_RE.sub(tag, text, count=1)
                elif "<pageSetup" in text:
                    text = text.replace("<pageSetup", tag + "<pageSetup", 1)
                elif "</worksheet>" in text:
                    text = text.replace("</worksheet>", tag + "</worksheet>", 1)
                chunk = text.encode("utf-8")
            zout.writestr(info, chunk)
    return out.getvalue()


@app.post("/admin/users/charts/export.xlsx")
@login_required
@admin_or_delegate_required
def admin_users_charts_export_xlsx():
    login_user = session_login_user()
    if not delegate_can(login_user, "charts"):
        flash("Grafik ekranını görme yetkiniz yok.", "error")
        return redirect(url_for("admin_users"))
    allowed_ids = allowed_user_ids_for(login_user)
    selected_user_ids = {int(v) for v in request.form.getlist("selected_user_ids") if str(v).isdigit()}
    selected_daire = request.form.get("daire", "").strip()
    selected_sube = request.form.get("sube", "").strip()

    date_range_query = db.session.query(db.func.min(OvertimeEntry.work_date), db.func.max(OvertimeEntry.work_date))
    if allowed_ids is not None:
        date_range_query = date_range_query.filter(OvertimeEntry.user_id.in_(list(allowed_ids) or [0]))
    min_wd, max_wd = date_range_query.first() or (None, None)
    start_options = build_start_options_from_date_range(min_wd, max_wd)
    start_options = sorted(set(start_options), reverse=True)
    years = sorted({period_year(y, m) for (y, m) in start_options}, reverse=True)
    default_year = years[0]
    selected_year = request.form.get("year", type=int) or default_year
    if selected_year not in years:
        selected_year = default_year
    period_options = [(y, m) for (y, m) in start_options if period_year(y, m) == selected_year] or [start_options[0]]
    default_start = period_options[0]
    selected_period = request.form.get("period", "").strip()
    active_start = default_start
    if selected_period and "-" in selected_period:
        try:
            sy, sm = (int(x) for x in selected_period.split("-"))
            if (sy, sm) in period_options:
                active_start = (sy, sm)
        except Exception:
            pass
    p_start, p_end = period_for_start(active_start[0], active_start[1])

    users_query = User.query.order_by(User.email.asc())
    users = users_query.all() if allowed_ids is None else users_query.filter(User.id.in_(list(allowed_ids) or [0])).all()
    if selected_user_ids:
        users = [u for u in users if u.id in selected_user_ids]
    # Hiç tik yoksa: seçilen yıl/dönem + daire/şube süzgecine uyan tüm yetkili kullanıcılar rapora girer.
    profiles = {
        p.user_id: p
        for p in UserProfile.query.filter(UserProfile.user_id.in_([u.id for u in users] or [0])).all()
    }
    if selected_daire or selected_sube:
        users = [
            u
            for u in users
            if (
                (
                    not selected_daire
                    or ((profiles.get(u.id) or UserProfile(user_id=u.id)).daire_baskanligi or "").strip() == selected_daire
                )
                and (
                    not selected_sube
                    or ((profiles.get(u.id) or UserProfile(user_id=u.id)).sube_mudurlugu or "").strip() == selected_sube
                )
            )
        ]
    candidate_ids = [u.id for u in users]
    delegate_perm = None if is_founder_user(login_user) else get_delegate_permission(login_user.id if login_user else 0)
    need_unit_scope = bool(delegate_perm and ((delegate_perm.scope_daire_baskanligi or "").strip() or (delegate_perm.scope_sube_mudurlugu or "").strip()))
    unit_changes_map = unit_changes_map_for_users(candidate_ids) if need_unit_scope else {}
    users = [
        u
        for u in users
        if include_user_for_selected_year((profiles.get(u.id) or UserProfile(user_id=u.id)), selected_year)
        and (
            (not need_unit_scope)
            or unit_scope_allows_user_for_year(
                login_user,
                u.id,
                selected_year,
                profile=(profiles.get(u.id) or UserProfile(user_id=u.id)),
                perm=delegate_perm,
                changes=unit_changes_map.get(u.id),
            )
        )
    ]
    users_ids = [u.id for u in users]

    if need_unit_scope:
        period_entries = (
            OvertimeEntry.query.filter(
                OvertimeEntry.work_date >= p_start,
                OvertimeEntry.work_date <= p_end,
                OvertimeEntry.user_id.in_(users_ids or [0]),
            )
            .order_by(OvertimeEntry.user_id.asc(), OvertimeEntry.work_date.asc(), OvertimeEntry.id.asc())
            .all()
        )
        period_agg = aggregate_entries_with_scope(
            period_entries,
            profiles,
            unit_changes_map,
            (delegate_perm.scope_daire_baskanligi if delegate_perm else ""),
            (delegate_perm.scope_sube_mudurlugu if delegate_perm else ""),
        )
    else:
        period_agg_rows = (
            db.session.query(
                OvertimeEntry.user_id,
                db.func.sum(OvertimeEntry.pct60),
                db.func.sum(OvertimeEntry.pct15),
                db.func.sum(OvertimeEntry.pazar),
                db.func.sum(OvertimeEntry.bayram),
            )
            .filter(
                OvertimeEntry.work_date >= p_start,
                OvertimeEntry.work_date <= p_end,
                OvertimeEntry.user_id.in_(users_ids or [0]),
            )
            .group_by(OvertimeEntry.user_id)
            .all()
        )
        period_agg = {
            int(uid): {"pct60": float(s60 or 0), "pct15": float(s15 or 0), "pazar": float(sp or 0), "bayram": float(sb or 0)}
            for uid, s60, s15, sp, sb in period_agg_rows
        }
    y_from, y_to = year_period_workdate_bounds(selected_year)
    if need_unit_scope:
        all_year_entries = (
            OvertimeEntry.query.filter(
                OvertimeEntry.user_id.in_(users_ids or [0]),
                OvertimeEntry.work_date >= y_from,
                OvertimeEntry.work_date <= y_to,
            )
            .order_by(OvertimeEntry.user_id.asc(), OvertimeEntry.work_date.asc(), OvertimeEntry.id.asc())
            .all()
        )
        year_agg = aggregate_entries_with_scope(
            all_year_entries,
            profiles,
            unit_changes_map,
            (delegate_perm.scope_daire_baskanligi if delegate_perm else ""),
            (delegate_perm.scope_sube_mudurlugu if delegate_perm else ""),
        )
    else:
        year_agg_rows = (
            db.session.query(
                OvertimeEntry.user_id,
                db.func.sum(OvertimeEntry.pct60),
                db.func.sum(OvertimeEntry.pct15),
                db.func.sum(OvertimeEntry.pazar),
                db.func.sum(OvertimeEntry.bayram),
            )
            .filter(
                OvertimeEntry.user_id.in_(users_ids or [0]),
                OvertimeEntry.work_date >= y_from,
                OvertimeEntry.work_date <= y_to,
            )
            .group_by(OvertimeEntry.user_id)
            .all()
        )
        year_agg = {
            int(uid): {"pct60": float(s60 or 0), "pct15": float(s15 or 0), "pazar": float(sp or 0), "bayram": float(sb or 0)}
            for uid, s60, s15, sp, sb in year_agg_rows
        }
    rows = []
    rows_period_only = []
    for u in users:
        p = profiles.get(u.id) or UserProfile(user_id=u.id)
        pa = period_agg.get(u.id, {"pct60": 0.0, "pct15": 0.0, "pazar": 0.0, "bayram": 0.0})
        ya = year_agg.get(u.id, {"pct60": 0.0, "pct15": 0.0, "pazar": 0.0, "bayram": 0.0})
        row_data = {"name": p.ad_soyad or "-", "period_hours": pa["pct60"], "year_hours": ya["pct60"], "year": ya}
        rows.append(row_data)
        if abs(float(pa.get("pct60", 0) or 0)) > 1e-9 or abs(float(pa.get("pct15", 0) or 0)) > 1e-9 or abs(float(pa.get("pazar", 0) or 0)) > 1e-9 or abs(float(pa.get("bayram", 0) or 0)) > 1e-9:
            rows_period_only.append(row_data)
    rows_period = sorted(rows_period_only, key=lambda x: x["period_hours"], reverse=True)
    rows_year = sorted(rows, key=lambda x: x["year_hours"], reverse=True)
    rows_year_pazar = sorted(rows, key=lambda x: float(x["year"].get("pazar", 0) or 0), reverse=True)
    rows_year_bayram = sorted(rows, key=lambda x: float(x["year"].get("bayram", 0) or 0), reverse=True)

    user_count = len(rows)
    meta_parts = [str(int(selected_year)), f"{user_count} kişi"]
    if selected_daire:
        meta_parts.append(selected_daire)
    if selected_sube:
        meta_parts.append(selected_sube)
    meta = " · ".join(meta_parts)
    title_period = f"Dönem (%60) {format_dmy(p_start)} - {format_dmy(p_end)} - {meta}"
    title_year = f"Yıl (%60) - {meta}"
    title_pazar = f"Pazar mesaisi (yıl) - {meta}"
    title_bayram = f"Bayram mesaisi (yıl) - {meta}"

    def _set_first_chart_title(ws, text: str) -> None:
        """Yalnızca grafik başlığı metnini günceller."""
        chs = getattr(ws, "_charts", None) or []
        if not chs or not (text or "").strip():
            return
        try:
            chs[0].title = text.strip()
        except Exception:
            pass

    def _sync_chart_series_to_data(ws, last_data_row: int) -> None:
        """Şablondaki sabit kısa aralık yüzünden eksik çubuk kalmasın: seriyi A3:B(last) ile hizala (renk/stil aynı)."""
        chs = getattr(ws, "_charts", None) or []
        if not chs or last_data_row < 3:
            return
        st = (ws.title or "Sheet1").replace("'", "''")
        rng_a = f"'{st}'!$A$3:$A${last_data_row}"
        rng_b = f"'{st}'!$B$3:$B${last_data_row}"
        try:
            ser = chs[0].series[0]
            ser.val.numRef.f = rng_b
            if ser.cat and getattr(ser.cat, "strRef", None) is not None:
                ser.cat.strRef.f = rng_a
        except Exception:
            pass

    template_path = os.path.join(os.path.dirname(__file__), "grafik.xlsx")
    wb = load_workbook(template_path)
    base_ws = wb["grafik"]
    ws2 = wb["grafik (2)"]
    ws3 = wb["grafik (3)"]
    ws4 = wb["grafik (4)"]

    def _apply_excel_print_margins_center(ws):
        """Excel Sayfa Yapısı > Kenar Boşlukları: üst/alt/sol/sağ ve üstbilgi/alttaki 0,5 cm; sayfada yatay+dikey ortala."""
        m_in = 0.5 / 2.54  # 0,5 cm -> inç (OOXML)
        ws.page_margins = PageMargins(
            left=m_in,
            right=m_in,
            top=m_in,
            bottom=m_in,
            header=m_in,
            footer=m_in,
        )
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = True
        ps = ws.page_setup
        if ps.paperSize is None:
            ps.paperSize = 9  # A4 (OOXML); yazdırmada tutarlı kağıt boyutu
        if ps.scale is None:
            ps.scale = 100

    for _pw in (base_ws, ws2, ws3, ws4):
        _apply_excel_print_margins_center(_pw)

    _DATA_COL_START = 1
    _DATA_COL_END = 2
    _DATA_FIRST_ROW = 3
    _DATA_CLEAR_LAST_ROW = 2000

    def fill_sheet(ws, data_rows, value_getter):
        """Yalnızca veri satırlarını doldurur; başlık satırı ve hücre stillerine dokunmaz."""
        for rr in range(_DATA_FIRST_ROW, _DATA_CLEAR_LAST_ROW + 1):
            c1 = ws.cell(row=rr, column=_DATA_COL_START)
            c2 = ws.cell(row=rr, column=_DATA_COL_END)
            c1.value = None
            c2.value = None

        row_num = _DATA_FIRST_ROW
        for r in data_rows:
            ws.cell(row=row_num, column=_DATA_COL_START).value = r["name"]
            ws.cell(row=row_num, column=_DATA_COL_END).value = float(value_getter(r))
            row_num += 1
        return max(_DATA_FIRST_ROW, row_num - 1)

    lr1 = fill_sheet(base_ws, rows_period, lambda r: r["period_hours"])
    _set_first_chart_title(base_ws, title_period)
    _sync_chart_series_to_data(base_ws, lr1)
    lr2 = fill_sheet(ws2, rows_year, lambda r: r["year_hours"])
    _set_first_chart_title(ws2, title_year)
    _sync_chart_series_to_data(ws2, lr2)
    lr3 = fill_sheet(ws3, rows_year_pazar, lambda r: float(r["year"].get("pazar", 0) or 0))
    _set_first_chart_title(ws3, title_pazar)
    _sync_chart_series_to_data(ws3, lr3)
    lr4 = fill_sheet(ws4, rows_year_bayram, lambda r: float(r["year"].get("bayram", 0) or 0))
    _set_first_chart_title(ws4, title_bayram)
    _sync_chart_series_to_data(ws4, lr4)

    # Kayıt öncesi tekrar: tüm kenar + üstbilgi/alttaki 0,5 cm (şablonda sıfır XML değeri Excel'de yanıltıcı olabiliyor).
    for _pw in (base_ws, ws2, ws3, ws4):
        _apply_excel_print_margins_center(_pw)

    mem = io.BytesIO()
    wb.save(mem)
    patched = _xlsx_patch_worksheets_page_margins_half_cm(mem.getvalue())
    return send_file(
        io.BytesIO(patched),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Grafik_Raporu_{selected_year}_{active_start[0]:04d}-{active_start[1]:02d}.xlsx",
    )


@app.get("/admin/users/<int:target_user_id>/show-password")
@login_required
@admin_or_delegate_required
def admin_show_password(target_user_id: int):
    login_user = session_login_user()
    allowed_ids = allowed_user_ids_for(login_user)
    if allowed_ids is not None and target_user_id not in allowed_ids:
        flash("Bu kullanıcı için yetkiniz yok.", "error")
        return redirect(url_for("admin_users"))
    delegate_perm = get_delegate_permission(login_user.id) if login_user else None
    can_view = bool(is_founder_user(login_user) or (delegate_perm.can_reset_password if delegate_perm else False))
    target = User.query.get(target_user_id)
    if not target:
        flash("Kullanıcı bulunamadı.", "error")
        return redirect(url_for("admin_users"))
    if not can_view:
        flash("Şifre görme yetkiniz yok.", "error")
        return redirect(url_for("admin_users"))
    flash(
        f"{target.email} için mevcut şifre güvenlik nedeniyle görüntülenemez (hashli saklanır). Bunun yerine 'Şifre Sıfırla' kullanın.",
        "error",
    )
    return redirect(url_for("admin_users"))


@app.post("/admin/users/<int:target_user_id>/reset-password")
@login_required
@admin_or_delegate_required
def admin_reset_password(target_user_id: int):
    login_user = session_login_user()
    allowed_ids = allowed_user_ids_for(login_user)
    if allowed_ids is not None and target_user_id not in allowed_ids:
        flash("Bu kullanıcı için yetkiniz yok.", "error")
        return redirect(url_for("admin_users"))
    delegate_perm = get_delegate_permission(login_user.id) if login_user else None
    can_reset = bool(is_founder_user(login_user) or (delegate_perm.can_reset_password if delegate_perm else False))
    if not can_reset:
        flash("Şifre sıfırlama yetkiniz yok.", "error")
        return redirect(url_for("admin_users"))

    target = User.query.get(target_user_id)
    if not target:
        flash("Kullanıcı bulunamadı.", "error")
        return redirect(url_for("admin_users"))

    # Kolay okunur geçici şifre üretimi (kullanıcı ilk girişte değiştirmeli).
    temp_password = secrets.token_urlsafe(9)[:12]
    target.password_hash = generate_password_hash(temp_password)
    db.session.commit()
    flash(
        f"{target.email} için geçici şifre oluşturuldu: {temp_password} (kullanıcı giriş yapınca şifresini değiştirmeli).",
        "success",
    )
    return redirect(url_for("admin_users"))


@app.post("/admin/users/<int:target_user_id>/change-email")
@login_required
@admin_or_delegate_required
def admin_change_email(target_user_id: int):
    login_user = session_login_user()
    allowed_ids = allowed_user_ids_for(login_user)
    if allowed_ids is not None and target_user_id not in allowed_ids:
        flash("Bu kullanıcı için yetkiniz yok.", "error")
        return redirect(url_for("admin_users"))
    if not delegate_can(login_user, "change_email"):
        flash("E-posta değiştirme yetkiniz yok.", "error")
        return redirect(url_for("admin_users"))

    target = User.query.get(target_user_id)
    if not target:
        flash("Kullanıcı bulunamadı.", "error")
        return redirect(url_for("admin_users"))

    new_email = (request.form.get("new_email") or "").strip().lower()
    if not new_email:
        flash("Yeni e-posta boş olamaz.", "error")
        return redirect(url_for("admin_users"))
    if "@" not in new_email or "." not in new_email:
        flash("Geçerli bir e-posta girin.", "error")
        return redirect(url_for("admin_users"))

    conflict = User.query.filter(User.email == new_email, User.id != target.id).first()
    if conflict:
        flash("Bu e-posta başka bir kullanıcıda kayıtlı.", "error")
        return redirect(url_for("admin_users"))

    old_email = target.email
    target.email = new_email
    target.username = new_email
    db.session.commit()
    flash(f"{old_email} kullanıcısının e-postası {new_email} olarak güncellendi.", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/permissions/<int:target_user_id>", methods=["GET", "POST"])
@login_required
@admin_required
def admin_edit_permission(target_user_id: int):
    founder = session_login_user()
    if not founder:
        return redirect(url_for("login"))
    target = User.query.get(target_user_id)
    if not target:
        flash("Kullanıcı bulunamadı.", "error")
        return redirect(url_for("admin_users"))
    if is_founder_user(target):
        flash("Kurucu kullanıcı için bu işlem yapılamaz.", "error")
        return redirect(url_for("admin_users"))

    perm = DelegatedAdminPermission.query.filter_by(owner_user_id=founder.id, delegate_user_id=target.id).first()
    if request.method == "POST":
        allowed_ids = [int(v) for v in request.form.getlist("allowed_user_ids") if str(v).isdigit()]
        can_reset_password = request.form.get("can_reset_password") == "1"
        can_view_users_screen = request.form.get("can_view_users_screen") == "1"
        can_view_charts = request.form.get("can_view_charts") == "1"
        can_view_filters = request.form.get("can_view_filters") == "1"
        can_add_user = request.form.get("can_add_user") == "1"
        can_change_email = request.form.get("can_change_email") == "1"
        can_period_lock = request.form.get("can_period_lock") == "1"
        can_bulk_entry = request.form.get("can_bulk_entry") == "1"
        can_view_terminated_users = request.form.get("can_view_terminated_users") == "1"
        can_unit_change = request.form.get("can_unit_change") == "1"
        scope_daire_baskanligi = request.form.get("scope_daire_baskanligi", "").strip()
        scope_sube_mudurlugu = request.form.get("scope_sube_mudurlugu", "").strip()
        if perm is None:
            perm = DelegatedAdminPermission(owner_user_id=founder.id, delegate_user_id=target.id)
            db.session.add(perm)
        perm.allowed_user_ids_json = json.dumps(sorted(set(allowed_ids)))
        perm.can_view_passwords = can_reset_password
        perm.can_reset_password = can_reset_password
        perm.can_view_users_screen = can_view_users_screen
        perm.can_view_charts = can_view_charts
        perm.can_view_filters = can_view_filters
        perm.can_add_user = can_add_user
        perm.can_change_email = can_change_email
        perm.can_period_lock = can_period_lock
        perm.can_bulk_entry = can_bulk_entry
        perm.can_view_terminated_users = can_view_terminated_users
        perm.can_unit_change = can_unit_change
        perm.scope_daire_baskanligi = scope_daire_baskanligi
        perm.scope_sube_mudurlugu = scope_sube_mudurlugu
        db.session.commit()
        invalidate_delegate_permission_cache(target.id)
        flash("Yetkiler kaydedildi.", "success")
        return redirect(url_for("admin_authorized_users"))

    current_allowed = set()
    if perm:
        try:
            current_allowed = {int(x) for x in json.loads(perm.allowed_user_ids_json or "[]") if str(x).isdigit()}
        except Exception:
            current_allowed = set()
    users = User.query.order_by(User.email.asc()).all()
    return render_template(
        "admin_permission_edit.html",
        target=target,
        users=users,
        profiles={p.user_id: p for p in UserProfile.query.all()},
        current_allowed=current_allowed,
        can_reset_password=bool(perm.can_reset_password) if perm else False,
        can_view_users_screen=bool(perm.can_view_users_screen) if perm else False,
        can_view_charts=bool(perm.can_view_charts) if perm else False,
        can_view_filters=bool(perm.can_view_filters) if perm else False,
        can_add_user=bool(perm.can_add_user) if perm else False,
        can_change_email=bool(perm.can_change_email) if perm else False,
        can_period_lock=bool(perm.can_period_lock) if perm else False,
        can_bulk_entry=bool(perm.can_bulk_entry) if perm else False,
        can_view_terminated_users=bool(perm.can_view_terminated_users) if perm else False,
        can_unit_change=bool(perm.can_unit_change) if perm else False,
        scope_daire_baskanligi=(perm.scope_daire_baskanligi if perm else ""),
        scope_sube_mudurlugu=(perm.scope_sube_mudurlugu if perm else ""),
        daire_options=DAIRE_OPTIONS,
        sube_options=SUBE_OPTIONS,
    )


@app.get("/admin/period-locks")
@login_required
@admin_or_delegate_required
def admin_period_locks():
    login_user = session_login_user()
    if not delegate_can(login_user, "period_lock"):
        flash("Dönem kilidi ekranı yetkiniz yok.", "error")
        return redirect(url_for("admin_users"))
    allowed_ids = allowed_user_ids_for(login_user)
    q = db.session.query(OvertimeEntry.work_date).distinct()
    if allowed_ids is not None:
        q = q.filter(OvertimeEntry.user_id.in_(list(allowed_ids) or [0]))
    work_dates = [row[0] for row in q.all()]
    start_options = sorted({period_start_key_for_date(wd) for wd in work_dates}, reverse=True)
    if not start_options:
        ps = period_start_for_date(date.today())
        start_options = [(ps.year, ps.month)]
    lock_rows = PeriodLock.query.all()
    lock_map = {(r.start_year, r.start_month): bool(r.is_locked) for r in lock_rows}
    grouped = {}
    for sy, sm in start_options:
        py = period_year(sy, sm)
        p_start, p_end = period_for_start(sy, sm)
        grouped.setdefault(py, []).append(
            {
                "start_year": sy,
                "start_month": sm,
                "start_label": format_dmy(p_start),
                "end_label": format_dmy(p_end),
                "is_locked": lock_map.get((sy, sm), False),
            }
        )
    year_groups = [(y, grouped[y]) for y in sorted(grouped.keys(), reverse=True)]
    return render_template("admin_period_locks.html", year_groups=year_groups)


@app.post("/admin/period-locks/toggle")
@login_required
@admin_or_delegate_required
def admin_period_locks_toggle():
    login_user = session_login_user()
    if not delegate_can(login_user, "period_lock"):
        flash("Dönem kilidi değiştirme yetkiniz yok.", "error")
        return redirect(url_for("admin_users"))
    try:
        sy = int(request.form.get("start_year", "0"))
        sm = int(request.form.get("start_month", "0"))
    except Exception:
        flash("Geçersiz dönem bilgisi.", "error")
        return redirect(url_for("admin_period_locks"))
    if sm < 1 or sm > 12 or sy < 1900:
        flash("Geçersiz dönem bilgisi.", "error")
        return redirect(url_for("admin_period_locks"))
    action = (request.form.get("action") or "").strip().lower()
    lock = PeriodLock.query.filter_by(start_year=sy, start_month=sm).first()
    if lock is None:
        lock = PeriodLock(start_year=sy, start_month=sm, is_locked=False)
        db.session.add(lock)
    lock.is_locked = action == "lock"
    db.session.commit()
    flash("Dönem kilitlendi." if lock.is_locked else "Dönem kilidi açıldı.", "success")
    return redirect(url_for("admin_period_locks"))


@app.post("/admin/users/<int:target_user_id>/set-terminated")
@login_required
@admin_or_delegate_required
def admin_set_terminated(target_user_id: int):
    login_user = session_login_user()
    if not delegate_can(login_user, "terminated_users"):
        flash("İşten ayrılanlar işlemi yetkiniz yok.", "error")
        return redirect(url_for("admin_users"))
    target = User.query.get(target_user_id)
    if not target:
        flash("Kullanıcı bulunamadı.", "error")
        return redirect(url_for("admin_users"))
    d_raw = (request.form.get("employment_end_date") or "").strip()
    try:
        end_date = datetime.strptime(d_raw, "%Y-%m-%d").date()
    except Exception:
        flash("Geçerli bir işten ayrılma tarihi seçin.", "error")
        return redirect(url_for("admin_users"))
    profile = get_or_create_profile(target.id)
    profile.employment_end_date = end_date
    db.session.commit()
    flash("İşten ayrılma tarihi kaydedildi. Kullanıcı girişi kapatıldı.", "success")
    return redirect(url_for("admin_users"))


@app.post("/admin/users/<int:target_user_id>/cancel-terminated")
@login_required
@admin_or_delegate_required
def admin_cancel_terminated(target_user_id: int):
    login_user = session_login_user()
    if not delegate_can(login_user, "terminated_users"):
        flash("İşten ayrılanlar işlemi yetkiniz yok.", "error")
        return redirect(url_for("admin_users"))
    target = User.query.get(target_user_id)
    if not target:
        flash("Kullanıcı bulunamadı.", "error")
        return redirect(url_for("admin_users"))
    profile = get_or_create_profile(target.id)
    profile.employment_end_date = None
    db.session.commit()
    flash("İşten ayrılma kaydı iptal edildi.", "success")
    return redirect(url_for("admin_terminated_users"))


@app.get("/admin/users/terminated")
@login_required
@admin_or_delegate_required
def admin_terminated_users():
    login_user = session_login_user()
    if not delegate_can(login_user, "terminated_users"):
        flash("İşten ayrılanlar ekranını görme yetkiniz yok.", "error")
        return redirect(url_for("admin_users"))
    allowed_ids = allowed_user_ids_for(login_user)
    profiles_q = UserProfile.query.filter(UserProfile.employment_end_date.isnot(None))
    if allowed_ids is not None:
        profiles_q = profiles_q.filter(UserProfile.user_id.in_(list(allowed_ids) or [0]))
    terminated_profiles = profiles_q.order_by(UserProfile.employment_end_date.desc(), UserProfile.ad_soyad.asc()).all()
    users = {u.id: u for u in User.query.filter(User.id.in_([p.user_id for p in terminated_profiles] or [0])).all()}
    rows = []
    for p in terminated_profiles:
        u = users.get(p.user_id)
        if not u:
            continue
        rows.append({"user": u, "profile": p})
    return render_template("admin_terminated_users.html", rows=rows, format_dmy=format_dmy)


@app.post("/admin/users/<int:target_user_id>/unit-change")
@login_required
@admin_or_delegate_required
def admin_add_unit_change(target_user_id: int):
    login_user = session_login_user()
    if not delegate_can(login_user, "unit_change"):
        flash("Birim değişikliği işlemi yetkiniz yok.", "error")
        return redirect(url_for("admin_users"))
    target = User.query.get(target_user_id)
    if not target:
        flash("Kullanıcı bulunamadı.", "error")
        return redirect(url_for("admin_users"))
    profile = get_or_create_profile(target.id)
    from_daire = (request.form.get("from_daire_baskanligi") or "").strip()
    from_sube = (request.form.get("from_sube_mudurlugu") or "").strip()
    to_daire = (request.form.get("to_daire_baskanligi") or "").strip()
    to_sube = (request.form.get("to_sube_mudurlugu") or "").strip()
    date_raw = (request.form.get("transfer_date") or "").strip()
    try:
        transfer_date = datetime.strptime(date_raw, "%Y-%m-%d").date()
    except Exception:
        flash("Geçerli bir tarih seçin.", "error")
        return redirect(url_for("admin_users"))
    if not from_daire or not from_sube or not to_daire or not to_sube:
        flash("Mevcut ve yeni birim bilgilerini eksiksiz seçin.", "error")
        return redirect(url_for("admin_users"))
    rec = UnitChange(
        user_id=target.id,
        transfer_date=transfer_date,
        from_daire_baskanligi=from_daire,
        from_sube_mudurlugu=from_sube,
        to_daire_baskanligi=to_daire,
        to_sube_mudurlugu=to_sube,
    )
    db.session.add(rec)
    # Profili güncel birime çek (sonraki normal ekranlar güncel birimi görsün)
    profile.daire_baskanligi = to_daire
    profile.sube_mudurlugu = to_sube
    db.session.commit()
    flash("Birim değişikliği kaydedildi.", "success")
    return redirect(url_for("admin_users"))


@app.get("/admin/users/unit-changes")
@login_required
@admin_or_delegate_required
def admin_unit_changes():
    login_user = session_login_user()
    if not delegate_can(login_user, "unit_change"):
        flash("Birim değişikliği ekranı yetkiniz yok.", "error")
        return redirect(url_for("admin_users"))
    allowed_ids = allowed_user_ids_for(login_user)
    q = UnitChange.query.order_by(UnitChange.transfer_date.desc(), UnitChange.id.desc())
    if allowed_ids is not None:
        q = q.filter(UnitChange.user_id.in_(list(allowed_ids) or [0]))
    rows_raw = q.all()
    users = {u.id: u for u in User.query.filter(User.id.in_([r.user_id for r in rows_raw] or [0])).all()}
    profiles = {p.user_id: p for p in UserProfile.query.filter(UserProfile.user_id.in_([r.user_id for r in rows_raw] or [0])).all()}
    rows = []
    for r in rows_raw:
        u = users.get(r.user_id)
        if not u:
            continue
        p = profiles.get(r.user_id) or UserProfile(user_id=r.user_id)
        rows.append({"change": r, "user": u, "profile": p})
    return render_template("admin_unit_changes.html", rows=rows, format_dmy=format_dmy)


@app.post("/admin/users/unit-changes/<int:change_id>/cancel")
@login_required
@admin_or_delegate_required
def admin_cancel_unit_change(change_id: int):
    login_user = session_login_user()
    if not delegate_can(login_user, "unit_change"):
        flash("Birim değişikliği iptal yetkiniz yok.", "error")
        return redirect(url_for("admin_users"))
    rec = UnitChange.query.get(change_id)
    if not rec:
        flash("Birim değişikliği kaydı bulunamadı.", "error")
        return redirect(url_for("admin_unit_changes"))
    db.session.delete(rec)
    db.session.commit()
    flash("Birim değişikliği kaydı iptal edildi.", "success")
    return redirect(url_for("admin_unit_changes"))


@app.route("/admin/users/new", methods=["GET", "POST"])
@login_required
@admin_or_delegate_required
def admin_add_user():
    login_user = session_login_user()
    if not login_user:
        return redirect(url_for("login"))
    if not delegate_can(login_user, "add_user"):
        flash("Kişi ekleme yetkiniz yok.", "error")
        return redirect(url_for("admin_users"))

    daire_options = list(DAIRE_OPTIONS)
    sube_options = list(SUBE_OPTIONS)

    if request.method == "POST":
        daire = request.form.get("daire_baskanligi", "").strip()
        sube = request.form.get("sube_mudurlugu", "").strip()
        ad_soyad = request.form.get("ad_soyad", "").strip()
        sicil_no = request.form.get("sicil_no", "").strip()
        ekip_kodu = request.form.get("ekip_kodu", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        password_confirm = request.form.get("password_confirm", "")

        if not email or "@" not in email:
            flash("Geçerli bir e-posta girin.", "error")
            return render_template("admin_user_add.html", daire_options=daire_options, sube_options=sube_options)
        if len(password) < 6:
            flash("Şifre en az 6 karakter olmalı.", "error")
            return render_template("admin_user_add.html", daire_options=daire_options, sube_options=sube_options)
        if password != password_confirm:
            flash("Şifre tekrar alanı uyuşmuyor.", "error")
            return render_template("admin_user_add.html", daire_options=daire_options, sube_options=sube_options)
        if User.query.filter((User.username == email) | (User.email == email)).first():
            flash("Bu e-posta zaten kayıtlı.", "error")
            return render_template("admin_user_add.html", daire_options=daire_options, sube_options=sube_options)

        user = User(username=email, email=email, password_hash=generate_password_hash(password))
        db.session.add(user)
        db.session.flush()
        profile = get_or_create_profile(user.id)
        profile.daire_baskanligi = daire
        profile.sube_mudurlugu = sube
        profile.ad_soyad = ad_soyad
        profile.sicil_no = sicil_no
        profile.ekip_kodu = ekip_kodu
        db.session.commit()
        flash("Yeni kişi eklendi.", "success")
        return redirect(url_for("admin_users"))

    return render_template("admin_user_add.html", daire_options=daire_options, sube_options=sube_options)


@app.get("/admin/authorized-users")
@login_required
@admin_required
def admin_authorized_users():
    founder = session_login_user()
    perms = DelegatedAdminPermission.query.filter_by(owner_user_id=founder.id).all()
    delegates = []
    for p in perms:
        u = User.query.get(p.delegate_user_id)
        if not u:
            continue
        try:
            allowed_count = len({int(x) for x in json.loads(p.allowed_user_ids_json or "[]") if str(x).isdigit()})
        except Exception:
            allowed_count = 0
        delegates.append({"user": u, "perm": p, "allowed_count": allowed_count})
    delegates.sort(key=lambda x: (x["user"].email or "").lower())
    return render_template("admin_authorized_users.html", delegates=delegates)


@app.post("/admin/authorized-users/<int:delegate_user_id>/remove")
@login_required
@admin_required
def admin_remove_authorized_user(delegate_user_id: int):
    founder = session_login_user()
    perm = DelegatedAdminPermission.query.filter_by(owner_user_id=founder.id, delegate_user_id=delegate_user_id).first()
    if perm:
        db.session.delete(perm)
        db.session.commit()
        invalidate_delegate_permission_cache(delegate_user_id)
        flash("Yetki kaldırıldı.", "success")
    else:
        flash("Yetki kaydı bulunamadı.", "error")
    return redirect(url_for("admin_authorized_users"))


@app.post("/admin/users/<int:target_user_id>/delete")
@login_required
@admin_required
def admin_delete_user(target_user_id: int):
    founder = session_login_user()
    target = User.query.get(target_user_id)
    if not target:
        flash("Kullanıcı bulunamadı.", "error")
        return redirect(url_for("admin_users"))
    if founder and target.id == founder.id:
        flash("Kurucu kullanıcı kendisini silemez.", "error")
        return redirect(url_for("admin_users"))

    try:
        # Kullanıcıya ait tüm verileri temizle
        OvertimeEntry.query.filter_by(user_id=target.id).delete()
        UserProfile.query.filter_by(user_id=target.id).delete()
        DelegatedAdminPermission.query.filter(
            (DelegatedAdminPermission.delegate_user_id == target.id)
            | (DelegatedAdminPermission.owner_user_id == target.id)
        ).delete(synchronize_session=False)
        db.session.delete(target)
        db.session.commit()
        invalidate_delegate_permission_cache(target.id)
        flash("Kullanıcı ve tüm verileri silindi.", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Kullanıcı silinemedi: {exc}", "error")
    return redirect(url_for("admin_users"))


@app.get("/admin/impersonate/<int:target_user_id>")
@login_required
@admin_or_delegate_required
def admin_impersonate(target_user_id: int):
    target = User.query.get(target_user_id)
    if not target:
        flash("Kullanıcı bulunamadı.", "error")
        return redirect(url_for("admin_users"))
    login_user = session_login_user()
    if not login_user:
        return redirect(url_for("login"))
    if not delegate_can(login_user, "impersonate"):
        flash("Kullanıcı ekranı görme yetkiniz yok.", "error")
        return redirect(url_for("admin_users"))
    allowed_ids = allowed_user_ids_for(login_user)
    if allowed_ids is not None and target.id not in allowed_ids:
        flash("Bu kullanıcıyı açma yetkiniz yok.", "error")
        return redirect(url_for("admin_users"))
    session["admin_original_user_id"] = login_user.id
    session["admin_impersonate_user_id"] = target.id
    session["user_id"] = login_user.id
    flash(f"{target.email} kullanıcısı olarak görüntüleme açıldı.", "success")
    return redirect(url_for("dashboard"))


@app.post("/admin/stop-impersonation")
@login_required
@admin_or_delegate_required
def admin_stop_impersonation():
    session.pop("admin_impersonate_user_id", None)
    session.pop("admin_original_user_id", None)
    flash("Kendi sayfanıza geri dönüldü.", "success")
    return redirect(url_for("admin_users"))


@app.post("/admin/users/export.xlsx")
@login_required
@admin_or_delegate_required
def admin_export_selected_users_xlsx():
    selected_ids = [int(v) for v in request.form.getlist("selected_user_ids") if str(v).isdigit()]
    year = request.form.get("year", type=int)
    period = request.form.get("period", "").strip()
    if not selected_ids:
        flash("Lütfen en az bir kişi seçin.", "error")
        return redirect(url_for("admin_users"))
    login_user = session_login_user()
    allowed_ids = allowed_user_ids_for(login_user)
    if allowed_ids is not None:
        selected_ids = [uid for uid in selected_ids if uid in allowed_ids]
        if not selected_ids:
            flash("Seçtiğiniz kullanıcılar için yetkiniz yok.", "error")
            return redirect(url_for("admin_users"))
    if not year or "-" not in period:
        flash("Yıl/dönem bilgisi eksik.", "error")
        return redirect(url_for("admin_users"))
    try:
        # admin_users ekraninda period degeri YYYY-MM olarak gelir;
        # Aralik/Ocak gibi donemlerde baslangic yili period degerinden alinmali.
        sy, sm = (int(x) for x in period.split("-"))
    except Exception:
        flash("Dönem formatı hatalı.", "error")
        return redirect(url_for("admin_users"))

    users = User.query.filter(User.id.in_(selected_ids)).all()
    profiles = {p.user_id: p for p in UserProfile.query.filter(UserProfile.user_id.in_(selected_ids)).all()}
    delegate_perm = None if is_founder_user(login_user) else get_delegate_permission(login_user.id if login_user else 0)
    need_unit_scope = bool(delegate_perm and ((delegate_perm.scope_daire_baskanligi or "").strip() or (delegate_perm.scope_sube_mudurlugu or "").strip()))
    unit_changes_map = unit_changes_map_for_users([u.id for u in users]) if need_unit_scope else {}
    users = [
        u
        for u in users
        if include_user_for_selected_year((profiles.get(u.id) or UserProfile(user_id=u.id)), year)
        and unit_scope_allows_user_for_year(
            login_user,
            u.id,
            year,
            profile=(profiles.get(u.id) or UserProfile(user_id=u.id)),
            perm=delegate_perm,
            changes=unit_changes_map.get(u.id),
        )
    ]
    selected_ids = [u.id for u in users]
    if not selected_ids:
        flash("Seçilen yıl için listelenebilir kullanıcı bulunamadı.", "error")
        return redirect(url_for("admin_users"))
    profiles = {uid: profiles.get(uid) or UserProfile(user_id=uid) for uid in selected_ids}
    p_start, p_end = period_for_start(sy, sm)

    sig_prefix = f"bulk_excel_sign_{login_user.id if login_user else 0}"
    chef_title = request.form.get("chef_title", "").strip()
    chef_name = request.form.get("chef_name", "").strip()
    manager_title = request.form.get("manager_title", "").strip()
    manager_name = request.form.get("manager_name", "").strip()
    director_title = request.form.get("director_title", "").strip()
    director_name = request.form.get("director_name", "").strip()
    set_setting_value(f"{sig_prefix}_chef_title", chef_title)
    set_setting_value(f"{sig_prefix}_chef_name", chef_name)
    set_setting_value(f"{sig_prefix}_manager_title", manager_title)
    set_setting_value(f"{sig_prefix}_manager_name", manager_name)
    set_setting_value(f"{sig_prefix}_director_title", director_title)
    set_setting_value(f"{sig_prefix}_director_name", director_name)
    db.session.commit()

    template_candidates = [
        os.path.join(os.path.dirname(__file__), "Toplu_Mesai_Sablon.xlsx"),
        os.path.join(os.path.dirname(__file__), "fazla_mesai_cizelge.xlsx"),
        os.path.join(os.path.dirname(__file__), "sablon.xlsx"),
        os.path.join(os.path.dirname(__file__), "..", "app", "src", "main", "assets", "sablon.xlsx"),
    ]
    template_path = next((p for p in template_candidates if os.path.exists(p)), "")
    if not template_path:
        flash("Toplu rapor şablonu bulunamadı (Toplu_Mesai_Sablon.xlsx).", "error")
        return redirect(url_for("admin_users"))
    wb = load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]

    month_names_upper = ["OCAK", "ŞUBAT", "MART", "NİSAN", "MAYIS", "HAZİRAN", "TEMMUZ", "AĞUSTOS", "EYLÜL", "EKİM", "KASIM", "ARALIK"]
    first_month_upper = month_names_upper[sm - 1]
    second_month_upper = month_names_upper[p_end.month - 1]
    period_year_value = p_end.year

    # G..AK (7..37) kolonları: 24..31 + 1..23
    day_numbers_in_sheet = list(range(24, 32)) + list(range(1, 24))
    day_col_map = {}
    first_month_visible_days = set()
    cur_y, cur_m = sy, sm
    prev_day_num = None
    for idx, day_num in enumerate(day_numbers_in_sheet):
        col = 7 + idx
        if prev_day_num is not None and day_num < prev_day_num:
            cur_y, cur_m = add_month(cur_y, cur_m)
        try:
            day_col_map[date(cur_y, cur_m, day_num).isoformat()] = col
            if cur_y == sy and cur_m == sm and 24 <= day_num <= 31:
                first_month_visible_days.add(day_num)
        except Exception:
            pass
        prev_day_num = day_num

    # Sadece L, M, N (29, 30, 31) icin kosullu gizleme yap.
    # Kritik: Genisligi bozmamak icin gorunur sutunlara hic dokunma.
    # Yalnizca gizlenecekse hidden=True uygula.
    for day_num in (29, 30, 31):
        if day_num in first_month_visible_days:
            continue
        col = 7 + (day_num - 24)  # 29->L, 30->M, 31->N
        letter = get_column_letter(col)
        ws.column_dimensions[letter].hidden = True

    export_rows = []
    for u in users:
        p = profiles.get(u.id) or UserProfile(user_id=u.id)
        if not is_founder_user(login_user):
            entries_raw = (
                OvertimeEntry.query.filter(
                    OvertimeEntry.user_id == u.id,
                    OvertimeEntry.work_date >= p_start,
                    OvertimeEntry.work_date <= p_end,
                )
                .order_by(OvertimeEntry.work_date.asc(), OvertimeEntry.start_time.asc(), OvertimeEntry.id.asc())
                .all()
            )
            scoped_entries = filter_entries_with_scope(
                entries_raw,
                u.id,
                p,
                unit_changes_map.get(u.id),
                (delegate_perm.scope_daire_baskanligi if delegate_perm else ""),
                (delegate_perm.scope_sube_mudurlugu if delegate_perm else ""),
            )
            rows = grouped_period_rows(scoped_entries)
        else:
            _, _, rows = report_period_rows_for_export(u.id, sy, sm)
        total60 = sum(float(r.get("pct60", 0) or 0) for r in rows)
        total15 = sum(float(r.get("pct15", 0) or 0) for r in rows)
        total_pazar = sum(float(r.get("pazar", 0) or 0) for r in rows)
        total_bayram = sum(float(r.get("bayram", 0) or 0) for r in rows)
        if abs(total60) < 1e-9 and abs(total15) < 1e-9 and abs(total_pazar) < 1e-9 and abs(total_bayram) < 1e-9:
            continue
        export_rows.append(
            {
                "user": u,
                "profile": p,
                "rows": rows,
                "total60": total60,
                "total15": total15,
                "total_pazar": total_pazar,
                "total_bayram": total_bayram,
                "name_sort": (p.ad_soyad or u.email or "").strip().lower(),
            }
        )

    export_rows.sort(key=lambda x: x["name_sort"])
    if not export_rows:
        flash("Seçtiğiniz kişilerde bu dönem için mesai kaydı bulunamadı.", "error")
        return redirect(url_for("admin_users"))

    base_row = 8
    max_row_for_people = 206
    row_step = 2
    person_capacity = ((max_row_for_people - base_row) // row_step) + 1
    for slot_idx in range(person_capacity):
        clear_bulk_mesai_template_person_slot(ws, base_row + (slot_idx * row_step))
    if len(export_rows) > person_capacity:
        flash(f"Şablon en fazla {person_capacity} personel destekliyor.", "error")
        return redirect(url_for("admin_users"))

    grand_60 = 0.0
    grand_15 = 0.0
    grand_pazar = 0.0
    grand_bayram = 0.0
    holiday_day_isos = set()

    for idx, item in enumerate(export_rows):
        row60 = base_row + (idx * row_step)
        row15 = row60 + 1
        p = item["profile"]
        u = item["user"]
        rows_by_day = {r["work_date"].isoformat(): r for r in item["rows"]}

        ws.cell(row=row60, column=3).value = p.sicil_no or ""  # C
        ws.cell(row=row60, column=4).value = p.ad_soyad or u.email  # D

        for day_iso, col in day_col_map.items():
            r = rows_by_day.get(day_iso)
            if not r:
                continue
            v60 = float(r.get("pct60", 0) or 0)
            v15 = float(r.get("pct15", 0) or 0)
            vp = float(r.get("pazar", 0) or 0)
            vb = float(r.get("bayram", 0) or 0)
            # Gunluk tabloda:
            # - Sadece pazar/bayram: 1
            # - Pazar/bayram + ek saat: 1+ekSaat
            # - Sadece %60: saat
            day_marker = vp + vb
            if day_marker > 0 and v60 > 0:
                ws.cell(row=row60, column=col).value = f"{fmt_num(day_marker)}+{fmt_num(v60)}"
            elif day_marker > 0:
                ws.cell(row=row60, column=col).value = day_marker
            else:
                ws.cell(row=row60, column=col).value = v60 if abs(v60) > 1e-9 else None
            ws.cell(row=row15, column=col).value = v15 if abs(v15) > 1e-9 else None
            if abs(vb) > 1e-9:
                holiday_day_isos.add(day_iso)

        total60 = float(item["total60"])
        total15 = float(item["total15"])
        total_pazar = float(item["total_pazar"])
        total_bayram = float(item["total_bayram"])
        ws.cell(row=row60, column=38).value = total60 if abs(total60) > 1e-9 else None  # AL
        ws.cell(row=row15, column=38).value = total15 if abs(total15) > 1e-9 else None  # AL
        ws.cell(row=row60, column=39).value = total_pazar if abs(total_pazar) > 1e-9 else None  # AM
        ws.cell(row=row60, column=40).value = total_bayram if abs(total_bayram) > 1e-9 else None  # AN

        grand_60 += total60
        grand_15 += total15
        grand_pazar += total_pazar
        grand_bayram += total_bayram

    def set_cell_value_safe(cell_ref: str, value):
        cell = ws[cell_ref]
        if not isinstance(cell, MergedCell):
            cell.value = value
            return
        for merged_range in ws.merged_cells.ranges:
            if cell_ref in merged_range:
                ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
                return
        # Beklenmeyen durumda yine de deneyelim
        ws[cell_ref].value = value

    first_profile = export_rows[0]["profile"]
    first_user = export_rows[0]["user"]
    first_unit = unit_at_date_for_user(first_user.id, p_end, profile=first_profile)
    set_cell_value_safe("B2", tr_upper((first_unit.get("sube_mudurlugu") or first_profile.sube_mudurlugu or "")))
    set_cell_value_safe("G5", first_month_upper)
    set_cell_value_safe("O5", second_month_upper)

    people_count = len(export_rows)
    set_cell_value_safe("D209", (
        f"Yukarıda adı soyadı yazılı {people_count} işçi, {period_year_value} yılı {first_month_upper} ve {second_month_upper} ayında toplam "
        f"{fmt_num(grand_60)} saat %60'lık, {fmt_num(grand_15)} saat %15'lik, {fmt_num(grand_pazar)} gün PAZAR, "
        f"{fmt_num(grand_bayram)} gün BAYRAM, olarak fazla çalışma yapmıştır."
    ))
    set_cell_value_safe("D213", chef_title or "")
    set_cell_value_safe("D214", chef_name or "")
    set_cell_value_safe("Q213", manager_title or "")
    set_cell_value_safe("Q214", manager_name or "")
    set_cell_value_safe("AC213", director_title or "")
    set_cell_value_safe("AC214", director_name or "")

    last_used_row60 = base_row + ((people_count - 1) * row_step)
    for r in range(last_used_row60 + 2, 208):
        ws.row_dimensions[r].hidden = True

    weekend_fill = PatternFill(fill_type="solid", start_color="FFD9D9D9", end_color="FFD9D9D9")
    for day_iso, col in day_col_map.items():
        try:
            d = parse_date(day_iso)
        except Exception:
            continue
        is_weekend = d.weekday() in (5, 6)
        is_official_holiday = day_iso in holiday_day_isos
        if is_weekend or is_official_holiday:
            for r in range(6, 208):
                ws.cell(row=r, column=col).fill = weekend_fill

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    return send_file(
        mem,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Toplu_Mesai_{year}_{sm:02d}_{format_dmy(p_start)}_{format_dmy(p_end)}.xlsx",
    )


@app.route("/admin/users/import-excel", methods=["GET", "POST"])
@login_required
@admin_required
def admin_import_period_excel():
    # Içe aktar ekraninda 2025-2026 tum donemler sabit listelensin.
    years = [2026, 2025]
    selected_year = request.args.get("year", type=int) or 2026
    if selected_year not in years:
        selected_year = 2026
    period_options = [(selected_year, m) for m in range(1, 13)]
    active_start = (selected_year, 1)
    if request.method == "GET":
        users = User.query.order_by(User.email.asc()).all()
        profiles = {p.user_id: p for p in UserProfile.query.filter(UserProfile.user_id.in_([u.id for u in users] or [0])).all()}
        rows = [{"user": u, "profile": profiles.get(u.id) or UserProfile(user_id=u.id)} for u in users]
        return render_template(
            "admin_import_excel.html",
            years=years,
            selected_year=selected_year,
            period_options=period_options,
            period_value=f"{active_start[0]:04d}-{active_start[1]:02d}",
            rows=rows,
        )

    upload = request.files.get("excel_file")
    year = request.form.get("year", type=int)
    period = (request.form.get("period") or "").strip()
    selected_ids = [int(v) for v in request.form.getlist("selected_user_ids") if str(v).isdigit()]
    import_all_users = (request.form.get("import_all_users") or "").strip() == "1"
    selected_id_set = set(selected_ids)
    if not upload or not upload.filename:
        flash("Lütfen bir Excel dosyası seçin.", "error")
        return redirect(url_for("admin_import_period_excel"))
    if not period:
        flash("Dönem seçimi eksik.", "error")
        return redirect(url_for("admin_import_period_excel"))
    try:
        if not year:
            flash("Yıl seçimi eksik.", "error")
            return redirect(url_for("admin_import_period_excel"))
        # Içe aktarmada secilen yil ana yil olarak baz alinir.
        # Period degeri yeni ekranda sadece ay (01..12) olarak gelir.
        # Geri uyumluluk icin YYYY-MM formatini da kabul et.
        if "-" in period:
            sm = int(period.split("-")[-1])
        else:
            sm = int(period)
        if sm < 1 or sm > 12:
            raise ValueError("invalid month")
        sy = resolve_period_start_year(int(year), int(sm))
    except Exception:
        flash("Dönem formatı hatalı.", "error")
        return redirect(url_for("admin_import_period_excel"))

    try:
        wb = load_workbook(upload, data_only=True)
    except Exception:
        flash("Excel dosyası okunamadı.", "error")
        return redirect(url_for("admin_import_period_excel"))
    ws = wb[wb.sheetnames[0]]

    # Secilen donemin gun sayisina gore G'den baslayip dinamik kolon ilerler.
    p_start, p_end = period_for_start(sy, sm)
    period_days = []
    cur = p_start
    while cur <= p_end:
        period_days.append(cur)
        cur = cur + timedelta(days=1)
    day_col_map = {7 + idx: d for idx, d in enumerate(period_days)}  # G + idx

    users = User.query.order_by(User.email.asc()).all()
    profiles = {p.user_id: p for p in UserProfile.query.all()}
    if not import_all_users and not selected_id_set:
        flash("İçe aktarma için en az bir personel seçin veya 'Tümünü seç' kutusunu işaretleyin.", "error")
        return redirect(url_for("admin_import_period_excel", year=year))

    def norm_sicil(v):
        t = str(v or "").strip()
        if not t:
            return ""
        if t.endswith(".0"):
            try:
                return str(int(float(t)))
            except Exception:
                return t
        return t

    user_by_sicil = {}
    for u in users:
        p = profiles.get(u.id)
        s = norm_sicil(p.sicil_no if p else "")
        if s:
            user_by_sicil[s] = u

    def parse_excel_cell(v):
        if v is None:
            return None, None
        raw = str(v).strip().upper().replace(",", ".").replace(" ", "")
        if not raw:
            return None, None
        if raw.startswith("P"):
            extra = 0.0
            if "+" in raw:
                try:
                    extra = max(0.0, float(raw.split("+", 1)[1] or "0"))
                except Exception:
                    extra = 0.0
            return "P", extra
        if raw.startswith("B"):
            extra = 0.0
            if "+" in raw:
                try:
                    extra = max(0.0, float(raw.split("+", 1)[1] or "0"))
                except Exception:
                    extra = 0.0
            return "B", extra
        if "+" in raw:
            left, right = (s.strip() for s in raw.split("+", 1))
            try:
                base_val = float(left)
                extra = max(0.0, float(right or "0"))
                if base_val in (0.0, 0.5, 1.0):
                    return "MPLUS", (base_val, extra)
            except Exception:
                pass
        try:
            return "N", float(raw)
        except Exception:
            return None, None

    def add_hours(hhmm: str, hours: float) -> str:
        base = hhmm_to_minutes(hhmm)
        if base is None:
            return hhmm
        mins = int(round(float(hours or 0.0) * 60))
        # Geceyi asan mesailerde saat ertesi gun 00:00/01:00... olarak devam eder.
        end_minutes = (base + mins) % (24 * 60)
        hh = end_minutes // 60
        mm = end_minutes % 60
        return f"{hh:02d}:{mm:02d}"

    matched_user_ids = set()
    rows_added = 0
    skipped_rows = 0
    duplicate_skipped = 0
    to_insert = []
    seen_keys = set()
    seen_day_state = {}
    for r in range(4, 2000):
        sicil = norm_sicil(ws.cell(row=r, column=3).value)  # C
        if not sicil:
            continue
        u = user_by_sicil.get(sicil)
        if not u:
            skipped_rows += 1
            continue
        if (not import_all_users) and (u.id not in selected_id_set):
            continue
        matched_user_ids.add(u.id)
        for col, work_d in day_col_map.items():
            kind, value = parse_excel_cell(ws.cell(row=r, column=col).value)
            if kind is None:
                continue

            defaults = day_defaults(work_d)
            start = str(defaults.get("start") or "08:00")
            end = str(defaults.get("end") or "17:00")
            pct60 = 0.0
            pct15 = 0.0
            pazar = 0.0
            bayram = 0.0

            if kind == "P":
                # P = tam pazar gunu; P+X = 17:00 sonrasi X saat %60, %15 yalnizca ek dilimden.
                extra_hours = max(0.0, float(value or 0.0))
                base_end = str(defaults.get("end") or "17:00")
                end = add_hours_hhmm(base_end, extra_hours) if extra_hours > 0 else base_end
                pct60 = extra_hours
                pct15 = float(calc_night_20_06(base_end, end) or 0.0) if extra_hours > 0 else 0.0
                pazar = 1.0
                bayram = 0.0
            elif kind == "B":
                # B = tam/yarim bayram; B+X = standart bitisten sonra X saat %60.
                extra_hours = max(0.0, float(value or 0.0))
                is_half_holiday = work_d in half_holiday_set(work_d.year)
                if is_half_holiday:
                    start = "13:00"
                    base_end = "17:00"
                    bayram = 0.5
                else:
                    base_end = str(defaults.get("end") or "17:00")
                    bayram = 1.0
                end = add_hours_hhmm(base_end, extra_hours) if extra_hours > 0 else base_end
                pct60 = extra_hours
                pct15 = float(calc_night_20_06(base_end, end) or 0.0) if extra_hours > 0 else 0.0
                pazar = 0.0
            elif kind == "MPLUS":
                # Excel disa aktarim: 1+5 = tam pazar/bayram + 17:00 sonrasi 5 saat %60.
                base_val, extra_hours = value
                extra_hours = max(0.0, float(extra_hours or 0.0))
                base_end = str(defaults.get("end") or "17:00")
                end = add_hours_hhmm(base_end, extra_hours) if extra_hours > 0 else base_end
                pct60 = extra_hours
                pct15 = float(calc_night_20_06(base_end, end) or 0.0) if extra_hours > 0 else 0.0
                if float(base_val) == 0.0:
                    pazar = 0.0
                    bayram = 0.0
                elif bool(defaults.get("isHoliday")):
                    bayram = float(base_val)
                    pazar = 0.0
                else:
                    pazar = float(base_val)
                    bayram = 0.0
            else:
                # Excel hucredeki sayi ogle dusulmus net %60 saatidir.
                num = max(0.0, float(value or 0.0))
                if num <= 0:
                    continue
                wd = work_d.weekday()
                is_special_day = bool(defaults.get("isHoliday")) or wd == 6
                is_sat_no_holiday = wd == 5 and not bool(defaults.get("isHoliday"))
                if is_special_day:
                    # P/B yoksa yalnizca %60 mesai; pazar/bayram gunu isareti yazilmaz.
                    hk_cell = holiday_kind_tr(work_d)
                    end = end_hhmm_for_bulk_special_target_pct60(start, num, wd, hk_cell)
                    sp_b = compute_mesai_split(start, end, wd, hk_cell)
                    pct60 = float(num)
                    pct15 = float(sp_b.get("pct15", 0) or 0)
                    pazar = 0.0
                    bayram = 0.0
                elif is_sat_no_holiday:
                    end = end_hhmm_for_saturday_net(start, num)
                    calc = day_defaults(work_d, end, start)
                    pct60 = float(calc.get("pct60", 0) or 0)
                    pct15 = float(calc.get("pct15", 0) or 0)
                    pazar = float(calc.get("pazar", 0) or 0)
                    bayram = float(calc.get("bayram", 0) or 0)
                else:
                    end = add_hours_hhmm(start, num)
                    calc = day_defaults(work_d, end, start)
                    pct60 = float(calc.get("pct60", 0) or 0)
                    pct15 = float(calc.get("pct15", 0) or 0)
                    pazar = float(calc.get("pazar", 0) or 0)
                    bayram = float(calc.get("bayram", 0) or 0)

            dup_key = (u.id, work_d.isoformat(), start, end)
            if dup_key in seen_keys:
                duplicate_skipped += 1
                continue
            # Ayni kullanici + ayni gun icin:
            # - Ayni %60 tekrar ise atla
            # - Pazar/Bayram zaten yazildiysa tekrarini atla
            day_key = (u.id, work_d.isoformat())
            state = seen_day_state.setdefault(day_key, {"pct60_values": set(), "has_pazar": False, "has_bayram": False})
            pct60_key = round(float(pct60 or 0.0), 4)
            if pct60_key > 0 and pct60_key in state["pct60_values"]:
                duplicate_skipped += 1
                continue
            if pazar > 0 and state["has_pazar"]:
                duplicate_skipped += 1
                continue
            if bayram > 0 and state["has_bayram"]:
                duplicate_skipped += 1
                continue

            seen_keys.add(dup_key)
            if pct60_key > 0:
                state["pct60_values"].add(pct60_key)
            if pazar > 0:
                state["has_pazar"] = True
            if bayram > 0:
                state["has_bayram"] = True

            to_insert.append(
                OvertimeEntry(
                    user_id=u.id,
                    work_date=work_d,
                    start_time=start,
                    end_time=end,
                    pct60=pct60,
                    pct15=pct15,
                    pazar=pazar,
                    bayram=bayram,
                    description="",
                )
            )
            rows_added += 1

    if not matched_user_ids:
        flash("Excelde seçilen personeller için eşleşen sicil numarası bulunamadı.", "error")
        return redirect(url_for("admin_import_period_excel"))

    p_start, p_end = period_for_start(sy, sm)
    old_period_rows = OvertimeEntry.query.filter(
        OvertimeEntry.user_id.in_(list(matched_user_ids)),
        OvertimeEntry.work_date >= p_start,
        OvertimeEntry.work_date <= p_end,
    ).all()
    actor_user = session_login_user()
    for old_row in old_period_rows:
        write_overtime_audit_log(
            action="delete",
            actor_user_id=(actor_user.id if actor_user else 0),
            target_user_id=old_row.user_id,
            old_entry=old_row,
            new_entry=None,
            source="import",
            note="excel_import_replace_period",
        )
        db.session.delete(old_row)
    if to_insert:
        db.session.add_all(to_insert)
        db.session.flush()
        for row in to_insert:
            write_overtime_audit_log(
                action="create",
                actor_user_id=(actor_user.id if actor_user else 0),
                target_user_id=row.user_id,
                old_entry=None,
                new_entry=row,
                source="import",
                note="excel_import_add",
            )
    db.session.commit()
    flash(
        f"İçe aktarma tamamlandı. {len(matched_user_ids)} kullanıcı için {rows_added} kayıt işlendi."
        + (f" {skipped_rows} satırda sicil eşleşmedi." if skipped_rows else "")
        + (f" {duplicate_skipped} mükerrer kayıt atlandı." if duplicate_skipped else ""),
        "success",
    )
    flash(f"İşlenen dönem: {format_dmy(p_start)} - {format_dmy(p_end)}", "success")
    return redirect(url_for("admin_users"))


@app.post("/admin/users/delete-period-all")
@login_required
@admin_required
def admin_delete_period_all():
    selected_ids = [int(v) for v in request.form.getlist("selected_user_ids") if str(v).isdigit()]
    if not selected_ids:
        flash("Veri silme için üst listeden en az bir personel seçin.", "error")
        return redirect(url_for("admin_users"))
    period = (request.form.get("period") or "").strip()
    if "-" not in period:
        flash("Dönem seçimi eksik.", "error")
        return redirect(url_for("admin_users"))
    try:
        sy, sm = (int(x) for x in period.split("-"))
    except Exception:
        flash("Dönem formatı hatalı.", "error")
        return redirect(url_for("admin_users"))
    p_start, p_end = period_for_start(sy, sm)
    actor_user = session_login_user()
    rows_to_delete = OvertimeEntry.query.filter(
        OvertimeEntry.user_id.in_(selected_ids),
        OvertimeEntry.work_date >= p_start,
        OvertimeEntry.work_date <= p_end,
    ).all()
    for row in rows_to_delete:
        write_overtime_audit_log(
            action="delete",
            actor_user_id=(actor_user.id if actor_user else 0),
            target_user_id=row.user_id,
            old_entry=row,
            new_entry=None,
            source="admin",
            note="period_delete_selected_users",
        )
        db.session.delete(row)
    deleted = len(rows_to_delete)
    db.session.commit()
    flash(f"Seçilen personeller için seçilen dönemde toplam {deleted} kayıt silindi.", "success")
    return redirect(url_for("admin_users"))


@app.route("/settings", methods=["GET", "POST"])
@login_required
def settings_page():
    user = ensure_user_or_redirect()
    if user is None:
        flash("Oturum süresi doldu, lütfen tekrar giriş yapın.", "error")
        return redirect(url_for("login"))

    if request.method == "POST":
        action = request.form.get("action", "").strip()
        try:
            if action == "change_password":
                old_password = request.form.get("old_password", "")
                new_password = request.form.get("new_password", "")
                new_password_confirm = request.form.get("new_password_confirm", "")
                if not check_password_hash(user.password_hash, old_password):
                    flash("Eski şifre hatalı.", "error")
                    return redirect(url_for("settings_page"))
                if len(new_password) < 6:
                    flash("Yeni şifre en az 6 karakter olmalı.", "error")
                    return redirect(url_for("settings_page"))
                if new_password != new_password_confirm:
                    flash("Yeni şifre tekrar alanı uyuşmuyor.", "error")
                    return redirect(url_for("settings_page"))
                user.password_hash = generate_password_hash(new_password)
                db.session.commit()
                flash("Şifre başarıyla değiştirildi.", "success")
            elif action == "apk_refresh":
                OvertimeEntry.query.filter_by(user_id=user.id).delete()
                p = get_or_create_profile(user.id)
                p.daire_baskanligi = ""
                p.sube_mudurlugu = ""
                p.ad_soyad = ""
                p.sicil_no = ""
                p.ekip_kodu = ""
                db.session.commit()
                flash("Web verileri temizlendi. APK'de giriş yapıp senkron yaptığınızda veriler yeniden yüklenecek.", "success")
            elif action == "clear_all":
                OvertimeEntry.query.filter_by(user_id=user.id).delete()
                p = get_or_create_profile(user.id)
                p.daire_baskanligi = ""
                p.sube_mudurlugu = ""
                p.ad_soyad = ""
                p.sicil_no = ""
                p.ekip_kodu = ""
                db.session.commit()
                flash("Web tarafındaki tüm veriler silindi.", "success")
            else:
                flash("Geçersiz işlem.", "error")
        except Exception as exc:
            db.session.rollback()
            flash(f"Ayar işlemi başarısız: {exc}", "error")
        return redirect(url_for("settings_page"))

    all_entries = OvertimeEntry.query.filter_by(user_id=user.id).order_by(OvertimeEntry.work_date.desc(), OvertimeEntry.id.desc()).all()
    start_options = sorted({(period_start_for_date(e.work_date).year, period_start_for_date(e.work_date).month) for e in all_entries}, reverse=True)
    if not start_options:
        ps = period_start_for_date(date.today())
        start_options = [(ps.year, ps.month)]
    selected_year = period_year(start_options[0][0], start_options[0][1])
    active_start = start_options[0]
    period_value = f"{active_start[0]:04d}-{active_start[1]:02d}"

    return render_template(
        "settings.html",
        selected_year=selected_year,
        period_value=period_value,
    )


@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    user = ensure_user_or_redirect()
    if user is None:
        flash("Oturum süresi doldu, lütfen tekrar giriş yapın.", "error")
        return redirect(url_for("login"))
    p = get_or_create_profile(user.id)
    if request.method == "POST":
        p.daire_baskanligi = request.form.get("daire_baskanligi", "").strip()
        p.sube_mudurlugu = request.form.get("sube_mudurlugu", "").strip()
        p.ad_soyad = request.form.get("ad_soyad", "").strip()
        p.sicil_no = request.form.get("sicil_no", "").strip()
        p.ekip_kodu = request.form.get("ekip_kodu", "").strip()
        db.session.commit()
        flash("Profil bilgileri güncellendi.", "success")
        return redirect(url_for("profile"))
    return render_template("profile.html", user=user, profile=p)


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    reset_url = None
    sent_via_smtp = False
    if request.method == "POST":
        ip = request.headers.get("X-Forwarded-For", request.remote_addr or "unknown")
        if is_rate_limited(f"forgot:{ip}", limit=8, window_sec=60):
            flash("Çok fazla deneme. Lütfen 1 dakika sonra tekrar deneyin.", "error")
            return render_template("forgot_password.html", reset_url=None, sent_via_smtp=False)
        email = request.form.get("email", "").strip()
        user = User.query.filter_by(email=email).first()
        if user:
            token = token_serializer.dumps({"uid": user.id}, salt="reset-password")
            reset_url = f"{app.config['SITE_BASE_URL']}{url_for('reset_password', token=token)}"
            try:
                sent_via_smtp = send_reset_email(user.email, reset_url)
            except Exception:
                sent_via_smtp = False
        flash("E-posta kayıtlıysa şifre sıfırlama bağlantısı oluşturuldu.", "success")
    return render_template("forgot_password.html", reset_url=reset_url if not sent_via_smtp else None, sent_via_smtp=sent_via_smtp)


@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password(token):
    max_age = app.config["RESET_TOKEN_EXPIRE_MIN"] * 60
    try:
        data = token_serializer.loads(token, salt="reset-password", max_age=max_age)
    except SignatureExpired:
        flash("Bağlantının süresi doldu.", "error")
        return redirect(url_for("forgot_password"))
    except BadSignature:
        flash("Geçersiz bağlantı.", "error")
        return redirect(url_for("forgot_password"))
    user = User.query.get(data["uid"])
    if not user:
        flash("Kullanıcı bulunamadı.", "error")
        return redirect(url_for("forgot_password"))
    if request.method == "POST":
        password = request.form.get("password", "")
        confirm = request.form.get("confirm_password", "")
        if len(password) < 6:
            flash("Şifre en az 6 karakter olmalı.", "error")
            return render_template("reset_password.html")
        if password != confirm:
            flash("Şifreler eşleşmiyor.", "error")
            return render_template("reset_password.html")
        user.password_hash = generate_password_hash(password)
        db.session.commit()
        flash("Şifre güncellendi. Giriş yapabilirsiniz.", "success")
        return redirect(url_for("login"))
    return render_template("reset_password.html")


@app.get("/api/day-defaults")
@login_required
def api_day_defaults_web():
    ymd = request.args.get("date", "")
    end_override = (request.args.get("endTime") or "").strip()
    start_override = (request.args.get("startTime") or "").strip()
    try:
        d = parse_date(ymd)
    except Exception:
        return jsonify({"error": "invalid_date"}), 400
    defaults = day_defaults(d, end_override or None, start_override or None)
    return jsonify(defaults)


@app.route("/dashboard", methods=["GET", "POST"])
@login_required
def dashboard():
    user = ensure_user_or_redirect()
    login_user = session_login_user()
    if user is None:
        flash("Oturum süresi doldu, lütfen tekrar giriş yapın.", "error")
        return redirect(url_for("login"))
    profile = get_or_create_profile(user.id)
    if request.method == "POST":
        try:
            entry = OvertimeEntry(
                user_id=user.id,
                work_date=parse_date(request.form.get("work_date", "")),
                start_time=request.form.get("start_time", "").strip(),
                end_time=request.form.get("end_time", "").strip(),
                pct60=parse_float(request.form.get("pct60", "0")),
                pct15=parse_float(request.form.get("pct15", "0")),
                pazar=parse_float(request.form.get("pazar", "0")),
                bayram=parse_float(request.form.get("bayram", "0")),
                description=request.form.get("description", "").strip(),
            )
            dup = OvertimeEntry.query.filter_by(
                user_id=user.id,
                work_date=entry.work_date,
                start_time=entry.start_time,
                end_time=entry.end_time,
            ).first()
            if dup:
                flash("Aynı gün ve saat için mükerrer mesai girilemez.", "error")
                return redirect(url_for("dashboard"))
            if find_overlapping_overtime_for_user(
                user.id, entry.work_date, entry.start_time, entry.end_time
            ):
                flash(MESAI_TIME_OVERLAP_MESSAGE, "overlap")
                return redirect(url_for("dashboard"))
            if is_period_locked(entry.work_date) and not can_bypass_period_lock(login_user):
                flash("Bu dönem kilitli. Mesai girişi yapılamaz.", "error")
                return redirect(url_for("dashboard"))
            db.session.add(entry)
            write_overtime_audit_log(
                action="create",
                actor_user_id=(login_user.id if login_user else user.id),
                target_user_id=user.id,
                old_entry=None,
                new_entry=entry,
                source="web",
                note="dashboard_add",
            )
            db.session.commit()
            flash("Mesai kaydı eklendi.", "success")
        except Exception as exc:
            db.session.rollback()
            flash(f"Kayıt eklenemedi: {exc}", "error")
        return redirect(url_for("dashboard"))
    entries = OvertimeEntry.query.filter_by(user_id=user.id).all()
    recent_items = build_recent_ui_items(entries)
    return render_template(
        "dashboard.html",
        user=user,
        profile=profile,
        recent_items=recent_items,
        apk_url=app.config["APK_URL"],
        api_token=session.get("api_token"),
    )


@app.route("/entries/<int:entry_id>/edit", methods=["GET", "POST"])
@login_required
def edit_entry(entry_id: int):
    user = ensure_user_or_redirect()
    login_user = session_login_user()
    if user is None:
        flash("Oturum süresi doldu, lütfen tekrar giriş yapın.", "error")
        return redirect(url_for("login"))
    entry = OvertimeEntry.query.filter_by(id=entry_id, user_id=user.id).first_or_404()
    if request.method == "POST":
        try:
            old_snapshot = OvertimeEntry(
                id=entry.id,
                user_id=entry.user_id,
                work_date=entry.work_date,
                start_time=entry.start_time,
                end_time=entry.end_time,
                pct60=entry.pct60,
                pct15=entry.pct15,
                pazar=entry.pazar,
                bayram=entry.bayram,
                description=entry.description,
            )
            new_work_date = parse_date(request.form.get("work_date", ""))
            new_start = request.form.get("start_time", "").strip()
            new_end = request.form.get("end_time", "").strip()
            if (
                (is_period_locked(entry.work_date) or is_period_locked(new_work_date))
                and not can_bypass_period_lock(login_user)
            ):
                flash("Bu dönem kilitli. Güncelleme yapılamaz.", "error")
                return redirect(url_for("dashboard"))
            if find_overlapping_overtime_for_user(
                user.id, new_work_date, new_start, new_end, exclude_entry_id=entry.id
            ):
                flash(MESAI_TIME_OVERLAP_MESSAGE, "overlap")
                return redirect(
                    url_for("edit_entry", entry_id=entry_id, back=request.form.get("back", "dashboard"))
                )
            entry.work_date = new_work_date
            entry.start_time = new_start
            entry.end_time = new_end
            entry.pct60 = parse_float(request.form.get("pct60", "0"))
            entry.pct15 = parse_float(request.form.get("pct15", "0"))
            entry.pazar = parse_float(request.form.get("pazar", "0"))
            entry.bayram = parse_float(request.form.get("bayram", "0"))
            entry.description = request.form.get("description", "").strip()

            # Aynı kullanıcıda aynı tarih+saat mükerrer kayıtlar varsa tek kayda indir.
            # Eski senkronlardan kalan kopyalar bu şekilde temizlenir.
            duplicates = OvertimeEntry.query.filter(
                OvertimeEntry.user_id == user.id,
                OvertimeEntry.work_date == entry.work_date,
                OvertimeEntry.start_time == entry.start_time,
                OvertimeEntry.end_time == entry.end_time,
                OvertimeEntry.id != entry.id,
            ).all()
            for d in duplicates:
                write_overtime_audit_log(
                    action="delete",
                    actor_user_id=(login_user.id if login_user else user.id),
                    target_user_id=user.id,
                    old_entry=d,
                    new_entry=None,
                    source="web",
                    note="duplicate_cleanup",
                )
                db.session.delete(d)

            write_overtime_audit_log(
                action="update",
                actor_user_id=(login_user.id if login_user else user.id),
                target_user_id=user.id,
                old_entry=old_snapshot,
                new_entry=entry,
                source="web",
                note="entry_edit",
            )
            db.session.commit()
            if duplicates:
                flash(f"Kayıt güncellendi. {len(duplicates)} mükerrer kayıt temizlendi.", "success")
            else:
                flash("Kayıt güncellendi.", "success")
            back = request.form.get("back", "dashboard")
            return redirect(url_for("reports") if back == "reports" else url_for("dashboard"))
        except Exception as exc:
            db.session.rollback()
            flash(f"Güncelleme başarısız: {exc}", "error")
    return render_template("entry_edit.html", entry=entry, back=request.args.get("back", "dashboard"))


@app.post("/entries/<int:entry_id>/delete")
@login_required
def delete_entry(entry_id: int):
    user = ensure_user_or_redirect()
    login_user = session_login_user()
    if user is None:
        flash("Oturum süresi doldu, lütfen tekrar giriş yapın.", "error")
        return redirect(url_for("login"))
    entry = OvertimeEntry.query.filter_by(id=entry_id, user_id=user.id).first_or_404()
    if is_period_locked(entry.work_date) and not can_bypass_period_lock(login_user):
        flash("Bu dönem kilitli. Silme işlemi yapılamaz.", "error")
        return redirect(url_for("dashboard"))
    write_overtime_audit_log(
        action="delete",
        actor_user_id=(login_user.id if login_user else user.id),
        target_user_id=user.id,
        old_entry=entry,
        new_entry=None,
        source="web",
        note="entry_delete",
    )
    db.session.delete(entry)
    db.session.commit()
    flash("Kayıt silindi.", "success")
    back = request.form.get("back", "dashboard")
    return redirect(url_for("reports") if back == "reports" else url_for("dashboard"))


@app.get("/reports")
@login_required
def reports():
    user = ensure_user_or_redirect()
    if user is None:
        flash("Oturum süresi doldu, lütfen tekrar giriş yapın.", "error")
        return redirect(url_for("login"))
    profile = get_or_create_profile(user.id)
    all_entries = OvertimeEntry.query.filter_by(user_id=user.id).order_by(OvertimeEntry.work_date.desc(), OvertimeEntry.id.desc()).all()
    start_options = sorted({(period_start_for_date(e.work_date).year, period_start_for_date(e.work_date).month) for e in all_entries}, reverse=True)
    if not start_options:
        ps = period_start_for_date(date.today())
        start_options = [(ps.year, ps.month)]
    years = sorted({period_year(y, m) for (y, m) in start_options}, reverse=True)
    selected_year = request.args.get("year", type=int) or years[0]
    if selected_year not in years:
        selected_year = years[0]
    period_options = [(y, m) for (y, m) in start_options if period_year(y, m) == selected_year] or [start_options[0]]
    selected_period = request.args.get("period", "")
    active_start = period_options[0]
    if selected_period and "-" in selected_period:
        sy, sm = (int(x) for x in selected_period.split("-"))
        if (sy, sm) in period_options:
            active_start = (sy, sm)
    p_start, p_end = period_for_start(active_start[0], active_start[1])
    period_entries_raw = (
        OvertimeEntry.query.filter(
            OvertimeEntry.user_id == user.id,
            OvertimeEntry.work_date >= p_start,
            OvertimeEntry.work_date <= p_end,
        )
        .order_by(OvertimeEntry.work_date.desc(), OvertimeEntry.start_time.desc(), OvertimeEntry.id.desc())
        .all()
    )
    period_total = {
        "pct60": sum(float(e.pct60 or 0) for e in period_entries_raw),
        "pct15": sum(float(e.pct15 or 0) for e in period_entries_raw),
        "pazar": sum(float(e.pazar or 0) for e in period_entries_raw),
        "bayram": sum(float(e.bayram or 0) for e in period_entries_raw),
    }
    yearly_entries = [e for e in all_entries if period_year(period_start_for_date(e.work_date).year, period_start_for_date(e.work_date).month) == selected_year]
    year_rows = grouped_period_rows(yearly_entries)
    year_total = {
        "pct60": sum(e["pct60"] for e in year_rows),
        "pct15": sum(e["pct15"] for e in year_rows),
        "pazar": sum(e["pazar"] for e in year_rows),
        "bayram": sum(e["bayram"] for e in year_rows),
    }
    return render_template(
        "reports.html",
        user=user,
        profile=profile,
        years=years,
        selected_year=selected_year,
        period_options=period_options,
        active_start=active_start,
        rows=period_entries_raw,
        period_start=p_start,
        period_end=p_end,
        period_total=period_total,
        year_total=year_total,
        weekday_tr=weekday_tr,
        format_dmy=format_dmy,
    )


@app.post("/reports/import")
@login_required
def import_reports_backup():
    user = ensure_user_or_redirect()
    login_user = session_login_user()
    if user is None:
        flash("Oturum süresi doldu, lütfen tekrar giriş yapın.", "error")
        return redirect(url_for("login"))
    profile = get_or_create_profile(user.id)
    f = request.files.get("backup_file")
    back = request.form.get("back", "reports")
    redirect_target = "settings_page" if back == "settings" else "reports"
    if f is None or f.filename == "":
        flash("İçe aktarma için dosya seçin.", "error")
        return redirect(url_for(redirect_target))
    try:
        raw = f.read()
        if not raw:
            flash("Dosya boş.", "error")
            return redirect(url_for(redirect_target))
        payload = json.loads(raw.decode("utf-8-sig", errors="strict"))
        if not isinstance(payload, dict):
            raise ValueError("Geçersiz JSON yapısı")
        prof = payload.get("profile", {})
        entries = payload.get("entries", [])
        if isinstance(prof, dict):
            profile.daire_baskanligi = str(prof.get("daireBaskanligi", profile.daire_baskanligi))
            profile.sube_mudurlugu = str(prof.get("subeMudurlugu", profile.sube_mudurlugu))
            profile.ad_soyad = str(prof.get("adSoyad", profile.ad_soyad))
            profile.sicil_no = str(prof.get("sicilNo", profile.sicil_no))
            profile.ekip_kodu = str(prof.get("ekipKodu", profile.ekip_kodu))
        inserted = 0
        for e in entries if isinstance(entries, list) else []:
            if not isinstance(e, dict):
                continue
            work_date = str(e.get("workDate", "")).strip()
            start_time = str(e.get("startTime", "")).strip()
            end_time = str(e.get("endTime", "")).strip()
            if not work_date or not start_time or not end_time:
                continue
            try:
                wd = parse_date(work_date)
            except Exception:
                continue
            if is_period_locked(wd) and not can_bypass_period_lock(login_user):
                continue
            dup = OvertimeEntry.query.filter_by(
                user_id=user.id,
                work_date=wd,
                start_time=start_time,
                end_time=end_time,
            ).first()
            if dup:
                continue
            row = OvertimeEntry(
                user_id=user.id,
                work_date=wd,
                start_time=start_time,
                end_time=end_time,
                pct60=float(e.get("pct60", 0) or 0),
                pct15=float(e.get("pct15", 0) or 0),
                pazar=float(e.get("pazar", 0) or 0),
                bayram=float(e.get("bayram", 0) or 0),
                description=str(e.get("description", "")),
            )
            db.session.add(row)
            inserted += 1
        db.session.commit()
        flash(f"İçe aktarma tamamlandı. Eklenen kayıt: {inserted}", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"İçe aktarma başarısız: {exc}", "error")
    return redirect(url_for(redirect_target))


def report_period_rows_for_export(user_id: int, sy: int, sm: int):
    p_start, p_end = period_for_start(sy, sm)
    entries = (
        OvertimeEntry.query.filter(
            OvertimeEntry.user_id == user_id,
            OvertimeEntry.work_date >= p_start,
            OvertimeEntry.work_date <= p_end,
        )
        .order_by(OvertimeEntry.work_date.asc(), OvertimeEntry.start_time.asc(), OvertimeEntry.id.asc())
        .all()
    )
    return p_start, p_end, grouped_period_rows(entries)


@app.get("/reports/export.csv")
@login_required
def export_reports_csv():
    user = ensure_user_or_redirect()
    if user is None:
        flash("Oturum süresi doldu, lütfen tekrar giriş yapın.", "error")
        return redirect(url_for("login"))
    year = request.args.get("year", type=int)
    period = request.args.get("period", "")
    if not year or "-" not in period:
        flash("Yıl/dönem bilgisi eksik.", "error")
        return redirect(url_for("reports"))
    sy, sm = (int(x) for x in period.split("-"))
    _, _, rows = report_period_rows_for_export(user.id, sy, sm)
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["Tarih", "Gun", "Baslama", "Bitis", "%60", "%15", "Pazar", "Bayram", "Aciklama"])
    for r in rows:
        writer.writerow(
            [
                format_dmy(r["work_date"]),
                weekday_tr(r["work_date"]),
                r["start_time"],
                r["end_time"],
                fmt_num(r["pct60"]),
                fmt_num(r["pct15"]),
                fmt_num(r["pazar"]),
                fmt_num(r["bayram"]),
                r["description"],
            ]
        )
    mem = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    mem.seek(0)
    return send_file(mem, mimetype="text/csv", as_attachment=True, download_name=f"Mesai_{year}_{sm:02d}.csv")


@app.get("/reports/export.xlsx")
@login_required
def export_reports_xlsx():
    user = ensure_user_or_redirect()
    if user is None:
        flash("Oturum süresi doldu, lütfen tekrar giriş yapın.", "error")
        return redirect(url_for("login"))
    profile = get_or_create_profile(user.id)
    year = request.args.get("year", type=int)
    period = request.args.get("period", "")
    if not year or "-" not in period:
        flash("Yıl/dönem bilgisi eksik.", "error")
        return redirect(url_for("reports"))
    sy, sm = (int(x) for x in period.split("-"))
    p_start, p_end, rows = report_period_rows_for_export(user.id, sy, sm)
    period_unit = unit_at_date_for_user(user.id, p_end, profile=profile)
    totals = {
        "pct60": sum(r["pct60"] for r in rows),
        "pct15": sum(r["pct15"] for r in rows),
        "pazar": sum(r["pazar"] for r in rows),
        "bayram": sum(r["bayram"] for r in rows),
    }

    template_candidates = [
        os.path.join(os.path.dirname(__file__), "sablon.xlsx"),
        os.path.join(os.path.dirname(__file__), "..", "app", "src", "main", "assets", "sablon.xlsx"),
    ]
    template_path = next((p for p in template_candidates if os.path.exists(p)), "")
    if not template_path:
        flash("Excel şablonu bulunamadı (sablon.xlsx). Lütfen şablon dosyasını web-portal klasörüne ekleyin.", "error")
        return redirect(url_for("reports", year=year, period=period))

    wb = load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]
    ws["D3"] = period_unit.get("daire_baskanligi") or profile.daire_baskanligi
    ws["D4"] = period_unit.get("sube_mudurlugu") or profile.sube_mudurlugu
    ws["D5"] = profile.ad_soyad
    ws["D6"] = profile.sicil_no
    end_month_name = ["OCAK", "ŞUBAT", "MART", "NİSAN", "MAYIS", "HAZİRAN", "TEMMUZ", "AĞUSTOS", "EYLÜL", "EKİM", "KASIM", "ARALIK"][p_end.month - 1]
    ws["H10"] = end_month_name
    ws["J10"] = p_end.year
    day_map = {r["work_date"].isoformat(): r for r in rows}
    next_y, next_m = add_month(sy, sm)
    for row_num in range(14, 45):
        day_num = 24 + (row_num - 14) if row_num <= 21 else (row_num - 21)
        cur_date = None
        try:
            if row_num <= 21:
                cur_date = date(sy, sm, day_num)
            else:
                cur_date = date(next_y, next_m, day_num)
        except Exception:
            cur_date = None
        data = day_map.get(cur_date.isoformat()) if cur_date else None
        ws[f"B{row_num}"] = data["start_time"] if data else None
        ws[f"C{row_num}"] = data["end_time"] if data else None
        ws[f"D{row_num}"] = data["pct60"] if data and abs(data["pct60"]) > 1e-9 else None
        ws[f"E{row_num}"] = data["pct15"] if data and abs(data["pct15"]) > 1e-9 else None
        ws[f"F{row_num}"] = data["pazar"] if data and abs(data["pazar"]) > 1e-9 else None
        ws[f"G{row_num}"] = data["bayram"] if data and abs(data["bayram"]) > 1e-9 else None
        ws[f"H{row_num}"] = data["description"] if data and data["description"] else None
        has_any = data and (
            data["start_time"] or data["end_time"] or abs(data["pct60"]) > 1e-9 or abs(data["pct15"]) > 1e-9 or abs(data["pazar"]) > 1e-9 or abs(data["bayram"]) > 1e-9 or data["description"]
        )
        ws[f"I{row_num}"] = profile.ekip_kodu if has_any else None
    ws["D45"] = totals["pct60"] if abs(totals["pct60"]) > 1e-9 else None
    ws["E45"] = totals["pct15"] if abs(totals["pct15"]) > 1e-9 else None
    ws["F45"] = totals["pazar"] if abs(totals["pazar"]) > 1e-9 else None
    ws["G45"] = totals["bayram"] if abs(totals["bayram"]) > 1e-9 else None
    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    return send_file(mem, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"Mesai_{year}_{sm:02d}.xlsx")


@app.get("/download-apk")
@login_required
def download_apk():
    user = ensure_user_or_redirect()
    if user is None:
        flash("Oturum süresi doldu, lütfen tekrar giriş yapın.", "error")
        return redirect(url_for("login"))

    # APK uygulamasiyla ayni guncelleme kaynagi: update manifest -> apkUrl.
    manifest_url = (app.config.get("UPDATE_MANIFEST_URL") or "").strip()
    if manifest_url:
        try:
            req = urllib.request.Request(
                manifest_url,
                headers={"User-Agent": "MesaiWebPortal/1.0"},
            )
            with urllib.request.urlopen(req, timeout=12) as resp:
                if resp.status == 200:
                    payload = json.loads(resp.read().decode("utf-8"))
                    apk_from_manifest = str(payload.get("apkUrl", "")).strip()
                    if apk_from_manifest:
                        return redirect(apk_from_manifest)
        except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, ValueError, KeyError):
            # Manifeste erisilemezse asagidaki fallback adimlari calissin.
            pass
        except Exception:
            pass

    # Dis URL tanimliysa (ornegin GitHub release), tek noktadan yonlendir.
    configured_apk_url = (app.config.get("APK_URL") or "").strip()
    if configured_apk_url and configured_apk_url != "/download-apk":
        return redirect(configured_apk_url)

    # Yerelde/depoda bulunan APK dosyalari arasindan en yeni olani indir.
    repo_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    candidate_patterns = [
        os.path.join(repo_root, "app", "build", "outputs", "apk", "release", "*.apk"),
        os.path.join(repo_root, "app", "build", "outputs", "apk", "debug", "*.apk"),
        os.path.join(repo_root, "web-portal", "static", "apk", "*.apk"),
        os.path.join(repo_root, "apk", "*.apk"),
    ]
    candidates = []
    for pattern in candidate_patterns:
        candidates.extend(glob.glob(pattern))
    candidates = [p for p in candidates if os.path.exists(p)]
    if candidates:
        latest_apk = max(candidates, key=lambda p: os.path.getmtime(p))
        return send_file(latest_apk, as_attachment=True, download_name=os.path.basename(latest_apk))

    flash("APK dosyası henüz üretilmemiş.", "error")
    return redirect(url_for("dashboard"))


@app.get("/apk-auto-login")
def apk_auto_login():
    token = request.args.get("token", "").strip()
    if not token:
        return redirect(url_for("login"))
    try:
        data = token_serializer.loads(token, max_age=60 * 60 * 24 * 30)
        uid = data.get("uid")
    except Exception:
        return redirect(url_for("login"))
    user = User.query.get(uid) if uid else None
    if not user:
        session.clear()
        return redirect(url_for("login"))
    session.clear()
    session["user_id"] = user.id
    # web oturumu için yeni token üret (nonce taze kalsın)
    session["api_token"] = token_serializer.dumps({"uid": user.id, "nonce": secrets.token_hex(8)})
    return redirect(url_for("dashboard"))


def bearer_user():
    header = request.headers.get("Authorization", "")
    if not header.startswith("Bearer "):
        return None
    token = header.replace("Bearer ", "", 1).strip()
    try:
        data = token_serializer.loads(token, max_age=60 * 60 * 24 * 30)
    except Exception:
        return None
    return User.query.get(data.get("uid"))


def api_auth_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        user = bearer_user()
        if not user:
            return jsonify({"error": "unauthorized"}), 401
        if is_user_terminated(int(user.id)):
            return jsonify({"error": "terminated_user"}), 403
        request.api_user = user
        return view_func(*args, **kwargs)

    return wrapped


@app.post("/api/login")
def api_login():
    ip = request.headers.get("X-Forwarded-For", request.remote_addr or "unknown")
    if is_rate_limited(f"api_login:{ip}", limit=20, window_sec=60):
        return jsonify({"error": "rate_limited"}), 429
    data = request.get_json(silent=True) or {}
    identity = str(data.get("email", data.get("usernameOrEmail", ""))).strip()
    password = str(data.get("password", ""))
    user = User.query.filter((User.username == identity) | (User.email == identity)).first()
    if not user or not check_password_hash(user.password_hash, password):
        return jsonify({"error": "invalid_credentials"}), 401
    if is_user_terminated(int(user.id)):
        return jsonify({"error": "terminated_user"}), 403
    token = token_serializer.dumps({"uid": user.id, "nonce": secrets.token_hex(8)})
    return jsonify({"token": token, "user": {"id": user.id, "email": user.email}})


@app.post("/api/register")
def api_register():
    ip = request.headers.get("X-Forwarded-For", request.remote_addr or "unknown")
    if is_rate_limited(f"api_register:{ip}", limit=10, window_sec=60):
        return jsonify({"error": "rate_limited"}), 429
    data = request.get_json(silent=True) or {}
    email = str(data.get("email", "")).strip()
    password = str(data.get("password", ""))
    if "@" not in email:
        return jsonify({"error": "invalid_email"}), 400
    if len(password) < 6:
        return jsonify({"error": "invalid_password"}), 400
    if User.query.filter((User.username == email) | (User.email == email)).first():
        return jsonify({"error": "already_exists"}), 409
    user = User(username=email, email=email, password_hash=generate_password_hash(password))
    db.session.add(user)
    db.session.commit()
    get_or_create_profile(user.id)
    token = token_serializer.dumps({"uid": user.id, "nonce": secrets.token_hex(8)})
    return jsonify({"token": token, "user": {"id": user.id, "email": user.email}}), 201


@app.post("/api/change-password")
@api_auth_required
def api_change_password():
    user = request.api_user
    data = request.get_json(silent=True) or {}
    old_password = str(data.get("oldPassword", ""))
    new_password = str(data.get("newPassword", ""))
    if not check_password_hash(user.password_hash, old_password):
        return jsonify({"error": "invalid_old_password"}), 400
    if len(new_password) < 6:
        return jsonify({"error": "invalid_new_password"}), 400
    user.password_hash = generate_password_hash(new_password)
    db.session.commit()
    return jsonify({"ok": True})


@app.get("/api/profile")
@api_auth_required
def api_profile_get():
    user = request.api_user
    p = get_or_create_profile(user.id)
    return jsonify({
        "daireBaskanligi": p.daire_baskanligi or "",
        "subeMudurlugu": p.sube_mudurlugu or "",
        "adSoyad": p.ad_soyad or "",
        "sicilNo": p.sicil_no or "",
        "ekipKodu": p.ekip_kodu or "",
    })


@app.get("/api/period-lock-status")
@api_auth_required
def api_period_lock_status():
    user = request.api_user
    work_date_raw = (request.args.get("workDate") or "").strip()
    if not work_date_raw:
        return jsonify({"error": "missing_work_date"}), 400
    try:
        wd = parse_date(work_date_raw)
    except Exception:
        return jsonify({"error": "invalid_work_date"}), 400
    return jsonify({
        "workDate": wd.isoformat(),
        "locked": bool(is_period_locked(wd)),
        "canBypass": bool(can_bypass_period_lock(user)),
    })


@app.put("/api/profile")
@api_auth_required
def api_profile_put():
    user = request.api_user
    p = get_or_create_profile(user.id)
    data = request.get_json(silent=True) or {}
    p.daire_baskanligi = str(data.get("daireBaskanligi", p.daire_baskanligi or "")).strip()
    p.sube_mudurlugu = str(data.get("subeMudurlugu", p.sube_mudurlugu or "")).strip()
    p.ad_soyad = str(data.get("adSoyad", p.ad_soyad or "")).strip()
    p.sicil_no = str(data.get("sicilNo", p.sicil_no or "")).strip()
    p.ekip_kodu = str(data.get("ekipKodu", p.ekip_kodu or "")).strip()
    db.session.commit()
    return jsonify({"ok": True})


@app.get("/api/entries")
@api_auth_required
def api_entries():
    user = request.api_user
    updated_after = request.args.get("updatedAfter")
    q = OvertimeEntry.query.filter_by(user_id=user.id)
    if updated_after:
        try:
            dt = datetime.fromisoformat(updated_after)
            q = q.filter(OvertimeEntry.updated_at > dt)
        except Exception:
            return jsonify({"error": "invalid_updatedAfter"}), 400
    entries = q.order_by(OvertimeEntry.updated_at.asc()).all()
    return jsonify([entry_to_dict(e) for e in entries])


@app.post("/api/entries")
@api_auth_required
def api_create_entry():
    user = request.api_user
    data = request.get_json(silent=True) or {}
    try:
        work_date = parse_date(str(data.get("workDate", "")))
        if is_period_locked(work_date) and not can_bypass_period_lock(user):
            return jsonify({"error": "period_locked"}), 423
        start_time = str(data.get("startTime", ""))
        end_time = str(data.get("endTime", ""))

        # APK/web tekrar aynı kaydı gönderirse mükerrer oluşturma
        existing = OvertimeEntry.query.filter_by(
            user_id=user.id,
            work_date=work_date,
            start_time=start_time,
            end_time=end_time,
        ).first()
        if existing:
            return jsonify(entry_to_dict(existing)), 200

        if find_overlapping_overtime_for_user(user.id, work_date, start_time, end_time):
            return jsonify({"error": "time_overlap", "message": MESAI_TIME_OVERLAP_MESSAGE}), 409

        entry = OvertimeEntry(
            user_id=user.id,
            work_date=work_date,
            start_time=start_time,
            end_time=end_time,
            pct60=float(data.get("pct60", 0)),
            pct15=float(data.get("pct15", 0)),
            pazar=float(data.get("pazar", 0)),
            bayram=float(data.get("bayram", 0)),
            description=str(data.get("description", "")),
        )
        db.session.add(entry)
        write_overtime_audit_log(
            action="create",
            actor_user_id=user.id,
            target_user_id=user.id,
            old_entry=None,
            new_entry=entry,
            source="apk",
            note="api_create",
        )
        db.session.commit()
        return jsonify(entry_to_dict(entry)), 201
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": str(exc)}), 400


@app.put("/api/entries/<int:entry_id>")
@api_auth_required
def api_update_entry(entry_id: int):
    user = request.api_user
    data = request.get_json(silent=True) or {}
    entry = OvertimeEntry.query.filter_by(id=entry_id, user_id=user.id).first()
    if not entry:
        return jsonify({"error": "not_found"}), 404
    try:
        old_snapshot = OvertimeEntry(
            id=entry.id,
            user_id=entry.user_id,
            work_date=entry.work_date,
            start_time=entry.start_time,
            end_time=entry.end_time,
            pct60=entry.pct60,
            pct15=entry.pct15,
            pazar=entry.pazar,
            bayram=entry.bayram,
            description=entry.description,
        )
        new_work_date = parse_date(str(data.get("workDate", entry.work_date.isoformat())))
        if (is_period_locked(entry.work_date) or is_period_locked(new_work_date)) and not can_bypass_period_lock(user):
            return jsonify({"error": "period_locked"}), 423
        new_start = str(data.get("startTime", entry.start_time))
        new_end = str(data.get("endTime", entry.end_time))
        if find_overlapping_overtime_for_user(
            user.id, new_work_date, new_start, new_end, exclude_entry_id=entry.id
        ):
            return jsonify({"error": "time_overlap", "message": MESAI_TIME_OVERLAP_MESSAGE}), 409
        entry.work_date = new_work_date
        entry.start_time = new_start
        entry.end_time = new_end
        entry.pct60 = float(data.get("pct60", entry.pct60))
        entry.pct15 = float(data.get("pct15", entry.pct15))
        entry.pazar = float(data.get("pazar", entry.pazar))
        entry.bayram = float(data.get("bayram", entry.bayram))
        entry.description = str(data.get("description", entry.description))
        write_overtime_audit_log(
            action="update",
            actor_user_id=user.id,
            target_user_id=user.id,
            old_entry=old_snapshot,
            new_entry=entry,
            source="apk",
            note="api_update",
        )
        db.session.commit()
        return jsonify(entry_to_dict(entry))
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": str(exc)}), 400


@app.delete("/api/entries/<int:entry_id>")
@api_auth_required
def api_delete_entry(entry_id: int):
    user = request.api_user
    entry = OvertimeEntry.query.filter_by(id=entry_id, user_id=user.id).first()
    if not entry:
        return jsonify({"error": "not_found"}), 404
    if is_period_locked(entry.work_date) and not can_bypass_period_lock(user):
        return jsonify({"error": "period_locked"}), 423
    write_overtime_audit_log(
        action="delete",
        actor_user_id=user.id,
        target_user_id=user.id,
        old_entry=entry,
        new_entry=None,
        source="apk",
        note="api_delete",
    )
    db.session.delete(entry)
    db.session.commit()
    return jsonify({"ok": True})


@app.cli.command("init-db")
def init_db():
    db.create_all()
    print("Database initialized.")


def sync_usernames_with_emails() -> int:
    users = User.query.all()
    changed = 0
    for u in users:
        email = (u.email or "").strip()
        if not email:
            continue
        if (u.username or "").strip() != email:
            u.username = email
            changed += 1
    if changed:
        db.session.commit()
    return changed


def ensure_delegated_permission_columns():
    inspector = db.inspect(db.engine)
    try:
        cols = {c["name"] for c in inspector.get_columns("delegated_admin_permission")}
    except Exception:
        return
    if "can_reset_password" not in cols:
        db.session.execute(db.text("ALTER TABLE delegated_admin_permission ADD COLUMN can_reset_password BOOLEAN NOT NULL DEFAULT FALSE"))
    if "can_view_users_screen" not in cols:
        db.session.execute(db.text("ALTER TABLE delegated_admin_permission ADD COLUMN can_view_users_screen BOOLEAN NOT NULL DEFAULT FALSE"))
    if "can_view_charts" not in cols:
        db.session.execute(db.text("ALTER TABLE delegated_admin_permission ADD COLUMN can_view_charts BOOLEAN NOT NULL DEFAULT FALSE"))
    if "can_view_filters" not in cols:
        db.session.execute(db.text("ALTER TABLE delegated_admin_permission ADD COLUMN can_view_filters BOOLEAN NOT NULL DEFAULT FALSE"))
    if "can_add_user" not in cols:
        db.session.execute(db.text("ALTER TABLE delegated_admin_permission ADD COLUMN can_add_user BOOLEAN NOT NULL DEFAULT FALSE"))
    if "can_change_email" not in cols:
        db.session.execute(db.text("ALTER TABLE delegated_admin_permission ADD COLUMN can_change_email BOOLEAN NOT NULL DEFAULT FALSE"))
    if "can_period_lock" not in cols:
        db.session.execute(db.text("ALTER TABLE delegated_admin_permission ADD COLUMN can_period_lock BOOLEAN NOT NULL DEFAULT FALSE"))
    if "can_bulk_entry" not in cols:
        db.session.execute(db.text("ALTER TABLE delegated_admin_permission ADD COLUMN can_bulk_entry BOOLEAN NOT NULL DEFAULT FALSE"))
    if "can_view_terminated_users" not in cols:
        db.session.execute(db.text("ALTER TABLE delegated_admin_permission ADD COLUMN can_view_terminated_users BOOLEAN NOT NULL DEFAULT FALSE"))
    if "can_unit_change" not in cols:
        db.session.execute(db.text("ALTER TABLE delegated_admin_permission ADD COLUMN can_unit_change BOOLEAN NOT NULL DEFAULT FALSE"))
    if "scope_daire_baskanligi" not in cols:
        db.session.execute(db.text("ALTER TABLE delegated_admin_permission ADD COLUMN scope_daire_baskanligi VARCHAR(255) NOT NULL DEFAULT ''"))
    if "scope_sube_mudurlugu" not in cols:
        db.session.execute(db.text("ALTER TABLE delegated_admin_permission ADD COLUMN scope_sube_mudurlugu VARCHAR(255) NOT NULL DEFAULT ''"))
    # eski kolon varsa yeni yapıya taşımak için bir kez eşitle
    if "can_view_passwords" in cols:
        db.session.execute(
            db.text(
                "UPDATE delegated_admin_permission "
                "SET can_reset_password = CASE WHEN can_view_passwords IS NULL THEN can_reset_password ELSE can_view_passwords END "
                "WHERE can_reset_password = FALSE"
            )
        )
        db.session.execute(
            db.text(
                "UPDATE delegated_admin_permission "
                "SET can_view_passwords = FALSE "
                "WHERE can_view_passwords IS NULL"
            )
        )
    db.session.commit()


def ensure_user_profile_columns():
    inspector = db.inspect(db.engine)
    try:
        cols = {c["name"] for c in inspector.get_columns("user_profile")}
    except Exception:
        return
    if "employment_end_date" not in cols:
        db.session.execute(db.text("ALTER TABLE user_profile ADD COLUMN employment_end_date DATE"))
    db.session.commit()


def ensure_performance_indexes():
    try:
        db.session.execute(db.text("CREATE INDEX IF NOT EXISTS idx_overtime_user_date ON overtime_entry (user_id, work_date)"))
        db.session.execute(db.text("CREATE INDEX IF NOT EXISTS idx_unit_change_user_date ON unit_change (user_id, transfer_date)"))
        db.session.execute(db.text("CREATE INDEX IF NOT EXISTS idx_audit_time ON audit_log (event_time)"))
        db.session.execute(db.text("CREATE INDEX IF NOT EXISTS idx_audit_actor_time ON audit_log (actor_user_id, event_time)"))
        db.session.execute(db.text("CREATE INDEX IF NOT EXISTS idx_audit_target_time ON audit_log (target_user_id, event_time)"))
        db.session.execute(db.text("CREATE INDEX IF NOT EXISTS idx_audit_period ON audit_log (period_start_year, period_start_month)"))
        db.session.commit()
    except Exception:
        db.session.rollback()


@app.cli.command("sync-usernames")
def sync_usernames_cmd():
    changed = sync_usernames_with_emails()
    print(f"Synced users: {changed}")


with app.app_context():
    db.create_all()
    try:
        ensure_user_profile_columns()
        ensure_delegated_permission_columns()
        ensure_performance_indexes()
        sync_usernames_with_emails()
    except Exception:
        db.session.rollback()


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
